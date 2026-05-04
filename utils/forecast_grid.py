"""SEO Channel Forecast grid builder (GAZMAN-style).

Generates a single-sheet xlsx "SEO Channel Forecast" with:
  - Rows 1-11: header/preamble block (title, last updated, fee rows)
  - Row 12: column sub-header strip (Forecast | Actuals | % Change per month)
  - Rows 13-19: 7-metric scaffold (BUDGET, REVENUE, ROAS, TRANSACTIONS, AOV, TRAFFIC, CVR)

Also provides `build_three_scenario_grid` — a four-sheet xlsx for budget-tier
presentations: Conservative / Moderate / Aggressive / Comparison.

Pure Python + openpyxl.  No Streamlit imports.
"""

from __future__ import annotations

import calendar
import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from engine.new_content_engine import get_ctr
from exporters.charts import add_charts_sheet

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_BOLD_FONT = Font(bold=True, size=11)
_RED_FONT = Font(color="FF0000")
_NUMBER_FMT = "#,##0"
_CURRENCY_FMT = "$#,##0.00"
_PERCENT_FMT = "0.00%"
_PCT_CHANGE_FMT = "0%"
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# GAZMAN column layout constants and helpers
# ---------------------------------------------------------------------------

_DATA_START = 3  # Col C — A=CHANNEL label, B=metric label, C+ = monthly data


def _fc(i: int) -> int:
    """Forecast column for month i (0-indexed)."""
    return _DATA_START + i * 3


def _ac(i: int) -> int:
    """Actuals column for month i."""
    return _DATA_START + i * 3 + 1


def _pc(i: int) -> int:
    """% Change column for month i."""
    return _DATA_START + i * 3 + 2


def _col_annual(m: int) -> int:
    return _DATA_START + m * 3


def _col_ytd(m: int) -> int:
    return _col_annual(m) + 1


def _col_ass(m: int) -> int:
    return _col_annual(m) + 2


def _col_prior(m: int) -> int:
    return _col_annual(m) + 3


def _col_yoy(m: int) -> int:
    return _col_annual(m) + 4


# ---------------------------------------------------------------------------
# Public column-range helper (reused by data-population sessions)
# ---------------------------------------------------------------------------


def _month_column_ranges(start_month: int, months: int) -> list[tuple[str, int, int, int]]:
    """Return [(month_name, forecast_col, actuals_col, pct_change_col), ...].

    1-based column indices.  First month's forecast col is 3 (A=channel, B=metric).
    """
    result = []
    for i in range(months):
        month_name = calendar.month_abbr[(start_month - 1 + i) % 12 + 1]
        result.append((month_name, _fc(i), _ac(i), _pc(i)))
    return result


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)


def _hcell(ws, row: int, col: int, value: str) -> None:
    """Write a styled header cell."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _HEADER_FONT
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center")


def _pct_change(forecast: float | None, actual: float | None) -> float | None:
    """Return (actual - forecast) / forecast, with edge-case handling.

    Rules:
      - None when either input is None.
      - None when forecast == 0 (undefined, including both-zero case).
      - -1.0 when actual == 0 and forecast != 0 (renders as -100%).
    """
    if actual is None or forecast is None:
        return None
    if forecast == 0:
        return None
    if actual == 0:
        return -1.0
    return (actual - forecast) / forecast


def _write_metric_rows(
    ws,
    months: int,
    col_ranges: list,
    monthly_traffic: list[float],
    monthly_transactions: list[float],
    monthly_revenue: list[float],
    monthly_cvr: list[float],
    monthly_aov: list[float],
    monthly_budget: list[float],
    actuals_traffic: list[float] | None,
    actuals_transactions: list[float] | None,
    actuals_revenue: list[float] | None,
    actuals_cvr: list[float] | None,
    actuals_aov: list[float] | None,
    actuals_budget: list[float] | None,
    prior_year_traffic: float | None,
    prior_year_transactions: float | None,
    prior_year_revenue: float | None,
    prior_year_cvr: float | None,
    prior_year_aov: float | None,
    prior_year_budget: float | None,
) -> None:
    """Fill rows 13-19 with forecast, actuals, % Change, and trailing totals."""
    ann_col = _col_annual(months)

    def _v(row: int, col: int, val, fmt: str, *, neg_red: bool = False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = fmt
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="right")
        if neg_red and val is not None and val < 0:
            cell.font = _RED_FONT
        return cell

    def _get(lst: list | None, i: int):
        return lst[i] if lst and i < len(lst) else None

    # Pre-compute per-month ROAS (forecast and actual)
    roas_fc = [
        monthly_revenue[i] / monthly_budget[i]
        if i < len(monthly_budget) and monthly_budget[i] != 0 else None
        for i in range(months)
    ]
    roas_ac: list[float | None] = []
    for i in range(months):
        rev_a = _get(actuals_revenue, i)
        bud_a = _get(actuals_budget, i)
        roas_ac.append(
            rev_a / bud_a if rev_a is not None and bud_a is not None and bud_a != 0
            else None
        )

    # (excel_row, fc_list, ac_list, prior_raw, fmt, is_cvr, agg)
    _METRICS = [
        (13, monthly_budget,       actuals_budget,       prior_year_budget,       "$#,##0",    False, "sum"),
        (14, monthly_revenue,      actuals_revenue,      prior_year_revenue,      "$#,##0",    False, "sum"),
        (15, roas_fc,              roas_ac,              None,                    "$#,##0.00", False, "roas"),
        (16, monthly_transactions, actuals_transactions, prior_year_transactions, "#,##0",     False, "sum"),
        (17, monthly_aov,          actuals_aov,          prior_year_aov,          "$#,##0.00", False, "avg"),
        (18, monthly_traffic,      actuals_traffic,      prior_year_traffic,      "#,##0",     False, "sum"),
        (19, monthly_cvr,          actuals_cvr,          prior_year_cvr,          "0.00%",     True,  "avg"),
    ]

    for row, fc_raw, ac_raw, prior_raw, fmt, is_cvr, agg in _METRICS:
        fc_cells: list[float | None] = []
        ac_cells: list[float | None] = []

        for i, (_, fc_col, ac_col, pc_col) in enumerate(col_ranges):
            fv_raw = _get(fc_raw, i)
            av_raw = _get(ac_raw, i)

            fv = fv_raw / 100.0 if is_cvr and fv_raw is not None else fv_raw
            av = av_raw / 100.0 if is_cvr and av_raw is not None else av_raw

            fc_cells.append(fv)
            ac_cells.append(av)

            if fv is not None:
                _v(row, fc_col, fv, fmt)
            if av is not None:
                _v(row, ac_col, av, fmt)

            pct = _pct_change(fv, av)
            if pct is not None:
                _v(row, pc_col, pct, _PCT_CHANGE_FMT, neg_red=True)

        # Annual Forecast
        if agg == "roas":
            total_rev = sum(monthly_revenue[:months])
            total_bud = sum(monthly_budget[:months])
            ann_fc = total_rev / total_bud if total_bud != 0 else None
        elif agg == "sum":
            nonnull = [v for v in fc_cells if v is not None]
            ann_fc = sum(nonnull) if nonnull else None
        else:  # avg (AOV, CVR)
            nonnull = [v for v in fc_cells if v is not None]
            ann_fc = sum(nonnull) / len(nonnull) if nonnull else None

        if ann_fc is not None:
            _v(row, ann_col, ann_fc, fmt)

        # YTD Actuals
        if agg == "roas":
            rev_vals = [_get(actuals_revenue, i) for i in range(months)]
            bud_vals = [_get(actuals_budget, i) for i in range(months)]
            rev_sum = sum(v for v in rev_vals if v is not None)
            bud_sum = sum(v for v in bud_vals if v is not None)
            ytd: float | None = rev_sum / bud_sum if bud_sum != 0 else None
        elif agg == "sum":
            nonnull_ac = [v for v in ac_cells if v is not None]
            ytd = sum(nonnull_ac) if nonnull_ac else None
        else:
            nonnull_ac = [v for v in ac_cells if v is not None]
            ytd = sum(nonnull_ac) / len(nonnull_ac) if nonnull_ac else None

        if ytd is not None:
            _v(row, _col_ytd(months), ytd, fmt)

        # Prior Year
        prior_cell = prior_raw / 100.0 if is_cvr and prior_raw is not None else prior_raw
        if prior_cell is not None:
            _v(row, _col_prior(months), prior_cell, fmt)

        # YoY %
        if ann_fc is not None and prior_cell is not None and prior_cell != 0:
            yoy = (ann_fc - prior_cell) / prior_cell
            _v(row, _col_yoy(months), yoy, _PCT_CHANGE_FMT, neg_red=True)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def build_seo_forecast_grid(
    monthly_traffic: list[float],
    monthly_transactions: list[float],
    monthly_revenue: list[float],
    monthly_cvr: list[float],
    monthly_aov: list[float],
    monthly_budget: list[float],
    actuals_traffic: list[float] | None = None,
    actuals_transactions: list[float] | None = None,
    actuals_revenue: list[float] | None = None,
    actuals_cvr: list[float] | None = None,
    actuals_aov: list[float] | None = None,
    actuals_budget: list[float] | None = None,
    prior_year_traffic: float | None = None,
    prior_year_transactions: float | None = None,
    prior_year_revenue: float | None = None,
    prior_year_cvr: float | None = None,
    prior_year_aov: float | None = None,
    prior_year_budget: float | None = None,
    months: int = 12,
    client_name: str = "",
    fy_label: str = "FY26",
    start_month: int = 1,
    last_updated: str | None = None,
    currency_notes: str = "",
    assumptions_text: str = "",
) -> io.BytesIO:
    """Build a GAZMAN-style SEO Channel Forecast xlsx.

    Returns a BytesIO buffer containing a single-sheet workbook with rows 13-19
    populated: per-month Forecast / Actuals / % Change, plus Annual Forecast,
    YTD Actuals, Prior Year, and YoY % trailing columns.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Channel Forecast"

    col_ranges = _month_column_ranges(start_month, months)
    ann_col = _col_annual(months)

    # Derive FY year for trailing column headers (FY26 -> 2026)
    try:
        fy_num = int(fy_label.replace("FY", "").strip())
        fy_year = 2000 + fy_num if fy_num < 100 else fy_num
        prior_year_label = fy_year - 1
    except ValueError:
        fy_year, prior_year_label = 2026, 2025

    # Row 2: Title
    title = (
        f"{client_name} | FORECAST | {fy_label}"
        if client_name
        else f"FORECAST | {fy_label}"
    )
    ws.cell(row=2, column=1, value=title).font = Font(bold=True, size=14)

    # Row 4: Last Updated
    if last_updated:
        ws.cell(row=4, column=2, value="Last Updated:")
        ws.cell(row=4, column=3, value=last_updated)

    # Row 5: Currency notes
    if currency_notes:
        ws.cell(row=5, column=2, value=currency_notes)

    # Row 7: Month name headers (each spanning 3 columns via merge)
    ws.cell(row=7, column=1, value="MONTH").font = _BOLD_FONT
    for m_name, fc, _ac_col, _pc_col in col_ranges:
        cell = ws.cell(row=7, column=fc, value=m_name)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=7, start_column=fc, end_row=7, end_column=fc + 2)

    # "TOTALS" merged header spanning all trailing columns
    cell = ws.cell(row=7, column=ann_col, value="TOTALS")
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = _HEADER_FILL
    cell.alignment = Alignment(horizontal="center")
    ws.merge_cells(
        start_row=7, start_column=ann_col,
        end_row=7, end_column=_col_yoy(months),
    )

    # Row 8: Total Forecasted Management Fee
    ws.cell(row=8, column=1, value="Total Forecasted Management Fee:")
    for i, (_, fc, _, _) in enumerate(col_ranges):
        val = monthly_budget[i] if i < len(monthly_budget) else None
        if val is not None:
            ws.cell(row=8, column=fc, value=val).number_format = _CURRENCY_FMT
    if monthly_budget:
        ws.cell(
            row=8, column=ann_col, value=sum(monthly_budget[:months])
        ).number_format = _CURRENCY_FMT

    # Row 9: Total Actual Management Fee
    ws.cell(row=9, column=1, value="Total Actual Management Fee:")
    if actuals_budget:
        for i, (_, fc, _, _) in enumerate(col_ranges):
            val = actuals_budget[i] if i < len(actuals_budget) else None
            if val is not None:
                ws.cell(row=9, column=fc, value=val).number_format = _CURRENCY_FMT

    # Row 10: Management Fee + Ad Spend (SEO-only = same as forecasted budget)
    ws.cell(row=10, column=1, value="Management Fee + Ad Spend")
    for i, (_, fc, _, _) in enumerate(col_ranges):
        val = monthly_budget[i] if i < len(monthly_budget) else None
        if val is not None:
            ws.cell(row=10, column=fc, value=val).number_format = _CURRENCY_FMT
    if monthly_budget:
        ws.cell(
            row=10, column=ann_col, value=sum(monthly_budget[:months])
        ).number_format = _CURRENCY_FMT

    # Row 12: Column sub-header strip
    _hcell(ws, 12, 1, "CHANNEL")
    for _, fc, ac_col, pc_col in col_ranges:
        _hcell(ws, 12, fc, "Forecast")
        _hcell(ws, 12, ac_col, "Actuals")
        _hcell(ws, 12, pc_col, "% Change")

    _hcell(ws, 12, ann_col, f"{fy_year} Forecast")
    _hcell(ws, 12, _col_ytd(months), f"{fy_year} YTD Actuals")
    _hcell(ws, 12, _col_ass(months), "SEO Assumptions:")
    _hcell(ws, 12, _col_prior(months), f"{prior_year_label} Actuals")
    _hcell(ws, 12, _col_yoy(months), "YoY %")

    # Rows 13-19: metric labels + data
    _METRIC_LABELS = ["BUDGET", "REVENUE", "ROAS", "TRANSACTIONS", "AOV", "TRAFFIC", "CVR"]
    for r_off, label in enumerate(_METRIC_LABELS):
        row = 13 + r_off
        if r_off == 0:
            ws.cell(row=row, column=1, value="SEO").font = _BOLD_FONT
        ws.cell(row=row, column=2, value=label).font = _BOLD_FONT

    _write_metric_rows(
        ws, months, col_ranges,
        monthly_traffic, monthly_transactions, monthly_revenue,
        monthly_cvr, monthly_aov, monthly_budget,
        actuals_traffic, actuals_transactions, actuals_revenue,
        actuals_cvr, actuals_aov, actuals_budget,
        prior_year_traffic, prior_year_transactions, prior_year_revenue,
        prior_year_cvr, prior_year_aov, prior_year_budget,
    )

    # Assumptions text column — merged vertically across rows 13-19
    ass_col = _col_ass(months)
    ass_cell = ws.cell(row=13, column=ass_col, value=assumptions_text)
    ws.merge_cells(
        start_row=13, start_column=ass_col,
        end_row=19, end_column=ass_col,
    )
    ass_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Fee rows 21-22
    for fee_row, fee_label in (
        (21, "SEO Management + Tech Fee"),
        (22, "SEO Total"),
    ):
        ws.cell(row=fee_row, column=1, value=fee_label).font = _BOLD_FONT
        for i, (_, fc, ac_col, _) in enumerate(col_ranges):
            fv = monthly_budget[i] if i < len(monthly_budget) else None
            if fv is not None:
                c = ws.cell(row=fee_row, column=fc, value=fv)
                c.number_format = _CURRENCY_FMT
                c.border = _THIN_BORDER
            if actuals_budget and i < len(actuals_budget):
                av = actuals_budget[i]
                if av is not None:
                    ca = ws.cell(row=fee_row, column=ac_col, value=av)
                    ca.number_format = _CURRENCY_FMT
                    ca.border = _THIN_BORDER
        if monthly_budget:
            ann_c = ws.cell(
                row=fee_row, column=ann_col,
                value=sum(monthly_budget[:months]),
            )
            ann_c.number_format = _CURRENCY_FMT
            ann_c.border = _THIN_BORDER

    ws.freeze_panes = "B13"
    _auto_width(ws)
    ws.column_dimensions[get_column_letter(ass_col)].width = 30

    # ── Charts sheet ─────────────────────────────────────────────────────────
    month_names = [name for name, _, _, _ in col_ranges]
    _chart_df = pd.DataFrame({
        "month_name": month_names,
        "baseline_traffic": [0] * months,
        "combined_p50": monthly_traffic[:months],
    })
    add_charts_sheet(wb, {"SEO Forecast": _chart_df}, month_names)
    if "Charts" in wb.sheetnames:
        wb["Charts"].sheet_properties.tabColor = "0F172A"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Three-scenario grid helpers
# ---------------------------------------------------------------------------

_TAB_COLORS: dict[str, str] = {
    "Conservative": "94A3B8",
    "Moderate": "2563EB",
    "Aggressive": "22C55E",
    "Comparison": "0F172A",
}

_SCENARIO_HEADER_FILLS: dict[str, PatternFill] = {
    name: PatternFill(start_color=color, end_color=color, fill_type="solid")
    for name, color in _TAB_COLORS.items()
}

_WHITE_FONT = Font(bold=True, color="FFFFFF", size=11)
_DARK_FONT = Font(bold=True, color="FFFFFF", size=11)

_COL_FORMATS: dict[str, str] = {
    "Baseline Traffic": "#,##0",
    "Positional Uplift": "#,##0",
    "New Content Uplift": "#,##0",
    "Decay": "#,##0",
    "Traffic P10": "#,##0",
    "Traffic P50": "#,##0",
    "Traffic P90": "#,##0",
    "Transactions": "#,##0",
    "Revenue": "$#,##0.00",
    "AOV Used": "$#,##0.00",
    "CVR Used (%)": "0.00%",
    "Avg Portfolio Position": "0.0",
    "Avg Portfolio CTR (%)": "0.00%",
    "Seasonality Modifier": "0.00%",
}

# Columns whose values are stored divided by 100 for Excel percentage display
_PCT_DIVIDE_COLS = {"CVR Used (%)", "Avg Portfolio CTR (%)"}


def _preset_subtitle(preset: dict) -> str:
    effort = preset.get("effort_level", "—").title()
    cadence = preset.get("content_cadence", 0)
    maint = int(preset.get("maintenance_coverage", 0.0) * 100)
    hours = preset.get("total_monthly_hours", 0.0)
    retainer = preset.get("retainer_aud_monthly", 0.0)
    return (
        f"{effort} effort | {cadence} posts/month | "
        f"{maint}% maintenance | {hours} hrs/month | "
        f"${retainer:,.0f}/month retainer"
    )


def _write_scenario_ws(
    ws,
    df: pd.DataFrame,
    scenario_name: str,
    preset: dict,
    fy_label: str,
) -> None:
    """Write a scenario DataFrame to an openpyxl worksheet (rows 1-5+)."""
    fill = _SCENARIO_HEADER_FILLS[scenario_name]

    # Row 1: Title
    title_cell = ws.cell(row=1, column=1, value=f"{scenario_name} Scenario — {fy_label} SEO Forecast")
    title_cell.font = Font(bold=True, size=14)

    # Row 2: Subtitle (preset summary)
    sub_cell = ws.cell(row=2, column=1, value=_preset_subtitle(preset))
    sub_cell.font = Font(italic=True, size=10)

    # Row 3: Blank — leave empty

    # Row 4: Column headers
    columns = list(df.columns)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font = _WHITE_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    # Freeze panes below header
    ws.freeze_panes = "A5"

    # Rows 5+: Data
    for row_idx, (_, data_row) in enumerate(df.iterrows(), start=5):
        for col_idx, col_name in enumerate(columns, start=1):
            raw_val = data_row[col_name]
            # Convert NaN / None to None for Excel
            if raw_val is None or (isinstance(raw_val, float) and pd.isna(raw_val)):
                val = None
            elif col_name in _PCT_DIVIDE_COLS and raw_val is not None:
                val = float(raw_val) / 100.0
            else:
                val = raw_val

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            fmt = _COL_FORMATS.get(col_name)
            if fmt and val is not None:
                cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")

    _auto_width(ws)


def build_scenario_grid_sheet(
    scenario_name: str,
    scenario_result: dict,
    preset: dict,
    cvr_series: list[float] | float,
    aov_series: list[float] | float,
    currency: str = "AUD",
    start_month: int = 7,
    include_seasonality_column: bool = True,
    seasonality: dict | None = None,
) -> pd.DataFrame:
    """Build the full monthly DataFrame for one scenario's sheet.

    Only forecast rows (is_forecast=True) from combined_df are included.
    Month labels and calendar months are derived from the date column in
    combined_df; the caller is responsible for ensuring the forecast starts
    at start_month if label alignment matters.

    Avg Portfolio Position is a static traffic-weighted mean of target_position
    across all ranked keywords — the same value appears in every forecast month.
    This is an approximation; the true per-month position ramps via the S-curve
    maturation schedule but is too expensive to compute per-row in a grid export.

    Args:
        scenario_name: "Conservative" / "Moderate" / "Aggressive"
        scenario_result: Dict from run_three_scenarios (has combined_df, positional_keyword_df, etc.)
        preset: The scenario's preset dict (effort, cadence, etc.)
        cvr_series: Per-month CVR as %, or single scalar.
        aov_series: Per-month AOV, or single scalar.
        currency: ISO code (currently informational only).
        start_month: Calendar month (1–12) of horizon month 1 (for label generation).
        include_seasonality_column: Add a "Seasonality Modifier" column (traffic_mod).
        seasonality: Monthly seasonality dict keyed by calendar month integer.

    Returns:
        DataFrame with one row per forecast month.
    """
    combined_df = scenario_result.get("combined_df", pd.DataFrame())
    if combined_df.empty or "is_forecast" not in combined_df.columns:
        return pd.DataFrame()

    forecast_df = combined_df[combined_df["is_forecast"]].reset_index(drop=True)
    n_months = len(forecast_df)
    if n_months == 0:
        return pd.DataFrame()

    # Normalise CVR / AOV to per-month lists
    if isinstance(cvr_series, (int, float)):
        cvr_list = [float(cvr_series)] * n_months
    else:
        cvr_list = list(cvr_series)[:n_months]
        if len(cvr_list) < n_months:
            cvr_list += [cvr_list[-1]] * (n_months - len(cvr_list))

    if isinstance(aov_series, (int, float)):
        aov_list = [float(aov_series)] * n_months
    else:
        aov_list = list(aov_series)[:n_months]
        if len(aov_list) < n_months:
            aov_list += [aov_list[-1]] * (n_months - len(aov_list))

    # Compute static avg portfolio position (traffic-weighted target position)
    kw = scenario_result.get("positional_keyword_df", pd.DataFrame())
    avg_position: float | None = None
    if (
        not kw.empty
        and "target_position" in kw.columns
        and "volume" in kw.columns
    ):
        weights = kw["volume"] * kw["target_position"].apply(lambda p: get_ctr(int(p)))
        total_weight = weights.sum()
        if total_weight > 0:
            avg_position = float((kw["target_position"] * weights).sum() / total_weight)

    avg_ctr: float | None = None
    if avg_position is not None:
        avg_ctr = get_ctr(round(avg_position))

    rows = []
    for i, (_, row) in enumerate(forecast_df.iterrows()):
        date = pd.Timestamp(row["date"])
        calendar_month = date.month
        month_label = date.strftime("%b %Y")

        baseline = int(row.get("baseline", 0) or 0)
        pos_uplift = int(row.get("positional_uplift_p50", row.get("positional_uplift", 0)) or 0)
        nc_uplift = int(row.get("new_content_uplift", 0) or 0)
        decay_raw = int(row.get("decay", 0) or 0)
        decay_display = -abs(decay_raw)

        p10 = int(row.get("combined_p10", row.get("combined", 0)) or 0)
        p50 = int(row.get("combined_p50", row.get("combined", 0)) or 0)
        p90 = int(row.get("combined_p90", row.get("combined", 0)) or 0)

        cvr = cvr_list[i]
        aov = aov_list[i]
        transactions = p50 * cvr / 100.0
        revenue = transactions * aov

        r: dict = {
            "Month Label": month_label,
            "Calendar Month": calendar_month,
            "Is Forecast": True,
            "Baseline Traffic": baseline,
            "Positional Uplift": pos_uplift,
            "New Content Uplift": nc_uplift,
            "Decay": decay_display,
            "Traffic P10": p10,
            "Traffic P50": p50,
            "Traffic P90": p90,
            "Transactions": round(transactions, 1),
            "Revenue": round(revenue, 2),
            "AOV Used": float(aov),
            "CVR Used (%)": float(cvr),
            "Avg Portfolio Position": avg_position,
            "Avg Portfolio CTR (%)": avg_ctr,
        }

        if include_seasonality_column:
            modifier = 0.0
            if seasonality:
                modifier = float(seasonality.get(calendar_month, {}).get("traffic_mod", 0.0))
            r["Seasonality Modifier"] = modifier

        rows.append(r)

    return pd.DataFrame(rows)


def build_three_scenario_grid(
    scenario_results: dict[str, dict],
    presets: dict[str, dict],
    cvr: float,
    aov: float,
    seasonality: dict | None = None,
    apply_seasonal_aov: bool = True,
    currency: str = "AUD",
    start_month: int = 7,
    client_name: str = "",
    fy_label: str = "FY26",
) -> io.BytesIO:
    """Build a four-sheet xlsx: Conservative, Moderate, Aggressive, Comparison.

    Each scenario sheet is produced by build_scenario_grid_sheet using monthly
    CVR and AOV series derived from start_month + seasonality offsets.

    Seasonal AOV: when apply_seasonal_aov=True and seasonality has 'aov_mod'
    values, per-month AOV = base_aov × (1 + aov_mod). Similarly for CVR via
    'cr_mod'. Months are indexed from start_month (e.g. start_month=7 → Jul,
    Aug, Sep, …) regardless of actual calendar dates in combined_df.

    Args:
        scenario_results: Output of run_three_scenarios().
        presets: Output of build_scenario_presets().
        cvr: Base conversion rate as % (e.g. 2.5 for 2.5%).
        aov: Base average order value.
        seasonality: Monthly seasonality dict keyed by int 1–12. Each entry
                     may have 'traffic_mod', 'cr_mod', 'aov_mod' (all floats).
        apply_seasonal_aov: Apply aov_mod from seasonality to AOV when True.
        currency: ISO currency code (informational).
        start_month: Calendar month (1–12) of forecast month 1.
        client_name: Optional client name for sheet titles.
        fy_label: Financial year label (e.g. "FY26").

    Returns:
        BytesIO buffer containing the xlsx workbook.
    """
    wb = Workbook()
    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    scenario_dfs: dict[str, pd.DataFrame] = {}

    for scenario_name in ("Conservative", "Moderate", "Aggressive"):
        result = scenario_results.get(scenario_name, {})
        if "error" in result or not result:
            scenario_dfs[scenario_name] = pd.DataFrame()
            continue

        preset_dict = presets.get(scenario_name, {})
        combined_df = result.get("combined_df", pd.DataFrame())
        if combined_df.empty or "is_forecast" not in combined_df.columns:
            scenario_dfs[scenario_name] = pd.DataFrame()
            continue

        n_months = int(combined_df["is_forecast"].sum())

        # Build per-month CVR and AOV series indexed by start_month offset
        aov_list: list[float] = []
        cvr_list: list[float] = []
        for i in range(n_months):
            cal_month = (start_month - 1 + i) % 12 + 1
            month_sea = (seasonality or {}).get(cal_month, {})
            aov_mod = float(month_sea.get("aov_mod", 0.0)) if (apply_seasonal_aov and seasonality) else 0.0
            cr_mod = float(month_sea.get("cr_mod", 0.0)) if seasonality else 0.0
            aov_list.append(aov * (1.0 + aov_mod))
            cvr_list.append(cvr * (1.0 + cr_mod))

        df = build_scenario_grid_sheet(
            scenario_name=scenario_name,
            scenario_result=result,
            preset=preset_dict,
            cvr_series=cvr_list,
            aov_series=aov_list,
            currency=currency,
            start_month=start_month,
            include_seasonality_column=True,
            seasonality=seasonality,
        )
        scenario_dfs[scenario_name] = df

        ws = wb.create_sheet(title=scenario_name)
        ws.sheet_properties.tabColor = _TAB_COLORS[scenario_name]
        _write_scenario_ws(ws, df, scenario_name, preset_dict, fy_label)

    # ── Comparison sheet ──────────────────────────────────────────────────
    comp_ws = wb.create_sheet(title="Comparison")
    comp_ws.sheet_properties.tabColor = _TAB_COLORS["Comparison"]

    title_text = (
        f"{client_name} | Scenario Comparison — {fy_label} SEO Forecast"
        if client_name
        else f"Scenario Comparison — {fy_label} SEO Forecast"
    )
    title_c = comp_ws.cell(row=1, column=1, value=title_text)
    title_c.font = Font(bold=True, size=14)

    # Row 2: blank

    # Row 3: Header
    comp_headers = ["Metric", "Conservative", "Moderate", "Aggressive"]
    comp_fill = PatternFill(
        start_color=_TAB_COLORS["Comparison"],
        end_color=_TAB_COLORS["Comparison"],
        fill_type="solid",
    )
    for col_idx, h in enumerate(comp_headers, start=1):
        cell = comp_ws.cell(row=3, column=col_idx, value=h)
        cell.font = _WHITE_FONT
        cell.fill = comp_fill
        cell.alignment = Alignment(horizontal="center")

    # Metric rows
    _COMP_METRICS = [
        "Total Traffic (P50)",
        "Total Uplift",
        "Total Revenue",
        "End Traffic (P50)",
        "Monthly Hours",
        "Retainer/month",
        "ROI multiplier",
    ]

    def _extract_comparison_values(sname: str) -> dict:
        df = scenario_dfs.get(sname, pd.DataFrame())
        preset_dict = presets.get(sname, {})
        result = scenario_results.get(sname, {})
        if df.empty or "error" in result:
            return {m: None for m in _COMP_METRICS}

        total_traffic = int(df["Traffic P50"].sum()) if "Traffic P50" in df.columns else 0
        total_uplift = (
            int(df["Positional Uplift"].sum() + df["New Content Uplift"].sum())
            if "Positional Uplift" in df.columns
            else 0
        )
        total_revenue = float(df["Revenue"].sum()) if "Revenue" in df.columns else 0.0
        end_traffic = int(df["Traffic P50"].iloc[-1]) if not df.empty and "Traffic P50" in df.columns else 0
        monthly_hours = float(preset_dict.get("total_monthly_hours", 0.0))
        retainer = float(preset_dict.get("retainer_aud_monthly", 0.0))
        annual_retainer = retainer * 12
        roi = round(total_revenue / annual_retainer, 2) if annual_retainer > 0 else None

        return {
            "Total Traffic (P50)": total_traffic,
            "Total Uplift": total_uplift,
            "Total Revenue": total_revenue,
            "End Traffic (P50)": end_traffic,
            "Monthly Hours": monthly_hours,
            "Retainer/month": retainer,
            "ROI multiplier": roi,
        }

    comp_data = {sn: _extract_comparison_values(sn) for sn in ("Conservative", "Moderate", "Aggressive")}

    _COMP_FMTS = {
        "Total Traffic (P50)": "#,##0",
        "Total Uplift": "#,##0",
        "Total Revenue": "$#,##0.00",
        "End Traffic (P50)": "#,##0",
        "Monthly Hours": "#,##0.0",
        "Retainer/month": "$#,##0.00",
        "ROI multiplier": "0.00",
    }

    for row_off, metric in enumerate(_COMP_METRICS, start=0):
        excel_row = 4 + row_off
        comp_ws.cell(row=excel_row, column=1, value=metric).font = _BOLD_FONT
        for col_idx, sname in enumerate(("Conservative", "Moderate", "Aggressive"), start=2):
            val = comp_data[sname].get(metric)
            cell = comp_ws.cell(row=excel_row, column=col_idx, value=val)
            if val is not None:
                cell.number_format = _COMP_FMTS.get(metric, "#,##0")
            cell.alignment = Alignment(horizontal="right")

    _auto_width(comp_ws)

    # ── Charts sheet ─────────────────────────────────────────────────────────
    _chart_tiers: dict[str, pd.DataFrame] = {}
    for _sname in ("Conservative", "Moderate", "Aggressive"):
        _sdf = scenario_dfs.get(_sname, pd.DataFrame())
        if not _sdf.empty and "Month Label" in _sdf.columns:
            _chart_tiers[_sname] = _sdf[["Month Label", "Baseline Traffic", "Traffic P50"]].rename(
                columns={
                    "Month Label": "month_name",
                    "Baseline Traffic": "baseline_traffic",
                    "Traffic P50": "combined_p50",
                }
            )

    if _chart_tiers:
        _month_labels = list(next(iter(_chart_tiers.values()))["month_name"])
        add_charts_sheet(wb, _chart_tiers, _month_labels)
        if "Charts" in wb.sheetnames:
            wb["Charts"].sheet_properties.tabColor = "0F172A"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
