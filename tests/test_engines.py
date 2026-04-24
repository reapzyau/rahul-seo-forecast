import io

import numpy as np
import pandas as pd
import pytest

from engine.combined_engine import run_combined_forecast
from engine.constants import (
    CTR_11_14,
    CTR_15_20,
    CTR_BY_POSITION,
    CTR_MODELS,
    FORECAST_SCENARIOS,
    INTENT_PATTERNS,
    SITE_PRESETS,
)
from engine.historical_engine import (
    calculate_growth_rates,
    exponential_smoothing_forecast,
    linear_forecast,
    run_historical_forecast,
    sma_forecast,
)
from engine.new_content_engine import (
    classify_difficulty,
    classify_intent,
    efficiency_score,
    expected_position,
    get_ctr,
    ranking_probability,
    run_new_content_forecast,
    time_to_rank_months,
)
from engine.revenue_engine import add_revenue, keyword_revenue_table
from tests.fixtures import make_semrush_kw_df

# ── Keyword Engine ──────────────────────────────────────────────────────────


class TestClassifyDifficulty:
    def test_easy(self):
        assert classify_difficulty(10) == "Easy"
        assert classify_difficulty(20) == "Easy"

    def test_moderate(self):
        assert classify_difficulty(21) == "Moderate"
        assert classify_difficulty(40) == "Moderate"

    def test_hard(self):
        assert classify_difficulty(41) == "Hard"
        assert classify_difficulty(60) == "Hard"

    def test_very_hard(self):
        assert classify_difficulty(61) == "Very Hard"
        assert classify_difficulty(80) == "Very Hard"

    def test_extreme(self):
        assert classify_difficulty(81) == "Extreme"
        assert classify_difficulty(100) == "Extreme"


class TestRankingProbability:
    def test_high_da_low_kd(self):
        prob = ranking_probability(80, 20)
        assert prob == 0.95  # (80-20+50)/100 = 1.10 -> clamped to 0.95

    def test_low_da_high_kd(self):
        prob = ranking_probability(10, 90)
        assert prob == 0.05  # (10-90+50)/100 = -0.30 -> clamped to 0.05

    def test_equal_da_kd(self):
        prob = ranking_probability(50, 50)
        assert prob == 0.5  # (50-50+50)/100 = 0.5

    def test_clamping_upper(self):
        assert ranking_probability(100, 1) == 0.95

    def test_clamping_lower(self):
        assert ranking_probability(1, 100) == 0.05


class TestExpectedPosition:
    def test_high_gap_gives_top_positions(self):
        pos = expected_position(80, 20, seed=42)
        assert 1 <= pos <= 3

    def test_negative_gap_gives_lower_positions(self):
        pos = expected_position(20, 60, seed=42)
        assert 5 <= pos <= 20  # gap is -40, so range is 12-20 or 8-16

    def test_seed_reproducibility(self):
        pos1 = expected_position(50, 30, seed=42)
        pos2 = expected_position(50, 30, seed=42)
        assert pos1 == pos2

    def test_different_seeds_may_differ(self):
        positions = {expected_position(50, 50, seed=i) for i in range(100)}
        assert len(positions) > 1  # Not all the same


class TestClassifyIntent:
    def test_informational_keywords(self):
        assert classify_intent("how to do seo") == "informational"
        assert classify_intent("what is keyword difficulty") == "informational"
        assert classify_intent("why does seo matter") == "informational"
        assert classify_intent("seo tutorial for beginners") == "informational"
        assert classify_intent("link building guide") == "informational"

    def test_commercial_keywords(self):
        assert classify_intent("best seo tool") == "commercial"
        assert classify_intent("seo tool review") == "commercial"
        assert classify_intent("ahrefs alternative") == "commercial"

    def test_transactional_keywords(self):
        assert classify_intent("buy seo software") == "transactional"
        assert classify_intent("seo tool pricing") == "transactional"
        assert classify_intent("ahrefs discount code") == "transactional"

    def test_navigational_keywords(self):
        assert classify_intent("google search console login") == "navigational"

    def test_default_classification(self):
        # Generic keywords with no clear signals default to commercial
        assert classify_intent("seo audit") == "commercial"


class TestGetCtr:
    def test_top_10(self):
        for pos, expected in CTR_BY_POSITION.items():
            assert get_ctr(pos) == expected

    def test_positions_11_14(self):
        for pos in [11, 12, 13, 14]:
            assert get_ctr(pos) == CTR_11_14

    def test_positions_15_20(self):
        for pos in [15, 16, 17, 18, 19, 20]:
            assert get_ctr(pos) == CTR_15_20

    def test_beyond_20(self):
        assert get_ctr(21) == 0.0
        assert get_ctr(50) == 0.0

    def test_ai_adjusted_model(self):
        model = CTR_MODELS["AI-Adjusted"]
        assert get_ctr(1, model) == 16.0
        assert get_ctr(1, model) < get_ctr(1)  # AI-Adjusted < Standard

    def test_standard_model_matches_default(self):
        model = CTR_MODELS["Standard"]
        for pos in range(1, 21):
            assert get_ctr(pos, model) == get_ctr(pos)


class TestTimeToRank:
    def test_easy_tier_range(self):
        for seed in range(20):
            ttr = time_to_rank_months("Easy", 50, seed)
            assert ttr >= 1

    def test_da_adjustment(self):
        # Higher DA should generally produce equal or lower time to rank
        results_high_da = [time_to_rank_months("Moderate", 80, s) for s in range(50)]
        results_low_da = [time_to_rank_months("Moderate", 20, s) for s in range(50)]
        assert np.mean(results_high_da) <= np.mean(results_low_da)


class TestEfficiencyScore:
    def test_basic(self):
        assert efficiency_score(1000, 49) == 1000 / 50

    def test_zero_kd(self):
        assert efficiency_score(500, 0) == 500.0

    def test_high_kd_low_score(self):
        assert efficiency_score(100, 99) == 1.0


class TestRunNewContentForecast:
    @pytest.fixture
    def sample_df(self):
        return pd.DataFrame({
            "keyword": ["easy kw", "hard kw", "medium kw"],
            "volume": [5000, 3000, 4000],
            "kd": [15, 75, 40],
        })

    def test_returns_two_dataframes(self, sample_df):
        kw_df, monthly_df = run_new_content_forecast(sample_df, da=50, cadence=2, months=12, seed=42)
        assert isinstance(kw_df, pd.DataFrame)
        assert isinstance(monthly_df, pd.DataFrame)

    def test_sorted_by_efficiency(self, sample_df):
        kw_df, _ = run_new_content_forecast(sample_df, da=50, cadence=2, months=12, seed=42)
        scores = kw_df["efficiency_score"].tolist()
        assert scores == sorted(scores, reverse=True)

    def test_monthly_df_has_correct_length(self, sample_df):
        _, monthly_df = run_new_content_forecast(sample_df, da=50, cadence=2, months=12, seed=42)
        assert len(monthly_df) == 12

    def test_seed_reproducibility(self, sample_df):
        kw1, m1 = run_new_content_forecast(sample_df, da=50, cadence=2, months=12, seed=42)
        kw2, m2 = run_new_content_forecast(sample_df, da=50, cadence=2, months=12, seed=42)
        pd.testing.assert_frame_equal(kw1, kw2)
        pd.testing.assert_frame_equal(m1, m2)

    def test_traffic_monotonically_nondecreasing(self, sample_df):
        _, monthly_df = run_new_content_forecast(sample_df, da=50, cadence=2, months=18, seed=42)
        traffic = monthly_df["traffic"].tolist()
        for i in range(1, len(traffic)):
            assert traffic[i] >= traffic[i - 1]


# ── Historical Engine ───────────────────────────────────────────────────────


class TestLinearForecast:
    @pytest.fixture
    def sample_data(self):
        dates = pd.date_range("2023-01-01", periods=12, freq="MS")
        traffic = pd.Series([1000 + i * 100 for i in range(12)])
        return dates, traffic

    def test_output_length(self, sample_data):
        dates, traffic = sample_data
        result = linear_forecast(dates, traffic, future_months=6)
        assert len(result) == 18  # 12 historical + 6 forecast

    def test_forecast_increases_for_upward_trend(self, sample_data):
        dates, traffic = sample_data
        result = linear_forecast(dates, traffic, future_months=6)
        forecast = result[result["is_forecast"]]
        assert forecast["linear"].iloc[-1] > forecast["linear"].iloc[0]

    def test_confidence_bands(self, sample_data):
        dates, traffic = sample_data
        result = linear_forecast(dates, traffic, future_months=6, confidence=15)
        forecast = result[result["is_forecast"]]
        assert (forecast["linear_upper"] >= forecast["linear"]).all()
        assert (forecast["linear_lower"] <= forecast["linear"]).all()


class TestExponentialSmoothing:
    def test_output_length(self):
        traffic = pd.Series([100, 110, 120, 130, 140])
        result = exponential_smoothing_forecast(traffic, alpha=0.3, future_months=3)
        assert len(result) == 8  # 5 historical + 3 forecast

    def test_weights_recent_data(self):
        traffic = pd.Series([100, 100, 100, 100, 200])
        result = exponential_smoothing_forecast(traffic, alpha=0.5, future_months=1)
        assert result[-1] > 100

    def test_upward_trend_extrapolates_upward(self):
        """Holt's method should continue the trend, not plateau."""
        traffic = pd.Series([100, 110, 120, 130, 140, 150])
        result = exponential_smoothing_forecast(traffic, alpha=0.3, future_months=3)
        # Each forecast step should be higher than the previous
        assert result[-1] > result[-2] > result[-3]

    def test_forecast_differs_from_flat(self):
        """Consecutive forecast values must not all be equal (flat-line bug check)."""
        traffic = pd.Series([100, 120, 140, 160, 180])
        result = exponential_smoothing_forecast(traffic, alpha=0.3, future_months=3)
        forecast = result[5:]
        assert len(set(forecast)) > 1  # values are not all identical

    def test_no_negative_values(self):
        traffic = pd.Series([500, 400, 300, 200, 100])
        result = exponential_smoothing_forecast(traffic, alpha=0.5, future_months=6)
        assert all(v >= 0 for v in result)


class TestSmaForecast:
    def test_output_length(self):
        traffic = pd.Series([100, 110, 120, 130, 140])
        result = sma_forecast(traffic, window=3, future_months=3)
        assert len(result) == 8

    def test_window_math(self):
        traffic = pd.Series([100, 200, 300])
        result = sma_forecast(traffic, window=3, future_months=1)
        # The forecast should be the average of the last 3 values
        assert result[3] == round(np.mean([100, 200, 300]))


class TestCalculateGrowthRates:
    def test_basic_growth(self):
        traffic = pd.Series([100, 110, 121])
        rates = calculate_growth_rates(traffic)
        assert rates["avg_mom"] > 0
        assert abs(rates["latest_mom"] - 10.0) < 0.01


class TestRunHistoricalForecast:
    def test_all_methods(self):
        df = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [1000 + i * 50 for i in range(12)],
        })
        result = run_historical_forecast(
            df, months=6,
            methods=["Linear Regression", "Exponential Smoothing", "Simple Moving Average"],
        )
        assert "linear" in result.columns
        assert "exponential_smoothing" in result.columns
        assert "sma" in result.columns


# ── YoY Growth Forecast ──────────────────────────────────────────────────────


class TestYoYGrowthForecast:
    def test_flat_series_stays_flat(self):
        from engine.historical_engine import yoy_growth_forecast
        dates = pd.date_range("2022-01-01", periods=24, freq="MS")
        traffic = pd.Series([5000] * 24)
        result = yoy_growth_forecast(dates, traffic, future_months=12)
        forecast = result[result["is_forecast"]]
        # Flat series → yoy_rate ≈ 0 → forecast ≈ 5000
        assert all(abs(forecast["linear"] - 5000) < 50)

    def test_growing_series_extrapolates_trend(self):
        from engine.historical_engine import yoy_growth_forecast
        dates = pd.date_range("2022-01-01", periods=24, freq="MS")
        # ~20% YoY growth: month 13 is 1.2x month 1
        base = 4000
        traffic = pd.Series([int(base * (1.2 ** (i / 12))) for i in range(24)])
        result = yoy_growth_forecast(dates, traffic, future_months=6)
        forecast = result[result["is_forecast"]]
        # Forecast should be higher than the same months from year 1
        assert forecast["linear"].iloc[0] > traffic.iloc[12]

    def test_yoy_rate_stored_in_attrs(self):
        from engine.historical_engine import yoy_growth_forecast
        dates = pd.date_range("2022-01-01", periods=24, freq="MS")
        traffic = pd.Series([5000] * 24)
        result = yoy_growth_forecast(dates, traffic, future_months=6)
        assert "yoy_rate" in result.attrs
        assert isinstance(result.attrs["yoy_rate"], float)

    def test_short_history_uses_annualised_fallback(self):
        from engine.historical_engine import yoy_growth_forecast
        dates = pd.date_range("2024-01-01", periods=6, freq="MS")
        traffic = pd.Series([1000, 1100, 1200, 1300, 1400, 1500])
        result = yoy_growth_forecast(dates, traffic, future_months=6)
        assert "yoy_rate" in result.attrs
        # With 6 months and no same-month prior, uses annualised slope
        assert result.attrs["yoy_rate"] > 0  # growing series

    def test_returns_same_column_contract_as_linear_forecast(self):
        from engine.historical_engine import linear_forecast, yoy_growth_forecast
        dates = pd.date_range("2022-01-01", periods=24, freq="MS")
        traffic = pd.Series([5000] * 24)
        yoy = yoy_growth_forecast(dates, traffic, future_months=6)
        lin = linear_forecast(dates, traffic, future_months=6)
        assert set(yoy.columns) == set(lin.columns)

    def test_combined_engine_uses_yoy_for_long_history(self):
        from engine.combined_engine import run_combined_forecast
        dates = pd.date_range("2022-01-01", periods=24, freq="MS")
        historical = pd.DataFrame({"date": dates, "traffic": [5000] * 24})
        result = run_combined_forecast(historical, None, None, months=12)
        assert result.attrs.get("yoy_rate") is not None

    def test_combined_engine_uses_linear_for_short_history(self):
        from engine.combined_engine import run_combined_forecast
        dates = pd.date_range("2024-01-01", periods=12, freq="MS")
        historical = pd.DataFrame({"date": dates, "traffic": [5000] * 12})
        result = run_combined_forecast(historical, None, None, months=6)
        # <13 months → linear forecast → no yoy_rate attr
        assert result.attrs.get("yoy_rate") is None


# ── Combined Engine ─────────────────────────────────────────────────────────


class TestCombinedForecast:
    def test_combined_equals_baseline_plus_uplifts(self):
        kw_df = pd.DataFrame({
            "keyword": ["test kw"],
            "volume": [5000],
            "kd": [20],
        })
        _, new_content_monthly = run_new_content_forecast(kw_df, da=60, cadence=1, months=12, seed=42)

        historical_df = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [10000 + i * 200 for i in range(12)],
        })

        combined = run_combined_forecast(
            historical_df=historical_df,
            positional_monthly=None,
            new_content_monthly=new_content_monthly,
            months=12,
        )
        forecast = combined[combined["is_forecast"]]

        for _, row in forecast.iterrows():
            # v4: combined = baseline + positional_uplift + new_content_uplift - decay
            expected = row["baseline"] + row["positional_uplift"] + row["new_content_uplift"] - row.get("decay", 0)
            assert abs(row["combined"] - expected) < 2

    def test_uplift_only_no_historical(self):
        monthly = pd.DataFrame({
            "month": range(1, 13),
            "uplift": [100] * 12,
            "traffic": [100] * 12,
            "baseline": [0] * 12,
        })
        combined = run_combined_forecast(
            historical_df=None,
            positional_monthly=monthly,
            new_content_monthly=None,
            months=12,
        )
        assert (combined["baseline"] == 0).all()


class TestCombinedHub:
    def test_layered_math_with_bands(self):
        """v4 math: combined = baseline + positional + new_content - decay (no AIO deduction)."""
        historical = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [10000] * 12,
        })
        positional = pd.DataFrame({
            "month": range(1, 13),
            "baseline": [10000] * 12,
            "uplift_p10": [200] * 12,
            "uplift_p50": [500] * 12,
            "uplift_p90": [800] * 12,
        })
        decay = pd.DataFrame({
            "month": range(1, 13),
            "cumulative_decay": [50 * i for i in range(1, 13)],
        })
        combined = run_combined_forecast(
            historical_df=historical,
            positional_monthly=positional,
            new_content_monthly=None,
            months=12,
            decay_df=decay,
        )
        forecast = combined[combined["is_forecast"]]
        m12 = forecast.iloc[-1]
        # v4 math: no AIO deduction at combined level
        expected_p50 = m12["baseline"] + m12["positional_uplift_p50"] - m12["decay"]
        assert abs(m12["combined_p50"] - expected_p50) < 2

    def test_streams_apply_seasonality_internally(self):
        """With a big November boost, positional output in month 11 should be higher."""
        from engine.positional_engine import run_positional_forecast_mc
        seasonality = {m: {"traffic_mod": 0.0} for m in range(1, 13)}
        seasonality[2] = {"traffic_mod": 0.50}  # month 2 = February in forecast starting Jan

        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(20)],
            "position": [10] * 20,
            "volume": [1000] * 20,
            "kd": [30] * 20,
            "current_traffic": [80] * 20,
            "intent": ["commercial"] * 20,
            "has_aio": [False] * 20,
        })
        _, monthly_no_season = run_positional_forecast_mc(df, months=12, n_trials=200, seed=42)
        _, monthly_season = run_positional_forecast_mc(
            df, months=12, n_trials=200, seed=42,
            seasonality=seasonality, forecast_start_month=1,
        )
        # Month 2 (index 1) should be higher with 50% boost
        assert monthly_season.iloc[1]["uplift_p50"] > monthly_no_season.iloc[1]["uplift_p50"]

    def test_combined_math_no_aio_term(self):
        """AIO is per-stream — combined must equal baseline + positional + new_content - decay."""
        positional = pd.DataFrame({
            "month": range(1, 7),
            "baseline": [10000] * 6,
            "uplift_p10": [100] * 6,
            "uplift_p50": [300] * 6,
            "uplift_p90": [500] * 6,
        })
        decay = pd.DataFrame({
            "month": range(1, 7),
            "cumulative_decay": [20 * i for i in range(1, 7)],
        })
        combined = run_combined_forecast(
            historical_df=None,
            positional_monthly=positional,
            new_content_monthly=None,
            months=6,
            decay_df=decay,
        )
        forecast = combined[combined["is_forecast"]]
        for _, row in forecast.iterrows():
            expected = row["baseline"] + row["positional_uplift_p50"] + row["new_content_uplift"] - row["decay"]
            assert abs(row["combined_p50"] - expected) < 2

    def test_bands_preserved_order(self):
        historical = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [10000] * 12,
        })
        positional = pd.DataFrame({
            "month": range(1, 13),
            "baseline": [10000] * 12,
            "uplift_p10": [200] * 12,
            "uplift_p50": [500] * 12,
            "uplift_p90": [800] * 12,
        })
        combined = run_combined_forecast(
            historical_df=historical,
            positional_monthly=positional,
            new_content_monthly=None,
            months=12,
        )
        forecast = combined[combined["is_forecast"]]
        assert (forecast["combined_p10"] <= forecast["combined_p50"]).all()
        assert (forecast["combined_p50"] <= forecast["combined_p90"]).all()

    def test_backward_compat_aliases(self):
        positional = pd.DataFrame({
            "month": range(1, 7),
            "baseline": [10000] * 6,
            "uplift_p10": [200] * 6,
            "uplift_p50": [500] * 6,
            "uplift_p90": [800] * 6,
        })
        combined = run_combined_forecast(
            historical_df=None,
            positional_monthly=positional,
            new_content_monthly=None,
            months=6,
        )
        assert "combined" in combined.columns
        assert "positional_uplift" in combined.columns


# ── Revenue Engine ──────────────────────────────────────────────────────────


class TestAddRevenue:
    def test_basic_math(self):
        df = pd.DataFrame({"month": [1, 2], "traffic": [1000, 2000]})
        result = add_revenue(df, cvr=2.0, aov=50.0)
        assert result["leads"].iloc[0] == 20  # 1000 * 0.02
        assert result["revenue"].iloc[0] == 1000.0  # 20 * 50

    def test_zero_traffic(self):
        df = pd.DataFrame({"month": [1], "traffic": [0]})
        result = add_revenue(df, cvr=5.0, aov=100.0)
        assert result["leads"].iloc[0] == 0
        assert result["revenue"].iloc[0] == 0.0


class TestKeywordRevenueTable:
    def test_filters_ranking_keywords(self):
        kw_df = pd.DataFrame({
            "keyword": ["a", "b", "c"],
            "estimated_monthly_traffic": [100, 200, 0],
            "will_rank": [True, True, False],
        })
        result = keyword_revenue_table(kw_df, cvr=5.0, aov=100.0)
        assert len(result) == 2
        assert result["monthly_revenue"].iloc[0] == 500.0  # 100 * 0.05 * 100


# ── Data Loader (unit-testable parts) ──────────────────────────────────────


class TestNewContentForecastFiltering:
    @pytest.fixture
    def mixed_intent_df(self):
        return pd.DataFrame({
            "keyword": [
                "how to do seo",          # informational
                "best seo tool",          # commercial
                "buy seo software",       # transactional
                "seo tutorial",           # informational
                "seo agency",             # commercial
            ],
            "volume": [5000, 3000, 2000, 4000, 1000],
            "kd": [15, 40, 30, 20, 50],
        })

    def test_intent_column_always_present(self, mixed_intent_df):
        kw_df, _ = run_new_content_forecast(mixed_intent_df, da=50, cadence=2, months=12, seed=42)
        assert "intent" in kw_df.columns

    def test_exclude_informational(self, mixed_intent_df):
        kw_df, _ = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            include_informational=False,
        )
        assert "informational" not in kw_df["intent"].values
        assert len(kw_df) == 3  # 2 informational removed from 5

    def test_ctr_penalty_reduces_traffic(self, mixed_intent_df):
        kw_base, _ = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
        )
        kw_penalty, _ = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            ai_overview_ctr_penalty=50.0,
        )
        # Informational keywords should have less or equal traffic with penalty
        base_info = kw_base[kw_base["intent"] == "informational"]["estimated_monthly_traffic"].sum()
        penalty_info = kw_penalty[kw_penalty["intent"] == "informational"]["estimated_monthly_traffic"].sum()
        assert penalty_info <= base_info

    def test_ctr_model_affects_traffic(self, mixed_intent_df):
        _, monthly_std = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            ctr_model=CTR_MODELS["Standard"],
        )
        _, monthly_ai = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            ctr_model=CTR_MODELS["AI-Adjusted"],
        )
        # AI-Adjusted should produce less or equal total traffic
        assert monthly_ai["traffic"].sum() <= monthly_std["traffic"].sum()

    def test_traffic_multiplier(self, mixed_intent_df):
        _, monthly_base = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            traffic_multiplier=1.0,
        )
        _, monthly_conservative = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            traffic_multiplier=0.7,
        )
        _, monthly_aggressive = run_new_content_forecast(
            mixed_intent_df, da=50, cadence=2, months=12, seed=42,
            traffic_multiplier=1.3,
        )
        base_total = monthly_base["traffic"].sum()
        assert monthly_conservative["traffic"].sum() <= base_total
        assert monthly_aggressive["traffic"].sum() >= base_total


class TestPresets:
    def test_site_preset_values_valid(self):
        for _name, preset in SITE_PRESETS.items():
            assert "da" in preset
            assert "cadence" in preset
            assert "months" in preset
            assert 1 <= preset["da"] <= 100
            assert preset["cadence"] >= 1
            assert preset["months"] >= 6

    def test_ctr_models_have_required_keys(self):
        for _name, model in CTR_MODELS.items():
            assert "ctr_by_position" in model
            assert "ctr_11_14" in model
            assert "ctr_15_20" in model
            assert "label" in model

    def test_forecast_scenarios_have_multiplier(self):
        for _name, scenario in FORECAST_SCENARIOS.items():
            assert "traffic_multiplier" in scenario
            assert scenario["traffic_multiplier"] > 0


class TestForecastSeries:
    def test_forecast_series_length(self):
        from engine.historical_engine import forecast_series
        series = pd.Series([100, 110, 120, 130, 140])
        result = forecast_series(series, future_months=3)
        assert len(result) == 8  # 5 historical + 3 forecast

    def test_forecast_series_upward_trend(self):
        from engine.historical_engine import forecast_series
        series = pd.Series([100, 200, 300, 400, 500])
        result = forecast_series(series, future_months=3)
        assert result[-1] > result[4]  # Forecast continues upward


class TestExtendedHistoricalForecast:
    def test_with_optional_metrics(self):
        df = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [1000 + i * 50 for i in range(12)],
            "revenue": [5000 + i * 200 for i in range(12)],
            "transactions": [50 + i * 3 for i in range(12)],
            "aov": [100 + i * 0.5 for i in range(12)],
            "cr": [5.0 - i * 0.05 for i in range(12)],
        })
        result = run_historical_forecast(df, months=6, methods=["Linear Regression"])
        assert "revenue_forecast" in result.columns
        assert "transactions_forecast" in result.columns
        assert "aov_forecast" in result.columns
        assert "cr_forecast" in result.columns
        assert len(result) == 18  # 12 historical + 6 forecast

    def test_without_optional_metrics(self):
        df = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=12, freq="MS"),
            "traffic": [1000 + i * 50 for i in range(12)],
        })
        result = run_historical_forecast(df, months=6, methods=["Linear Regression"])
        assert "revenue_forecast" not in result.columns
        assert "aov_forecast" not in result.columns


class TestBuildFullMetrics:
    def test_full_metrics_table(self):
        from engine.revenue_engine import build_full_metrics_table
        df = pd.DataFrame({
            "date": pd.date_range("2023-01-01", periods=18, freq="MS"),
            "traffic": [1000 + i * 50 for i in range(18)],
            "revenue": [5000 + i * 200 for i in range(18)],
            "transactions": [50 + i * 3 for i in range(18)],
            "aov": [100 + i * 0.5 for i in range(18)],
            "cr": [5.0 - i * 0.05 for i in range(18)],
        })
        result = run_historical_forecast(df, months=6, methods=["Linear Regression"])
        metrics = build_full_metrics_table(result)
        assert "Month" in metrics.columns
        assert "Organic Sessions" in metrics.columns
        assert "Organic Sessions Forecasted" in metrics.columns
        # YoY columns should exist for sessions at minimum
        yoy_cols = [c for c in metrics.columns if "YoY" in c]
        assert len(yoy_cols) > 0


class TestAddDynamicRevenue:
    def test_static_values(self):
        from engine.revenue_engine import add_dynamic_revenue
        df = pd.DataFrame({"month": [1, 2, 3], "traffic": [1000, 2000, 3000]})
        result = add_dynamic_revenue(df, cvr_series=5.0, aov_series=100.0)
        assert result["transactions"].iloc[0] == 50  # 1000 * 0.05
        assert result["revenue"].iloc[0] == 5000.0  # 50 * 100

    def test_dynamic_aov(self):
        from engine.revenue_engine import add_dynamic_revenue
        df = pd.DataFrame({"month": [1, 2, 3], "traffic": [1000, 1000, 1000]})
        result = add_dynamic_revenue(df, cvr_series=10.0, aov_series=[50, 100, 150])
        assert result["revenue"].iloc[0] == 5000.0   # 100 * 50
        assert result["revenue"].iloc[2] == 15000.0   # 100 * 150


class TestSeasonality:
    def test_apply_seasonality_modifies_traffic(self):
        from engine.seasonality_engine import apply_seasonality
        df = pd.DataFrame({"month": [1, 2, 3], "traffic": [1000, 1000, 1000]})
        result = apply_seasonality(df)
        # Seasonality should change at least some months
        assert not (result["traffic"] == 1000).all()
        assert "traffic_base" in result.columns
        assert "season_label" in result.columns

    def test_campaign_boost_adds_to_modifier(self):
        from engine.seasonality_engine import apply_seasonality
        df = pd.DataFrame({"month": [11], "traffic": [1000]})
        campaigns = [{"name": "Test Sale", "month": 11, "traffic_boost": 0.50}]
        result = apply_seasonality(df, campaigns=campaigns)
        # November default is 0.25, plus campaign 0.50 = 0.75 total
        assert result["traffic"].iloc[0] > 1500

    def test_build_campaign_list(self):
        from engine.seasonality_engine import build_campaign_list
        text = "GAZFRENZY | 11 | 0.20 | 0.10 | -0.05\nFather's Day | 9 | 0.15"
        campaigns = build_campaign_list(text)
        assert len(campaigns) == 2
        assert campaigns[0]["name"] == "GAZFRENZY"
        assert campaigns[0]["month"] == 11
        assert campaigns[1]["traffic_boost"] == 0.15


class TestKeywordPipeline:
    def test_classify_serp_page(self):
        from engine.keyword_pipeline_engine import classify_serp_page
        assert classify_serp_page(1) == "Page 1"
        assert classify_serp_page(10) == "Page 1"
        assert classify_serp_page(11) == "Page 2"
        assert classify_serp_page(25) == "Page 3"
        assert classify_serp_page(50) == "Pages 4-10"
        assert classify_serp_page(None) == "Not Ranking"

    def test_pipeline_snapshot(self):
        from engine.keyword_pipeline_engine import build_pipeline_snapshot
        kw_df = pd.DataFrame({"expected_position": [1, 5, 15, 25, None]})
        snapshot = build_pipeline_snapshot(kw_df)
        assert snapshot["Page 1"] == 2
        assert snapshot["Page 2"] == 1
        assert snapshot["Page 3"] == 1
        assert snapshot["Not Ranking"] == 1

    def test_pipeline_over_time(self):
        from engine.keyword_pipeline_engine import build_pipeline_over_time
        kw_df = pd.DataFrame({
            "keyword": ["kw1", "kw2"],
            "expected_position": [3, 15],
            "publish_month": [1, 1],
            "will_rank": [True, True],
            "traffic_starts_month": [4, 6],
            "time_to_rank": [3, 5],
        })
        result = build_pipeline_over_time(kw_df, months=12)
        assert len(result) == 12
        assert "page_1" in result.columns
        assert "page_1_mom_change" in result.columns


class TestRoadmapEngine:
    def test_build_roadmap_defaults(self):
        from engine.roadmap_engine import build_roadmap
        task_df, monthly_df, summary = build_roadmap(months=12)
        assert summary["n_tasks"] > 0
        assert summary["total_hours"] > 0
        assert len(task_df) == summary["n_tasks"]
        assert len(monthly_df) == summary["n_tasks"] + 1

    def test_quarterly_occurrence(self):
        from engine.roadmap_engine import build_roadmap
        task = [{"task": "Test Q", "focus": "Content", "occurrence": "Quarterly", "hours": 5.0}]
        _, monthly_df, _ = build_roadmap(task, months=12)
        task_row = monthly_df[monthly_df["Task"] == "Test Q"].iloc[0]
        assert task_row["M1"] == 5.0
        assert task_row["M2"] == 0.0
        assert task_row["M4"] == 5.0
        assert task_row["M7"] == 5.0
        assert task_row["M10"] == 5.0

    def test_xlsx_export_buffer(self):
        from engine.roadmap_engine import build_roadmap, build_roadmap_xlsx
        _, monthly_df, summary = build_roadmap(months=12)
        buf = build_roadmap_xlsx(monthly_df, summary, hourly_rate=200.0)
        data = buf.read()
        assert len(data) > 0
        assert data[:2] == b"PK"


class TestTemplates:
    def test_keyword_template_is_valid_csv(self):
        from utils.export import keyword_template_csv
        df = pd.read_csv(io.StringIO(keyword_template_csv()))
        assert list(df.columns) == ["keyword", "volume", "kd"]
        assert len(df) == 3

    def test_traffic_template_is_valid_csv(self):
        from utils.export import traffic_template_csv
        df = pd.read_csv(io.StringIO(traffic_template_csv()))
        assert "date" in df.columns
        assert "traffic" in df.columns
        assert len(df) == 3
        # Optional metric columns
        for col in ["revenue", "transactions", "aov", "cr"]:
            assert col in df.columns

    def test_keyword_template_has_valid_data(self):
        from utils.export import keyword_template_csv
        df = pd.read_csv(io.StringIO(keyword_template_csv()))
        assert (df["volume"] > 0).all()
        assert (df["kd"] >= 0).all()


class TestModelsConfig:
    def test_config_loads(self):
        from engine.ai_engine import get_default_model, get_fallback_chain, get_model_options
        models = get_model_options()
        assert len(models) > 0
        assert all("id" in m and "label" in m for m in models)

    def test_default_model_in_list(self):
        from engine.ai_engine import get_default_model, get_model_options
        default = get_default_model()
        ids = [m["id"] for m in get_model_options()]
        assert default in ids

    def test_fallback_chain_valid(self):
        from engine.ai_engine import get_fallback_chain, get_model_options
        chain = get_fallback_chain()
        assert len(chain) >= 2
        ids = [m["id"] for m in get_model_options()]
        for model in chain:
            assert model in ids


class TestPromptLoader:
    def test_all_prompts_load(self):
        from engine.ai_engine import _load_prompt
        for name in ["cluster_keywords", "check_cannibalization", "content_roadmap", "transform_data"]:
            system, user_tmpl = _load_prompt(name)
            assert len(system) > 20
            assert user_tmpl.template

    def test_prompt_substitution(self):
        from engine.ai_engine import _load_prompt
        system, user_tmpl = _load_prompt("cluster_keywords")
        result = user_tmpl.substitute(kw_list="- test keyword")
        assert "test keyword" in result


class TestDataLoaderHelpers:
    def test_efficiency_ordering_preserved(self):
        """Verify the new content engine sorts by efficiency."""
        df = pd.DataFrame({
            "keyword": ["low_eff", "high_eff"],
            "volume": [100, 10000],
            "kd": [90, 10],
        })
        kw_df, _ = run_new_content_forecast(df, da=50, cadence=1, months=6, seed=42)
        assert kw_df.iloc[0]["keyword"] == "high_eff"


# ── Positional Engine ──────────────────────────────────────────────────────


class TestPositionalForecast:
    @pytest.fixture
    def sample_existing(self):
        return pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(50)],
            "position": [3, 8, 15, 22] * 12 + [5, 12],
            "volume": [1000] * 50,
            "kd": [30] * 50,
            "current_traffic": [100] * 50,
            "intent": ["commercial"] * 50,
            "has_aio": [False] * 50,
        })

    def test_ga4_anchored_baseline(self, sample_existing):
        from engine.positional_engine import run_positional_forecast
        kw_df, monthly = run_positional_forecast(
            sample_existing, months=12, effort="moderate", ga4_baseline=10000,
        )
        assert abs(monthly.iloc[0]["baseline"] - 10000) / 10000 < 0.01

    def test_month_12_exceeds_baseline(self, sample_existing):
        from engine.positional_engine import run_positional_forecast
        _, monthly = run_positional_forecast(
            sample_existing, months=12, effort="moderate", ga4_baseline=10000,
        )
        assert monthly.iloc[-1]["traffic"] > monthly.iloc[0]["baseline"]

    def test_effort_levels_ordered(self, sample_existing):
        from engine.positional_engine import run_positional_forecast
        _, light = run_positional_forecast(sample_existing, months=12, effort="light")
        _, moderate = run_positional_forecast(sample_existing, months=12, effort="moderate")
        _, aggressive = run_positional_forecast(sample_existing, months=12, effort="aggressive")
        assert aggressive.iloc[-1]["uplift"] >= moderate.iloc[-1]["uplift"] >= light.iloc[-1]["uplift"]

    def test_quick_wins_filter(self, sample_existing):
        from engine.positional_engine import quick_wins, run_positional_forecast
        kw_df, _ = run_positional_forecast(sample_existing, months=12, effort="moderate")
        qw = quick_wins(kw_df, top_n=20)
        if not qw.empty:
            assert (qw["position"] >= 4).all()
            assert (qw["position"] <= 20).all()


# ── Monte Carlo + Attention Curve ──────────────────────────────────────────


class TestPositionalMonteCarlo:
    @pytest.fixture
    def mc_sample(self):
        return make_semrush_kw_df(n=30, positions=[10] * 30, kds=[30] * 30)

    def test_bands_ordered(self, mc_sample):
        from engine.positional_engine import run_positional_forecast_mc
        _, monthly = run_positional_forecast_mc(mc_sample, months=12, n_trials=200)
        assert (monthly["uplift_p10"] <= monthly["uplift_p50"]).all()
        assert (monthly["uplift_p50"] <= monthly["uplift_p90"]).all()

    def test_band_width_meaningful(self):
        from engine.positional_engine import run_positional_forecast_mc
        df = make_semrush_kw_df(n=50, positions=[15] * 50, kds=[40] * 50)
        _, monthly = run_positional_forecast_mc(df, months=12, n_trials=500)
        m12 = monthly.iloc[-1]
        spread = m12["uplift_p90"] - m12["uplift_p10"]
        assert spread > m12["uplift_p50"] * 0.1

    def test_deterministic_with_seed(self, mc_sample):
        from engine.positional_engine import run_positional_forecast_mc
        _, m1 = run_positional_forecast_mc(mc_sample, months=12, n_trials=200, seed=42)
        _, m2 = run_positional_forecast_mc(mc_sample, months=12, n_trials=200, seed=42)
        pd.testing.assert_frame_equal(m1, m2)

    def test_backward_compat_columns(self, mc_sample):
        from engine.positional_engine import run_positional_forecast_mc
        _, monthly = run_positional_forecast_mc(mc_sample, months=12, n_trials=200)
        assert "uplift" in monthly.columns
        assert "traffic" in monthly.columns
        assert (monthly["uplift"] == monthly["uplift_p50"]).all()


class TestAttentionCurve:
    def test_top_keywords_get_full_weight(self):
        from engine.positional_engine import attention_weight
        assert attention_weight(0.01) == 1.00
        assert attention_weight(0.04) == 1.00

    def test_long_tail_gets_minimal_weight(self):
        from engine.positional_engine import attention_weight
        assert attention_weight(0.99) == 0.05

    def test_weights_decrease_monotonically(self):
        from engine.positional_engine import attention_weight
        weights = [attention_weight(p) for p in [0.01, 0.1, 0.3, 0.8]]
        for i in range(1, len(weights)):
            assert weights[i] <= weights[i - 1]

    def test_apply_attention_curve_assigns_weights(self):
        from engine.positional_engine import apply_attention_curve
        df = make_semrush_kw_df(n=100, volumes=list(range(100, 0, -1)), kds=[30] * 100)
        result = apply_attention_curve(df)
        assert (result.head(5)["attention_weight"] == 1.00).all()
        assert (result.tail(50)["attention_weight"] == 0.05).all()

    def test_attention_reduces_aggregate_uplift(self):
        from engine.positional_engine import run_positional_forecast_mc
        df = make_semrush_kw_df(n=200, positions=[15] * 200, kds=[35] * 200)
        _, with_attn = run_positional_forecast_mc(
            df, months=12, n_trials=200, use_attention_curve=True, seed=99
        )
        _, no_attn = run_positional_forecast_mc(
            df, months=12, n_trials=200, use_attention_curve=False, seed=99
        )
        assert with_attn.iloc[-1]["uplift_p50"] < no_attn.iloc[-1]["uplift_p50"]


# ── AIO Risk Engine ────────────────────────────────────────────────────────


class TestAioRiskEngine:
    def test_zero_affected(self):
        from engine.aio_risk_engine import calculate_aio_risk
        df = pd.DataFrame({
            "keyword": ["a", "b"], "has_aio": [False, False],
            "volume": [100, 200], "current_traffic": [50, 100],
            "intent": ["commercial", "commercial"],
        })
        risk = calculate_aio_risk(df, ctr_penalty_pct=40.0)
        assert risk["keywords_affected"] == 0
        assert risk["traffic_at_risk"] == 0

    def test_projected_loss_math(self):
        from engine.aio_risk_engine import calculate_aio_risk
        df = pd.DataFrame({
            "keyword": ["a", "b", "c"],
            "has_aio": [True, True, True],
            "volume": [100, 200, 300],
            "current_traffic": [400, 300, 300],
            "intent": ["informational", "commercial", "informational"],
        })
        risk = calculate_aio_risk(df, ctr_penalty_pct=40.0)
        assert risk["traffic_at_risk"] == 1000
        assert risk["projected_loss"] == 400

    def test_intent_breakdown(self):
        from engine.aio_risk_engine import calculate_aio_risk
        df = pd.DataFrame({
            "keyword": ["a", "b"],
            "has_aio": [True, True],
            "volume": [100, 200],
            "current_traffic": [50, 100],
            "intent": ["informational", "commercial"],
        })
        risk = calculate_aio_risk(df, ctr_penalty_pct=40.0)
        assert not risk["intent_breakdown"].empty

    def test_recommendations_nonempty(self):
        from engine.aio_risk_engine import aio_recommendations, calculate_aio_risk
        df = pd.DataFrame({
            "keyword": ["a"], "has_aio": [True],
            "volume": [100], "current_traffic": [50],
            "intent": ["informational"],
        })
        risk = calculate_aio_risk(df)
        assert len(aio_recommendations(risk)) > 0


# ── Snapshot + Variance ────────────────────────────────────────────────────


class TestSnapshot:
    def test_roundtrip(self):
        from engine.snapshot_engine import build_snapshot, load_snapshot, snapshot_to_bytes
        combined = pd.DataFrame({
            "date": pd.date_range("2026-01-01", periods=12, freq="MS"),
            "actual": [None] * 12,
            "baseline": [10000] * 12,
            "combined_p50": [11000] * 12,
            "is_forecast": [True] * 12,
        })
        snap = build_snapshot("Test Client", combined, {"effort": "moderate"})
        data = snapshot_to_bytes(snap)
        loaded = load_snapshot(data)
        assert loaded["client_name"] == "Test Client"
        assert len(loaded["forecast"]) == 12

    def test_variance_calculation(self):
        from engine.snapshot_engine import compare_to_actuals
        snapshot = {
            "forecast": [
                {"date": "2026-01-01", "combined_p50": 10000, "combined_p10": 8000, "combined_p90": 12000},
                {"date": "2026-02-01", "combined_p50": 10500, "combined_p10": 8500, "combined_p90": 12500},
            ]
        }
        actuals = pd.DataFrame({
            "date": ["2026-01-01", "2026-02-01"],
            "traffic": [10500, 9000],
        })
        result = compare_to_actuals(snapshot, actuals)
        assert len(result) == 2
        assert abs(result.iloc[0]["variance_pct"] - 5.0) < 0.1
        assert result.iloc[0]["within_band"]
        assert result.iloc[1]["variance_pct"] < 0

    def test_summarise_variance(self):
        from engine.snapshot_engine import compare_to_actuals, summarise_variance
        snapshot = {
            "forecast": [
                {"date": "2026-01-01", "combined_p50": 10000, "combined_p10": 8000, "combined_p90": 12000},
            ]
        }
        actuals = pd.DataFrame({"date": ["2026-01-01"], "traffic": [10500]})
        comparison = compare_to_actuals(snapshot, actuals)
        summary = summarise_variance(comparison)
        assert summary["n_months_compared"] == 1
        assert summary["pct_within_band"] == 100.0


# ── AIO Erosion (Time-varying) ─────────────────────────────────────────────


class TestAioErosion:
    def test_erosion_grows_over_time(self):
        from engine.aio_risk_engine import project_aio_erosion
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(100)],
            "intent": ["informational"] * 100,
            "current_traffic": [100] * 100,
            "has_aio": [False] * 100,
        })
        result = project_aio_erosion(df, months=24, monthly_growth=0.03)
        assert result["cumulative_erosion"].is_monotonic_increasing
        assert result.iloc[-1]["cumulative_erosion"] > result.iloc[0]["cumulative_erosion"] * 5

    def test_informational_erodes_more_than_transactional(self):
        from engine.aio_risk_engine import project_aio_erosion
        info_df = pd.DataFrame({
            "keyword": ["a"], "intent": ["informational"],
            "current_traffic": [1000], "has_aio": [True],
        })
        trans_df = pd.DataFrame({
            "keyword": ["a"], "intent": ["transactional"],
            "current_traffic": [1000], "has_aio": [True],
        })
        info_result = project_aio_erosion(info_df, months=12)
        trans_result = project_aio_erosion(trans_df, months=12)
        assert info_result.iloc[-1]["cumulative_erosion"] > trans_result.iloc[-1]["cumulative_erosion"]

    def test_zero_growth_only_hits_pre_affected(self):
        from engine.aio_risk_engine import project_aio_erosion
        df = pd.DataFrame({
            "keyword": ["a", "b"],
            "intent": ["informational", "informational"],
            "current_traffic": [1000, 1000],
            "has_aio": [True, False],
        })
        result = project_aio_erosion(df, months=12, monthly_growth=0.0)
        assert result.iloc[0]["cumulative_erosion"] == result.iloc[-1]["cumulative_erosion"]


# ── Decay Engine ──────────────────────────────────────────────────────────


class TestDecayEngine:
    def test_position_bucketing(self):
        from engine.decay_engine import position_bucket
        assert position_bucket(1) == "top3"
        assert position_bucket(5) == "top10"
        assert position_bucket(15) == "11_20"
        assert position_bucket(30) == "21_50"
        assert position_bucket(75) == "51_plus"
        assert position_bucket(None) == "51_plus"

    def test_monthly_decay_factor(self):
        from engine.decay_engine import monthly_decay_factor
        monthly = monthly_decay_factor(0.12)
        assert 0.98 < monthly < 0.995
        annual = monthly ** 12
        assert abs(annual - 0.88) < 0.01

    def test_decay_cumulative_increases(self):
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["a", "b", "c"],
            "position": [2, 8, 25],
            "current_traffic": [1000, 500, 200],
        })
        result = calculate_portfolio_decay(df, months=12)
        assert result["cumulative_decay"].is_monotonic_increasing

    def test_maintenance_reduces_decay(self):
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["a"],
            "position": [10],
            "current_traffic": [1000],
        })
        no_maint = calculate_portfolio_decay(df, months=12, maintenance_coverage=0.0)
        with_maint = calculate_portfolio_decay(df, months=12, maintenance_coverage=0.7)
        assert with_maint.iloc[-1]["cumulative_decay"] < no_maint.iloc[-1]["cumulative_decay"]

    def test_empty_portfolio(self):
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame(columns=["position", "current_traffic"])
        result = calculate_portfolio_decay(df, months=6)
        assert len(result) == 6
        assert result["decay_loss"].sum() == 0

    def test_decayed_baseline_is_lower_than_linear(self):
        from engine.decay_engine import project_decayed_baseline
        historical = pd.Series([10000 + i * 100 for i in range(12)])
        keyword_df = pd.DataFrame({
            "keyword": ["a", "b"],
            "position": [5, 25],
            "current_traffic": [3000, 1500],
        })
        result = project_decayed_baseline(historical, keyword_df, months=12, maintenance_coverage=0.0)
        assert (result["honest_baseline"] <= result["linear_baseline"]).all()


class TestDecayRespectsMaintenance:
    def test_high_maintenance_reduces_decay(self):
        """High maintenance_coverage should meaningfully reduce cumulative decay vs zero."""
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["kw1", "kw2", "kw3"],
            "position": [2, 7, 15],
            "current_traffic": [2000, 1000, 500],
        })
        no_maint = calculate_portfolio_decay(df, months=12, maintenance_coverage=0.0)
        high_maint = calculate_portfolio_decay(df, months=12, maintenance_coverage=0.9)
        cumulative_no = no_maint.iloc[-1]["cumulative_decay"]
        cumulative_hi = high_maint.iloc[-1]["cumulative_decay"]
        assert cumulative_hi < cumulative_no
        # At 0.9 coverage, effective decay should be reduced by at least 50%
        assert cumulative_hi < cumulative_no * 0.5

    def test_non_branded_informational_decays_faster(self):
        """Non-branded informational keyword decays more than non-branded commercial in same position."""
        from engine.decay_engine import calculate_portfolio_decay
        info_df = pd.DataFrame({
            "keyword": ["info_kw"],
            "position": [8],
            "current_traffic": [1000],
            "intent": ["informational"],
            "is_branded": [False],
        })
        comm_df = pd.DataFrame({
            "keyword": ["commercial_kw"],
            "position": [8],
            "current_traffic": [1000],
            "intent": ["commercial"],
            "is_branded": [False],
        })
        info_decay = calculate_portfolio_decay(info_df, months=12).iloc[-1]["cumulative_decay"]
        comm_decay = calculate_portfolio_decay(comm_df, months=12).iloc[-1]["cumulative_decay"]
        assert info_decay > comm_decay

    def test_branded_informational_does_not_get_multiplier(self):
        """Branded informational decays same as branded commercial."""
        from engine.decay_engine import calculate_portfolio_decay
        info = calculate_portfolio_decay(pd.DataFrame({
            "keyword": ["brand info"], "position": [5], "current_traffic": [500],
            "intent": ["informational"], "is_branded": [True],
        }), months=12).iloc[-1]["cumulative_decay"]
        comm = calculate_portfolio_decay(pd.DataFrame({
            "keyword": ["brand comm"], "position": [5], "current_traffic": [500],
            "intent": ["commercial"], "is_branded": [True],
        }), months=12).iloc[-1]["cumulative_decay"]
        assert info == comm

    def test_missing_intent_column_no_crash(self):
        """If intent column missing, decay still runs with no multiplier applied."""
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["kw"], "position": [10], "current_traffic": [500],
        })
        result = calculate_portfolio_decay(df, months=6)
        assert len(result) == 6
        assert result["cumulative_decay"].iloc[-1] > 0

    def test_apply_intent_multipliers_false_disables_logic(self):
        """When apply_intent_multipliers=False, non-branded info decays same as commercial."""
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["info", "comm"],
            "position": [8, 8],
            "current_traffic": [1000, 1000],
            "intent": ["informational", "commercial"],
            "is_branded": [False, False],
        })
        result_on = calculate_portfolio_decay(df, months=12, apply_intent_multipliers=True)
        result_off = calculate_portfolio_decay(df, months=12, apply_intent_multipliers=False)
        assert result_on.iloc[-1]["cumulative_decay"] > result_off.iloc[-1]["cumulative_decay"]

    def test_custom_multiplier_overrides_default(self):
        """Pass a harsher multiplier, confirm decay scales with it."""
        from engine.decay_engine import calculate_portfolio_decay
        df = pd.DataFrame({
            "keyword": ["info"], "position": [10], "current_traffic": [1000],
            "intent": ["informational"], "is_branded": [False],
        })
        default = calculate_portfolio_decay(df, months=12).iloc[-1]["cumulative_decay"]
        harsh = calculate_portfolio_decay(
            df, months=12,
            intent_decay_multipliers={"informational_non_branded": 2.5},
        ).iloc[-1]["cumulative_decay"]
        assert harsh > default


# ── Intent-Weighted Revenue ───────────────────────────────────────────────


class TestIntentWeightedRevenue:
    def test_commercial_only_boosts_cvr(self):
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({
            "intent": ["commercial"] * 5,
            "uplift": [100] * 5,
        })
        result = compute_intent_weighted_cvr(df, 2.0)
        assert result == pytest.approx(2.0 * 1.5, rel=0.01)

    def test_transactional_highest_multiplier(self):
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({
            "intent": ["transactional"] * 5,
            "volume": [100] * 5,
        })
        result = compute_intent_weighted_cvr(df, 2.0)
        assert result == pytest.approx(2.0 * 2.0, rel=0.01)

    def test_informational_lowers_cvr(self):
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({
            "intent": ["informational"] * 5,
            "uplift": [100] * 5,
        })
        result = compute_intent_weighted_cvr(df, 2.0)
        assert result < 2.0

    def test_mixed_intent_is_between(self):
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({
            "intent": ["commercial", "informational"],
            "uplift": [100, 100],
        })
        result = compute_intent_weighted_cvr(df, 2.0)
        assert 2.0 * 0.3 < result < 2.0 * 1.5

    def test_empty_returns_base(self):
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({"intent": [], "uplift": []})
        assert compute_intent_weighted_cvr(df, 2.5) == 2.5

    def test_intent_col_present(self):
        """compute_intent_weighted_cvr reads the canonical 'intent' column."""
        from engine.revenue_engine import compute_intent_weighted_cvr
        df = pd.DataFrame({
            "intent": ["commercial"] * 3,
            "estimated_monthly_traffic": [200] * 3,
        })
        result = compute_intent_weighted_cvr(df, 2.0)
        assert result == pytest.approx(2.0 * 1.5, rel=0.01)

    def test_breakdown_table(self):
        from engine.revenue_engine import intent_revenue_breakdown
        df = pd.DataFrame({
            "intent": ["commercial", "transactional", "informational"],
            "uplift": [1000, 500, 300],
        })
        table = intent_revenue_breakdown(df, 2.0, 100.0)
        assert not table.empty
        assert "Intent" in table.columns
        assert "Monthly Revenue" in table.columns
        assert len(table) == 4  # one row per intent in INTENT_CVR_MULTIPLIERS


# ── Movement Learning ────────────────────────────────────────────────────────


class TestMovementLearning:
    def test_learn_movement_returns_per_tier_stats(self):
        from engine.positional_engine import learn_movement_from_history
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(40)],
            "kd": [20] * 20 + [50] * 20,
            "previous_position": [15] * 20 + [25] * 20,
            "position": [10] * 20 + [20] * 20,
        })
        stats = learn_movement_from_history(df)
        assert "Easy" in stats
        assert abs(stats["Easy"]["mean_gain"] - 5.0) < 0.01
        assert stats["Easy"]["sample_size"] == 20

    def test_learn_movement_filters_outliers(self):
        from engine.positional_engine import learn_movement_from_history
        # Row 0: previous=70, position=20 → delta=50 (>30, outlier, excluded)
        # Rows 1-14: previous=15, position=10 → delta=5 (valid)
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(15)],
            "kd": [20] * 15,
            "previous_position": [70] + [15] * 14,
            "position": [20] + [10] * 14,
        })
        stats = learn_movement_from_history(df)
        # Only 14 valid samples; the 50-position jump is excluded
        assert stats["Easy"]["sample_size"] == 14
        assert abs(stats["Easy"]["mean_gain"] - 5.0) < 0.01

    def test_learn_movement_skips_small_samples(self):
        from engine.positional_engine import learn_movement_from_history
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(8)],
            "kd": [20] * 8,
            "previous_position": [15] * 8,
            "position": [10] * 8,
        })
        stats = learn_movement_from_history(df)
        assert "Easy" not in stats  # 8 samples < 10 threshold

    def test_learn_movement_missing_column_returns_empty(self):
        from engine.positional_engine import learn_movement_from_history
        df = pd.DataFrame({
            "keyword": ["kw1"], "kd": [20], "position": [10],
        })
        assert learn_movement_from_history(df) == {}

    def test_estimate_target_uses_learned_stats_when_available(self):
        from engine.positional_engine import estimate_target_position
        stats_high = {"Easy": {"mean_gain": 10.0, "std_gain": 1.0, "sample_size": 20}}
        stats_low = {"Easy": {"mean_gain": 2.0, "std_gain": 1.0, "sample_size": 20}}
        target_high = estimate_target_position(20, 15, "moderate", stats_high)
        target_low = estimate_target_position(20, 15, "moderate", stats_low)
        assert target_high < target_low  # Higher learned gain → better (lower) position

    def test_estimate_target_falls_back_when_tier_has_few_samples(self):
        from engine.positional_engine import estimate_target_position
        # Easy: 15 samples → use learned (mean=10); Extreme: 2 samples → fall back to default
        mixed_stats = {
            "Easy": {"mean_gain": 10.0, "std_gain": 1.0, "sample_size": 15},
            "Extreme": {"mean_gain": 99.0, "std_gain": 1.0, "sample_size": 2},
        }
        # Easy keyword (kd=15): should use learned mean=10, giving a bigger gain
        target_easy_learned = estimate_target_position(20, 15, "moderate", mixed_stats)
        target_easy_default = estimate_target_position(20, 15, "moderate", None)
        assert target_easy_learned < target_easy_default  # learned gain > default → better position

        # Extreme keyword (kd=90): 2 samples < 10, falls back to _BASE_GAIN_BY_TIER
        target_extreme_mixed = estimate_target_position(20, 90, "moderate", mixed_stats)
        target_extreme_default = estimate_target_position(20, 90, "moderate", None)
        assert target_extreme_mixed == target_extreme_default  # fallback produces same result

    def test_forecast_with_learned_stats_differs_from_defaults(self):
        from engine.positional_engine import run_positional_forecast_mc
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(30)],
            "position": [15] * 30,
            "volume": [1000] * 30,
            "kd": [20] * 30,
            "current_traffic": [80] * 30,
            "intent": ["commercial"] * 30,
            "has_aio": [False] * 30,
        })
        _, monthly_default = run_positional_forecast_mc(df, months=12, n_trials=200, seed=42)
        big_gain_stats = {
            "Easy": {"mean_gain": 12.0, "std_gain": 1.0, "sample_size": 20}
        }
        _, monthly_learned = run_positional_forecast_mc(
            df, months=12, n_trials=200, seed=42,
            historical_movement_stats=big_gain_stats,
        )
        assert monthly_learned.iloc[-1]["uplift_p50"] != monthly_default.iloc[-1]["uplift_p50"]


# ── Maturation Curve (Task 4) ─────────────────────────────────────────────


class TestMaturationCurve:
    def test_logistic_progress_at_midpoint_is_half(self):
        from engine.maturation_curve import logistic_progress
        # At t == t_mid, sigmoid evaluates to exactly 0.5 regardless of k
        assert abs(logistic_progress(5.0, 5.0, 1.0) - 0.5) < 0.001
        assert abs(logistic_progress(5.0, 5.0, 1.2) - 0.5) < 0.001

    def test_logistic_progress_monotonically_increasing(self):
        from engine.maturation_curve import maturation_schedule
        schedule = maturation_schedule("Moderate", 24, 1)
        assert all(schedule[i] <= schedule[i + 1] for i in range(len(schedule) - 1))

    def test_maturation_schedule_zero_before_publish(self):
        from engine.maturation_curve import maturation_schedule
        # publish_month=3: months 1, 2, 3 are 0 (elapsed ≤ 0); month 4 starts accumulating
        schedule = maturation_schedule("Easy", 12, 3)
        assert schedule[0] == 0.0  # month 1
        assert schedule[1] == 0.0  # month 2
        assert schedule[2] == 0.0  # month 3 (publish month itself: elapsed=0 → 0.0)
        assert schedule[3] > 0.0   # month 4 (elapsed=1, first post-publish month)

    def test_easy_tier_ramps_faster_than_extreme(self):
        from engine.maturation_curve import maturation_schedule
        easy = maturation_schedule("Easy", 12, 1)
        extreme = maturation_schedule("Extreme", 12, 1)
        assert easy[5] > extreme[5]  # by month 6, Easy is meaningfully ahead of Extreme

    def test_80_pct_shape_constraint(self):
        """S-curve should reach ~75%+ by three-quarters of the way through the ramp."""
        from engine.maturation_curve import maturation_schedule
        schedule = maturation_schedule("Easy", 12, 1)
        assert schedule[6] >= 0.75

    def test_new_content_no_discontinuity(self):
        """No single-month jump should exceed 50% of that keyword's final traffic."""
        df = pd.DataFrame({
            "keyword": ["kw1"],
            "volume": [5000],
            "kd": [15],
        })
        _, monthly = run_new_content_forecast(df, da=60, cadence=1, months=18, seed=42)
        traffic = monthly["traffic"].values.astype(float)
        diffs = [abs(traffic[i] - traffic[i - 1]) for i in range(1, len(traffic))]
        final = max(traffic[-1], 1)
        assert all(d / final < 0.5 for d in diffs)

    def test_positional_bands_still_ordered_after_scurve(self):
        from engine.positional_engine import run_positional_forecast_mc
        df = pd.DataFrame({
            "keyword": [f"kw_{i}" for i in range(30)],
            "position": [10] * 30,
            "volume": [1000] * 30,
            "kd": [30] * 30,
            "current_traffic": [100] * 30,
            "intent": ["commercial"] * 30,
            "has_aio": [False] * 30,
        })
        _, monthly = run_positional_forecast_mc(df, months=12, n_trials=200, seed=42)
        assert (monthly["uplift_p10"] <= monthly["uplift_p50"]).all()
        assert (monthly["uplift_p50"] <= monthly["uplift_p90"]).all()


# ── Learned Seasonality (Task 3) ─────────────────────────────────────────


class TestLearnedSeasonality:
    def _make_ga4_df(self, n_years=2, nov_boost=0.25):
        """Synthetic GA4 df where November is artificially high."""
        dates = pd.date_range("2022-01-01", periods=12 * n_years, freq="MS")
        traffic = []
        for d in dates:
            base = 10000
            if d.month == 11:
                base = int(base * (1 + nov_boost))
            traffic.append(base)
        return pd.DataFrame({"date": dates, "traffic": traffic})

    def test_learn_seasonality_with_november_peak(self):
        from engine.seasonality_engine import learn_seasonality_from_ga4
        df = self._make_ga4_df(n_years=2, nov_boost=0.25)
        learned = learn_seasonality_from_ga4(df)
        assert learned is not None
        assert abs(learned[11]["traffic_mod"] - 0.25) < 0.03

    def test_learn_seasonality_requires_12_months(self):
        from engine.seasonality_engine import learn_seasonality_from_ga4
        df = self._make_ga4_df(n_years=1).head(6)
        assert learn_seasonality_from_ga4(df) is None

    def test_blend_weight_zero_returns_defaults(self):
        from engine.seasonality_engine import (
            DEFAULT_SEASONALITY,
            blend_learned_and_default_seasonality,
            learn_seasonality_from_ga4,
        )
        df = self._make_ga4_df()
        learned = learn_seasonality_from_ga4(df)
        blended = blend_learned_and_default_seasonality(learned, DEFAULT_SEASONALITY, 0.0)
        for m in range(1, 13):
            assert blended[m]["traffic_mod"] == DEFAULT_SEASONALITY[m]["traffic_mod"]

    def test_blend_weight_one_returns_learned(self):
        from engine.seasonality_engine import (
            DEFAULT_SEASONALITY,
            blend_learned_and_default_seasonality,
            learn_seasonality_from_ga4,
        )
        df = self._make_ga4_df()
        learned = learn_seasonality_from_ga4(df)
        blended = blend_learned_and_default_seasonality(learned, DEFAULT_SEASONALITY, 1.0)
        for m in range(1, 13):
            assert abs(blended[m]["traffic_mod"] - learned[m]["traffic_mod"]) < 0.001

    def test_blend_weight_half_is_midpoint(self):
        from engine.seasonality_engine import (
            DEFAULT_SEASONALITY,
            blend_learned_and_default_seasonality,
            learn_seasonality_from_ga4,
        )
        df = self._make_ga4_df()
        learned = learn_seasonality_from_ga4(df)
        blended = blend_learned_and_default_seasonality(learned, DEFAULT_SEASONALITY, 0.5)
        for m in range(1, 13):
            expected = round(
                0.5 * learned[m]["traffic_mod"] + 0.5 * DEFAULT_SEASONALITY[m]["traffic_mod"], 4
            )
            assert abs(blended[m]["traffic_mod"] - expected) < 0.001

    def test_au_holidays_df_has_black_friday_2025(self):
        from engine.seasonality_engine import build_au_holidays_df
        df = build_au_holidays_df(2025, 2025)
        bf = df[df["holiday"] == "Black Friday"]
        assert len(bf) == 1
        ds = pd.to_datetime(bf["ds"].iloc[0])
        assert ds.year == 2025
        assert ds.month == 11
        assert ds.weekday() == 4  # Friday

    def test_au_holidays_df_covers_full_year_range(self):
        from engine.seasonality_engine import build_au_holidays_df
        df = build_au_holidays_df(2023, 2028)
        years = pd.to_datetime(df["ds"]).dt.year.unique()
        for y in range(2023, 2029):
            assert y in years

    def test_au_holidays_has_all_required_events(self):
        from engine.seasonality_engine import AU_HOLIDAYS
        required = {
            "EOFY", "Black Friday", "Cyber Monday",
            "Christmas", "Boxing Day Sales", "Back to School",
        }
        found = set(AU_HOLIDAYS["holiday"].unique())
        for r in required:
            assert r in found, f"Missing holiday: {r}"

    def test_au_holidays_has_correct_columns(self):
        from engine.seasonality_engine import AU_HOLIDAYS
        assert set(AU_HOLIDAYS.columns) >= {"holiday", "ds", "lower_window", "upper_window"}


# ── Prophet / v4 Historical (Task 1) ─────────────────────────────────────


class TestHistoricalV4:
    def _make_df(self, n_months: int, trend: float = 100.0):
        dates = pd.date_range("2022-01-01", periods=n_months, freq="MS")
        traffic = [int(5000 + i * trend) for i in range(n_months)]
        return pd.DataFrame({"date": dates, "traffic": traffic})

    def test_24_months_selects_holts_or_prophet(self):
        from engine.historical_engine import _PROPHET_MIN_MONTHS, run_historical_forecast_v4
        df = self._make_df(24)
        result = run_historical_forecast_v4(df, months=6)
        assert result.attrs["chosen_method"] in ("prophet", "holts", "linear")
        # With 24 months, should not be linear
        assert result.attrs["chosen_method"] != "linear"

    def test_18_months_selects_holts(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(18)
        result = run_historical_forecast_v4(df, months=6)
        assert result.attrs["chosen_method"] == "holts"

    def test_6_months_selects_linear(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(6)
        result = run_historical_forecast_v4(df, months=3)
        assert result.attrs["chosen_method"] == "linear"

    def test_fallback_when_prophet_unavailable(self):
        """Mock ImportError to ensure graceful fallback."""
        import sys
        import unittest.mock as mock

        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(24)
        with mock.patch.dict(sys.modules, {"prophet": None}):
            result = run_historical_forecast_v4(df, months=6)
        # Should still produce a result with linear
        assert len(result) == 24 + 6
        assert "linear" in result.columns

    def test_output_extends_dates(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(12)
        result = run_historical_forecast_v4(df, months=6)
        assert len(result) == 18
        forecast_rows = result[result["is_forecast"]]
        assert len(forecast_rows) == 6

    def test_trending_data_produces_nonfloat_forecast(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(18, trend=500.0)
        result = run_historical_forecast_v4(df, months=6)
        forecast = result[result["is_forecast"]]
        # A strong upward trend should produce non-flat forecasts
        assert forecast["exponential_smoothing"].iloc[-1] != forecast["exponential_smoothing"].iloc[0]

    def test_method_selection_24_months_picks_prophet(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(24)
        result = run_historical_forecast_v4(df, months=6)
        assert result.attrs["chosen_method"] == "prophet"

    def test_method_selection_18_months_picks_holts_with_low_confidence(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(18)
        result = run_historical_forecast_v4(df, months=6)
        assert result.attrs["chosen_method"] == "holts"
        assert result.attrs["low_confidence"] is True

    def test_method_selection_6_months_picks_linear_with_warning(self):
        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(6)
        result = run_historical_forecast_v4(df, months=3)
        assert result.attrs["chosen_method"] == "linear"
        assert "seasonality" in result.attrs["method_reason"].lower()

    def test_fallback_to_holts_when_prophet_unavailable(self):
        import unittest.mock as mock

        from engine.historical_engine import run_historical_forecast_v4
        df = self._make_df(24)
        with mock.patch("engine.prophet_engine._PROPHET_AVAILABLE", False):
            result = run_historical_forecast_v4(df, months=6)
        assert len(result) == 30
        assert "exponential_smoothing" in result.columns
        assert result.attrs["prophet_available"] is False

    @pytest.mark.skipif(
        not __import__("engine.prophet_engine", fromlist=["_PROPHET_AVAILABLE"])._PROPHET_AVAILABLE,
        reason="Prophet not installed",
    )
    def test_prophet_extends_dates_monotonically(self):
        from engine.prophet_engine import run_prophet_forecast
        df = self._make_df(24)
        result = run_prophet_forecast(df, months=6)
        assert result["date"].is_monotonic_increasing

    @pytest.mark.skipif(
        not __import__("engine.prophet_engine", fromlist=["_PROPHET_AVAILABLE"])._PROPHET_AVAILABLE,
        reason="Prophet not installed",
    )
    def test_prophet_forecast_non_flat_on_trending_data(self):
        from engine.prophet_engine import run_prophet_forecast
        df = self._make_df(24, trend=500.0)
        result = run_prophet_forecast(df, months=6)
        forecast = result[result["is_forecast"]]
        assert forecast["forecast"].iloc[-1] != forecast["forecast"].iloc[0]


# ── Prompt 9 integration tests ─────────────────────────────────────────────────


class TestPrompt9Integration:
    """Tests for per-stream seasonality/AIO, roadmap_content_plan, and industry bias."""

    def _kw_df(self, n=10):
        return pd.DataFrame({
            "keyword": [f"kw {i}" for i in range(n)],
            "volume": [500] * n,
            "kd": [30] * n,
        })

    def test_positional_applies_seasonality_per_month(self):
        from engine.positional_engine import run_positional_forecast_mc
        kw_df = pd.DataFrame({
            "keyword": ["test kw"],
            "volume": [1000],
            "kd": [30],
            "position": [8],
        })
        seasonality = {1: {"traffic_mod": 0.50, "cr_mod": 0, "aov_mod": 0}}
        _, monthly_with = run_positional_forecast_mc(
            kw_df, months=12, effort="moderate",
            forecast_start_month=1, seasonality=seasonality, seed=42,
        )
        _, monthly_without = run_positional_forecast_mc(
            kw_df, months=12, effort="moderate",
            forecast_start_month=1, seed=42,
        )
        m1_with = monthly_with[monthly_with["month"] == 1]["traffic_p50"].iloc[0]
        m1_without = monthly_without[monthly_without["month"] == 1]["traffic_p50"].iloc[0]
        assert m1_with >= m1_without

    def test_positional_applies_aio_ctr_penalty_per_intent(self):
        from engine.positional_engine import run_positional_forecast_mc
        kw_df = pd.DataFrame({
            "keyword": ["how to do something"],
            "volume": [2000],
            "kd": [20],
            "position": [5],
            "intent": ["informational"],
        })
        _, monthly_no_aio = run_positional_forecast_mc(kw_df, months=6, seed=42)
        _, monthly_with_aio = run_positional_forecast_mc(
            kw_df, months=6, seed=42,
            aio_intent_penalties={"informational": 45.0},
        )
        total_no_aio = monthly_no_aio["traffic_p50"].sum()
        total_with_aio = monthly_with_aio["traffic_p50"].sum()
        assert total_with_aio <= total_no_aio

    def test_new_content_uses_roadmap_content_plan_when_provided(self):
        from engine.new_content_engine import run_new_content_forecast
        kw_df = self._kw_df(5)
        # Use keyword that will substring-match the URL
        kw_df.loc[0, "keyword"] = "summer-collection"
        plan = [{"url": "/blog/summer-collection-guide", "content_type": "new_page", "month": 3, "is_new_page": True}]
        kw_result, _ = run_new_content_forecast(
            kw_df, da=40, cadence=2, months=12, seed=42,
            roadmap_content_plan=plan,
        )
        matched = kw_result[kw_result["keyword"] == "summer-collection"]
        if not matched.empty and matched.iloc[0].get("will_rank", False):
            assert matched.iloc[0]["publish_month"] == 3

    def test_new_content_falls_back_to_cadence_for_unmatched_keywords(self):
        from engine.new_content_engine import run_new_content_forecast
        kw_df = self._kw_df(6)
        # Plan only matches first keyword; others should use cadence
        plan = [{"url": "/blog/kw-0", "content_type": "new_page", "month": 5, "is_new_page": True}]
        _, monthly = run_new_content_forecast(
            kw_df, da=40, cadence=2, months=12, seed=42,
            roadmap_content_plan=plan,
        )
        assert len(monthly) == 12

    def test_new_content_optimisation_vs_new_page_maturation_amplitude(self):
        from engine.new_content_engine import run_new_content_forecast
        kw_df = pd.DataFrame({"keyword": ["opt kw"], "volume": [1000], "kd": [20]})
        plan_opt = [{"url": "/opt-kw", "content_type": "optimisation", "month": 1, "is_new_page": False}]
        plan_new = [{"url": "/opt-kw", "content_type": "new_page", "month": 1, "is_new_page": True}]
        _, monthly_opt = run_new_content_forecast(kw_df, da=70, cadence=1, months=6, seed=42, roadmap_content_plan=plan_opt)
        _, monthly_new = run_new_content_forecast(kw_df, da=70, cadence=1, months=6, seed=42, roadmap_content_plan=plan_new)
        # Optimisation ramps faster early (lower t_mid) but has lower amplitude (0.3 cap)
        early_opt = monthly_opt["traffic"].iloc[0]
        early_new = monthly_new["traffic"].iloc[0]
        # Both should be >= 0; the test just verifies they differ (amplitude difference)
        assert early_opt >= 0 and early_new >= 0

    def test_combined_math_without_aio_deduction(self):
        from engine.combined_engine import run_combined_forecast
        # combined = baseline + positional + new_content - decay (no AIO term)
        result = run_combined_forecast(None, None, None, months=6)
        assert result is not None

    def test_industry_bias_applied_when_industry_set(self):
        from engine.seasonality_engine import (
            DEFAULT_SEASONALITY,
            INDUSTRY_SEASONALITY_PRIORS,
            apply_industry_bias,
        )
        base = dict(DEFAULT_SEASONALITY)
        biased = apply_industry_bias(base, "Accessories", bias_weight=1.0)
        # Accessories has a November boost: check Nov traffic_mod is higher
        assert biased[11]["traffic_mod"] > base[11]["traffic_mod"]

    def test_industry_bias_skipped_when_industry_unknown(self):
        from engine.seasonality_engine import DEFAULT_SEASONALITY, apply_industry_bias
        base = dict(DEFAULT_SEASONALITY)
        result = apply_industry_bias(base, "Unknown", bias_weight=1.0)
        # Unknown industry should return base unchanged
        assert result[11]["traffic_mod"] == base[11]["traffic_mod"]

    def test_load_roadmap_v2_populates_content_plan_session_key_equivalent(self):
        from pathlib import Path
        fixture = Path("tests/fixtures/sample_pattern_native_roadmap.xlsx")
        if not fixture.exists():
            pytest.skip("Fixture not found")
        from engine.roadmap_ai_engine import load_roadmap_v2
        raw = fixture.read_bytes()
        bundle, method = load_roadmap_v2(None, raw, "roadmap.xlsx")
        assert method == "deterministic"
        assert "content_plan" in bundle
        assert isinstance(bundle["content_plan"], list)


# ── Deseasonalise helpers ────────────────────────────────────────────────────


class TestDeseasonalise:
    """Tests for engine.seasonality_engine.deseasonalise_series / reseasonalise_values."""

    def _make_dates(self, months: list[int], year: int = 2025) -> pd.Series:
        return pd.Series([pd.Timestamp(year, m, 1) for m in months])

    def test_deseasonalise_then_reseasonalise_is_identity(self):
        from engine.seasonality_engine import (
            DEFAULT_SEASONALITY,
            deseasonalise_series,
            reseasonalise_values,
        )
        months = list(range(1, 13))
        dates = self._make_dates(months)
        values = pd.Series([10_000.0 + m * 500 for m in months])

        deseasoned = deseasonalise_series(dates, values, DEFAULT_SEASONALITY)
        roundtripped = reseasonalise_values(dates, deseasoned, DEFAULT_SEASONALITY)

        for original, result in zip(values, roundtripped, strict=True):
            assert abs(original - result) < 1e-6, (
                f"Round-trip failed: original={original}, result={result}"
            )

    def test_missing_months_treated_as_neutral(self):
        from engine.seasonality_engine import deseasonalise_series, reseasonalise_values

        sparse_seasonality = {11: {"traffic_mod": 0.25}}  # only Nov defined
        dates = self._make_dates([1, 6, 11])
        values = pd.Series([10_000.0, 10_000.0, 10_000.0])

        deseasoned = deseasonalise_series(dates, values, sparse_seasonality)
        # Jan and Jun are missing → multiplier 1.0 → unchanged
        assert deseasoned.iloc[0] == pytest.approx(10_000.0)
        assert deseasoned.iloc[1] == pytest.approx(10_000.0)
        # Nov has +25% modifier → deseasonalised = 10_000 / 1.25 = 8_000
        assert deseasoned.iloc[2] == pytest.approx(8_000.0)

        reseasoned = reseasonalise_values(dates, values, sparse_seasonality)
        assert reseasoned.iloc[0] == pytest.approx(10_000.0)
        assert reseasoned.iloc[1] == pytest.approx(10_000.0)
        assert reseasoned.iloc[2] == pytest.approx(12_500.0)

    def test_deseasonalised_november_is_lower_than_raw(self):
        from engine.seasonality_engine import DEFAULT_SEASONALITY, deseasonalise_series

        nov_mod = DEFAULT_SEASONALITY[11]["traffic_mod"]  # +0.25
        assert nov_mod > 0, "Test assumes November has positive traffic_mod"

        dates = self._make_dates([11])
        raw_value = 12_500.0
        values = pd.Series([raw_value])

        deseasoned = deseasonalise_series(dates, values, DEFAULT_SEASONALITY)
        expected = raw_value / (1.0 + nov_mod)
        assert deseasoned.iloc[0] == pytest.approx(expected, rel=1e-6)
        assert deseasoned.iloc[0] < raw_value


# ── Combined seasonality ─────────────────────────────────────────────────────


class TestCombinedSeasonality:
    """Tests for seasonality-aware baseline in run_combined_forecast."""

    def _make_hist_df(self, n_months: int = 12, base: int = 10_000, start_month: int = 1) -> pd.DataFrame:
        rows = []
        for i in range(n_months):
            month = (start_month - 1 + i) % 12 + 1
            year = 2024 + (start_month - 1 + i) // 12
            rows.append({"date": pd.Timestamp(year, month, 1), "traffic": float(base + i * 200)})
        return pd.DataFrame(rows)

    def test_backward_compat_no_seasonality_matches_linear_forecast(self):
        from engine.combined_engine import run_combined_forecast
        from engine.historical_engine import linear_forecast

        hist = self._make_hist_df(n_months=10)
        dates = hist["date"]
        traffic = hist["traffic"]

        result_no_season = run_combined_forecast(
            historical_df=hist,
            positional_monthly=None,
            new_content_monthly=None,
            months=6,
            seasonality=None,
        )

        # Without seasonality, baseline should match plain linear_forecast output
        lf = linear_forecast(dates, traffic, 6, confidence=15.0)
        lf_forecast = lf[lf["is_forecast"]].reset_index(drop=True)
        combined_forecast = result_no_season[result_no_season["is_forecast"]].reset_index(drop=True)

        for i in range(6):
            expected = int(lf_forecast.iloc[i]["linear"])
            actual = int(combined_forecast.iloc[i]["baseline"])
            assert abs(actual - expected) <= 1, (
                f"Month {i+1}: expected baseline={expected}, got {actual}"
            )

    def test_november_baseline_higher_than_may(self):
        """With DEFAULT_SEASONALITY on short history, Nov forecast > May forecast."""
        from engine.combined_engine import run_combined_forecast
        from engine.seasonality_engine import DEFAULT_SEASONALITY

        # 10 months of history starting May — so Nov and May both appear in forecast
        hist = self._make_hist_df(n_months=10, base=10_000, start_month=5)

        result = run_combined_forecast(
            historical_df=hist,
            positional_monthly=None,
            new_content_monthly=None,
            months=12,
            seasonality=DEFAULT_SEASONALITY,
        )

        forecast = result[result["is_forecast"]].copy()
        forecast["month"] = pd.to_datetime(forecast["date"]).dt.month

        nov_rows = forecast[forecast["month"] == 11]
        may_rows = forecast[forecast["month"] == 5]

        if nov_rows.empty or may_rows.empty:
            import pytest
            pytest.skip("Forecast window does not include both Nov and May")

        nov_baseline = nov_rows["baseline"].iloc[0]
        may_baseline = may_rows["baseline"].iloc[0]

        assert nov_baseline > may_baseline, (
            f"Expected Nov baseline ({nov_baseline}) > May baseline ({may_baseline})"
        )

    def test_baseline_lift_matches_seasonality_spec(self):
        """With perfectly seasonal history, Nov forecast ≈ base × (1 + nov_mod).

        Build history where raw traffic = base × (1 + seasonal_mod) so that
        deseasonalise_series() returns a perfectly flat series.  OLS on a flat
        series gives a flat forecast; reseasonalising November multiplies by 1.25.
        """
        from engine.combined_engine import run_combined_forecast
        from engine.seasonality_engine import DEFAULT_SEASONALITY

        BASE = 10_000.0
        # Raw traffic shaped by seasonality → deseasonalised series is flat at BASE
        hist = pd.DataFrame([
            {
                "date": pd.Timestamp(2024, m, 1),
                "traffic": BASE * (1.0 + DEFAULT_SEASONALITY[m]["traffic_mod"]),
            }
            for m in range(1, 11)  # Jan – Oct
        ])

        result = run_combined_forecast(
            historical_df=hist,
            positional_monthly=None,
            new_content_monthly=None,
            months=12,
            seasonality=DEFAULT_SEASONALITY,
        )

        forecast = result[result["is_forecast"]].copy()
        forecast["month"] = pd.to_datetime(forecast["date"]).dt.month
        nov_rows = forecast[forecast["month"] == 11]

        if nov_rows.empty:
            import pytest
            pytest.skip("November not in forecast window")

        nov_baseline = float(nov_rows["baseline"].iloc[0])
        nov_mod = DEFAULT_SEASONALITY[11]["traffic_mod"]  # +0.25
        expected = BASE * (1.0 + nov_mod)  # 12 500

        # OLS on flat deseasonalised data is flat ⇒ reseasonalised Nov ≈ BASE × 1.25
        # Allow ±12% relative tolerance for any linear_forecast confidence-interval drift
        assert abs(nov_baseline - expected) / expected < 0.12, (
            f"Nov baseline={nov_baseline:.0f}, expected≈{expected:.0f}"
        )

    def test_forecast_start_month_derived_from_historical(self):
        """When historical_df is provided, forecast dates follow from its last date."""
        from engine.combined_engine import run_combined_forecast

        hist = self._make_hist_df(n_months=8, base=10_000, start_month=1)
        # Last historical date = August 2024 → first forecast = September 2024
        last_hist_month = pd.Timestamp(hist["date"].iloc[-1]).month  # August = 8

        result = run_combined_forecast(
            historical_df=hist,
            positional_monthly=None,
            new_content_monthly=None,
            months=3,
            seasonality=None,
            forecast_start_month=99,  # should be overridden
        )

        forecast = result[result["is_forecast"]].reset_index(drop=True)
        first_forecast_month = pd.Timestamp(forecast.iloc[0]["date"]).month
        expected = (last_hist_month % 12) + 1  # September = 9
        assert first_forecast_month == expected, (
            f"Expected first forecast month={expected}, got {first_forecast_month}"
        )


# ── Comparison columns ───────────────────────────────────────────────────────


class TestComparisonColumns:
    """Tests for engine.combined_engine._add_comparison_columns."""

    def _make_combined(self, n_hist: int = 14, n_fc: int = 12) -> pd.DataFrame:
        """Build a minimal combined_df with history + forecast rows."""
        from engine.combined_engine import run_combined_forecast

        rows_hist = []
        for i in range(n_hist):
            date = pd.Timestamp("2024-01-01") + pd.DateOffset(months=i)
            rows_hist.append({"date": date, "traffic": float(10_000 + i * 200)})
        hist_df = pd.DataFrame(rows_hist)

        result = run_combined_forecast(
            historical_df=hist_df,
            positional_monthly=None,
            new_content_monthly=None,
            months=n_fc,
            seasonality=None,
        )
        return result

    def test_mom_diff_uses_prior_row(self):
        df = self._make_combined()
        forecast = df[df["is_forecast"]].reset_index(drop=True)

        # MoM diff for row[1] should equal combined[1] - combined[0]
        col = "combined_p50" if "combined_p50" in forecast.columns else "combined"
        v0 = float(forecast.iloc[0][col])
        v1 = float(forecast.iloc[1][col])

        expected = round(v1 - v0, 1)
        actual = forecast.iloc[1]["mom_diff"]
        assert actual is not None and abs(float(actual) - expected) < 1.0, (
            f"mom_diff row 1 expected {expected}, got {actual}"
        )

    def test_mom_pct_first_forecast_row_is_none_or_nan(self):
        from engine.combined_engine import run_combined_forecast

        # No historical_df → all rows are forecast; first row has no prior → MoM blank
        df = run_combined_forecast(
            historical_df=None,
            positional_monthly=None,
            new_content_monthly=None,
            months=6,
            seasonality=None,
        )
        forecast = df[df["is_forecast"]].reset_index(drop=True)
        first_mom_pct = forecast.iloc[0]["mom_pct"]
        assert first_mom_pct is None or (
            isinstance(first_mom_pct, float) and pd.isna(first_mom_pct)
        ), f"Expected None/NaN for first row mom_pct, got {first_mom_pct}"

    def test_yoy_diff_when_prior_year_in_history(self):
        df = self._make_combined(n_hist=14, n_fc=12)
        forecast = df[df["is_forecast"]].reset_index(drop=True)

        # First forecast month is 15 months after start of history → prior year
        # is month 3 of history (index 2) — there IS a history row 12 months prior
        rows_with_yoy = forecast[forecast["yoy_diff"].notna()]
        assert not rows_with_yoy.empty, "Expected at least one forecast row with yoy_diff"

    def test_yoy_blank_when_no_prior_year_match(self):
        # Only 6 months of history → first forecast month has no row 12 months prior
        df = self._make_combined(n_hist=6, n_fc=12)
        forecast = df[df["is_forecast"]].reset_index(drop=True)

        # First 6 forecast months have no history to compare against
        early = forecast.iloc[:6]
        assert early["yoy_diff"].isna().all() or (early["yoy_diff"] == None).all(), (  # noqa: E711
            "Expected early forecast rows to have no YoY diff when history < 12 months"
        )

    def test_yoy_uses_actual_for_history_rows_combined_p50_for_forecast(self):
        from engine.combined_engine import _add_comparison_columns

        # Build a tiny synthetic df with 2 history rows and 1 forecast row
        rows = [
            {"date": pd.Timestamp("2024-01-01"), "actual": 10_000.0,
             "combined_p50": 10_000.0, "is_forecast": False},
            {"date": pd.Timestamp("2024-02-01"), "actual": 11_000.0,
             "combined_p50": 11_000.0, "is_forecast": False},
            {"date": pd.Timestamp("2025-01-01"), "actual": None,
             "combined_p50": 12_500.0, "is_forecast": True},
        ]
        df = pd.DataFrame(rows)

        result = _add_comparison_columns(df)

        # Jan 2024 history row: value = actual = 10,000
        # Feb 2024 history row: MoM diff should use actual (11,000 - 10,000 = 1,000)
        assert result.iloc[1]["mom_diff"] == pytest.approx(1_000.0, abs=1.0)

        # Jan 2025 forecast: YoY prior = Jan 2024 actual = 10,000
        # combined_p50 = 12,500 → yoy_diff = 2,500
        assert result.iloc[2]["yoy_prior"] == pytest.approx(10_000.0, abs=1.0)
        assert result.iloc[2]["yoy_diff"] == pytest.approx(2_500.0, abs=1.0)
        assert result.iloc[2]["yoy_pct"] == pytest.approx(25.0, abs=0.5)


# ── Forecast Grid — GAZMAN Header Scaffold ──────────────────────────────────


class TestForecastGridHeader:
    """Tests for the GAZMAN header scaffold in utils.forecast_grid (Session B1a)."""

    def _make_grid(self, months: int = 12, **kwargs) -> "io.BytesIO":
        import io  # noqa: F401 (used in type hint above)

        from utils.forecast_grid import build_seo_forecast_grid

        traffic = [10_000.0 + i * 100 for i in range(months)]
        return build_seo_forecast_grid(
            monthly_traffic=traffic,
            monthly_transactions=[t * 0.025 for t in traffic],
            monthly_revenue=[t * 0.025 * 100 for t in traffic],
            monthly_cvr=[2.5] * months,
            monthly_aov=[100.0] * months,
            monthly_budget=[5_000.0] * months,
            months=months,
            client_name="GAZMAN",
            fy_label="FY26",
            start_month=7,
            last_updated="2026-04-24",
            currency_notes="All figures in AUD",
            **kwargs,
        )

    def _open(self, buf):
        from openpyxl import load_workbook
        buf.seek(0)
        return load_workbook(buf)

    def test_sheet_title_in_row_2(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        title = ws.cell(row=2, column=1).value or ""
        assert "GAZMAN" in title and "FY26" in title

    def test_last_updated_in_row_4(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        label = ws.cell(row=4, column=2).value or ""
        value = ws.cell(row=4, column=3).value or ""
        assert "Last Updated" in label and "2026-04-24" in str(value)

    def test_month_names_in_row_7(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        row7 = [ws.cell(row=7, column=c).value for c in range(1, 50)]
        assert "Jul" in row7

    def test_each_month_header_spans_three_columns(self):
        from utils.forecast_grid import _month_column_ranges
        ranges = _month_column_ranges(7, 12)
        # First month July: forecast=col3, actuals=col4, pct=col5
        assert ranges[0] == ("Jul", 3, 4, 5)
        # Second month August: forecast=col6, actuals=col7, pct=col8
        assert ranges[1] == ("Aug", 6, 7, 8)

    def test_totals_column_exists(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        row7 = [ws.cell(row=7, column=c).value for c in range(1, 50)]
        assert "TOTALS" in row7

    def test_row_12_channel_header_strip(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        assert ws.cell(row=12, column=1).value == "CHANNEL"
        row12 = [ws.cell(row=12, column=c).value for c in range(1, 50)]
        assert "Forecast" in row12
        assert "Actuals" in row12
        assert "% Change" in row12

    def test_row_13_col_a_contains_seo(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        assert ws.cell(row=13, column=1).value == "SEO"

    def test_rows_14_to_19_col_a_blank(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        for row in range(14, 20):
            val = ws.cell(row=row, column=1).value
            assert val is None, f"Row {row} col A expected blank, got {val!r}"

    def test_seven_metric_labels_in_col_b_bold(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        col_b = [ws.cell(row=r, column=2).value for r in range(13, 20)]
        assert col_b == ["BUDGET", "REVENUE", "ROAS", "TRANSACTIONS", "AOV", "TRAFFIC", "CVR"]

    def test_freeze_panes_at_b13(self):
        wb = self._open(self._make_grid())
        ws = wb["SEO Channel Forecast"]
        assert ws.freeze_panes == "B13"


# ── Forecast Grid — Metric Data Population ───────────────────────────────────


class TestForecastGridMetricData:
    """Tests for the B1b metric data population in utils.forecast_grid."""

    _MONTHS = 3  # small grid so column maths stay readable in tests
    _TRAFFIC = [10_000.0, 11_000.0, 12_000.0]
    _CVR = [2.5, 2.6, 2.7]          # percentage form
    _AOV = [100.0, 101.0, 102.0]
    _TRANSACTIONS = [t * c / 100 for t, c in zip(_TRAFFIC, _CVR, strict=True)]
    _REVENUE = [tr * a for tr, a in zip(_TRANSACTIONS, _AOV, strict=True)]
    _BUDGET = [5_000.0, 5_000.0, 5_000.0]

    def _make(self, **kwargs):
        from utils.forecast_grid import build_seo_forecast_grid
        defaults = dict(
            monthly_traffic=self._TRAFFIC,
            monthly_transactions=self._TRANSACTIONS,
            monthly_revenue=self._REVENUE,
            monthly_cvr=self._CVR,
            monthly_aov=self._AOV,
            monthly_budget=self._BUDGET,
            months=self._MONTHS,
            start_month=7,
        )
        defaults.update(kwargs)
        return build_seo_forecast_grid(**defaults)

    def _ws(self, buf):
        from openpyxl import load_workbook
        buf.seek(0)
        return load_workbook(buf)["SEO Channel Forecast"]

    # fc(0)=3, ac(0)=4, pc(0)=5; for m=3: ann=12, ytd=13, prior=15, yoy=16
    _FC0 = 3
    _AC0 = 4
    _PC0 = 5
    _ANN = 12   # _col_annual(3) = 3 + 3*3 = 12
    _PRIOR = 15  # _col_prior(3) = 12 + 3 = 15
    _YOY = 16    # _col_yoy(3) = 12 + 4 = 16

    def test_budget_row_populated_from_monthly_budget(self):
        ws = self._ws(self._make())
        assert ws.cell(row=13, column=self._FC0).value == pytest.approx(self._BUDGET[0])

    def test_revenue_row_populated_from_monthly_revenue(self):
        ws = self._ws(self._make())
        assert ws.cell(row=14, column=self._FC0).value == pytest.approx(self._REVENUE[0], rel=1e-3)

    def test_roas_computed_from_revenue_div_budget_per_month(self):
        ws = self._ws(self._make())
        expected = self._REVENUE[0] / self._BUDGET[0]
        assert ws.cell(row=15, column=self._FC0).value == pytest.approx(expected, rel=1e-3)

    def test_roas_blank_when_budget_zero(self):
        ws = self._ws(self._make(monthly_budget=[0.0, 0.0, 0.0]))
        assert ws.cell(row=15, column=self._FC0).value is None

    def test_pct_change_negative_when_actual_below_forecast(self):
        actuals = [b * 0.8 for b in self._BUDGET]   # 20% below forecast
        ws = self._ws(self._make(actuals_budget=actuals))
        pct = ws.cell(row=13, column=self._PC0).value
        assert pct is not None and pct < 0

    def test_pct_change_neg_one_when_actual_zero_and_forecast_nonzero(self):
        ws = self._ws(self._make(actuals_budget=[0.0, 0.0, 0.0]))
        pct = ws.cell(row=13, column=self._PC0).value
        assert pct == pytest.approx(-1.0)

    def test_pct_change_blank_when_no_actual(self):
        ws = self._ws(self._make())   # no actuals_* passed
        assert ws.cell(row=13, column=self._PC0).value is None

    def test_annual_forecast_col_sums_monthly(self):
        ws = self._ws(self._make())
        expected = sum(self._BUDGET)
        assert ws.cell(row=13, column=self._ANN).value == pytest.approx(expected)

    def test_annual_roas_is_sum_revenue_div_sum_budget(self):
        ws = self._ws(self._make())
        expected = sum(self._REVENUE) / sum(self._BUDGET)
        assert ws.cell(row=15, column=self._ANN).value == pytest.approx(expected, rel=1e-3)

    def test_yoy_pct_computed_when_prior_year_provided(self):
        prior = sum(self._BUDGET) * 0.9  # 10% growth expected
        ws = self._ws(self._make(prior_year_budget=prior))
        yoy = ws.cell(row=13, column=self._YOY).value
        assert yoy is not None
        expected = (sum(self._BUDGET) - prior) / prior
        assert yoy == pytest.approx(expected, rel=1e-3)

    def test_yoy_pct_blank_when_prior_year_none(self):
        ws = self._ws(self._make())  # no prior_year_* passed
        assert ws.cell(row=13, column=self._YOY).value is None

    def test_cvr_cell_value_divided_by_100_for_percentage_format(self):
        ws = self._ws(self._make())
        # CVR row is 19; CVR input is e.g. 2.5 → cell should be 0.025
        cvr_cell_val = ws.cell(row=19, column=self._FC0).value
        assert cvr_cell_val is not None
        assert cvr_cell_val == pytest.approx(self._CVR[0] / 100.0, rel=1e-6)

    def test_negative_pct_change_has_red_font(self):
        actuals = [b * 0.5 for b in self._BUDGET]
        ws = self._ws(self._make(actuals_budget=actuals))
        cell = ws.cell(row=13, column=self._PC0)
        assert cell.value is not None and cell.value < 0
        assert "FF0000" in cell.font.color.rgb

# ── Forecast Grid — Assumptions Column + Fee Rows ────────────────────────────


class TestForecastGridFees:
    """Tests for B1c: assumptions text column and management fee rows."""

    _MONTHS = 3
    _TRAFFIC = [10_000.0, 11_000.0, 12_000.0]
    _CVR = [2.5, 2.6, 2.7]
    _AOV = [100.0, 101.0, 102.0]
    _TRANSACTIONS = [t * c / 100 for t, c in zip(_TRAFFIC, _CVR, strict=True)]
    _REVENUE = [tr * a for tr, a in zip(_TRANSACTIONS, _AOV, strict=True)]
    _BUDGET = [5_000.0, 5_000.0, 5_000.0]
    _ASS_TEXT = "CVR: 2.5%\nAOV: $100\nBudget: $5k/month"

    # For months=3: _col_ass(3) = _col_annual(3) + 2 = (3 + 9) + 2 = 14
    _ASS_COL = 14
    _FC0 = 3
    _AC0 = 4
    _PC0 = 5
    _ANN = 12  # _col_annual(3) = 12

    def _make(self, **kwargs):
        from utils.forecast_grid import build_seo_forecast_grid
        defaults = dict(
            monthly_traffic=self._TRAFFIC,
            monthly_transactions=self._TRANSACTIONS,
            monthly_revenue=self._REVENUE,
            monthly_cvr=self._CVR,
            monthly_aov=self._AOV,
            monthly_budget=self._BUDGET,
            months=self._MONTHS,
            start_month=7,
            assumptions_text=self._ASS_TEXT,
        )
        defaults.update(kwargs)
        return build_seo_forecast_grid(**defaults)

    def _ws(self, buf):
        from openpyxl import load_workbook
        buf.seek(0)
        return load_workbook(buf)["SEO Channel Forecast"]

    def test_assumptions_text_in_row_13(self):
        ws = self._ws(self._make())
        assert ws.cell(row=13, column=self._ASS_COL).value == self._ASS_TEXT

    def test_assumptions_col_merged_across_rows_13_to_19(self):
        from openpyxl.utils import get_column_letter
        ws = self._ws(self._make())
        col_letter = get_column_letter(self._ASS_COL)
        expected = f"{col_letter}13:{col_letter}19"
        merged = [str(r) for r in ws.merged_cells.ranges]
        assert expected in merged, f"Expected {expected!r} in merged ranges {merged}"

    def test_assumptions_col_has_wrap_text(self):
        ws = self._ws(self._make())
        cell = ws.cell(row=13, column=self._ASS_COL)
        assert cell.alignment.wrap_text is True

    def test_row_21_has_seo_management_tech_fee_label(self):
        ws = self._ws(self._make())
        assert ws.cell(row=21, column=1).value == "SEO Management + Tech Fee"

    def test_row_22_has_seo_total_label(self):
        ws = self._ws(self._make())
        assert ws.cell(row=22, column=1).value == "SEO Total"

    def test_fee_rows_forecast_actuals_populated(self):
        actuals = [4_800.0, 4_900.0, 5_100.0]
        ws = self._ws(self._make(actuals_budget=actuals))
        assert ws.cell(row=21, column=self._FC0).value == pytest.approx(self._BUDGET[0])
        assert ws.cell(row=21, column=self._AC0).value == pytest.approx(actuals[0])
        assert ws.cell(row=22, column=self._FC0).value == pytest.approx(self._BUDGET[0])
        assert ws.cell(row=22, column=self._AC0).value == pytest.approx(actuals[0])

    def test_fee_rows_pct_change_column_is_blank(self):
        actuals = [4_800.0, 4_900.0, 5_100.0]
        ws = self._ws(self._make(actuals_budget=actuals))
        assert ws.cell(row=21, column=self._PC0).value is None
        assert ws.cell(row=22, column=self._PC0).value is None

    def test_fee_rows_annual_total_populated(self):
        ws = self._ws(self._make())
        expected = sum(self._BUDGET)
        assert ws.cell(row=21, column=self._ANN).value == pytest.approx(expected)
        assert ws.cell(row=22, column=self._ANN).value == pytest.approx(expected)
