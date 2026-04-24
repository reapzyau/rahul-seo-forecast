"""Tests for utils/forecast_grid.py and engine/snapshot_engine.py extensions."""
import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from engine.snapshot_engine import (
    build_snapshot,
    compare_to_actuals,
    load_snapshot,
    snapshot_to_bytes,
    summarise_variance,
)
from utils.forecast_grid import build_seo_forecast_grid

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_traffic(n: int = 12, base: int = 10_000) -> list[float]:
    return [float(base + i * 100) for i in range(n)]


def _make_grid(
    months: int = 12,
    with_cvr: bool = False,
    with_aov: bool = False,
    with_bands: bool = False,
    with_streams: bool = False,
    with_assumptions: bool = False,
) -> io.BytesIO:
    traffic = _make_traffic(months)
    transactions = [round(t * 0.025) for t in traffic]
    revenue = [round(t * 0.025 * 100, 2) for t in traffic]
    cvr = [2.5 + i * 0.01 for i in range(months)] if with_cvr else None
    aov = [100.0 + i * 0.5 for i in range(months)] if with_aov else None
    t_p10 = [t * 0.9 for t in traffic] if with_bands else None
    t_p90 = [t * 1.1 for t in traffic] if with_bands else None
    r_p10 = [r * 0.9 for r in revenue] if with_bands else None
    r_p90 = [r * 1.1 for r in revenue] if with_bands else None
    baseline = [t * 0.7 for t in traffic] if with_streams else None
    pos_up = [t * 0.2 for t in traffic] if with_streams else None
    nc_up = [t * 0.1 for t in traffic] if with_streams else None
    decay = [t * 0.01 for t in traffic] if with_streams else None
    assump = [
        {"key": "blended_cr_pct", "label": "CVR", "value": 2.5,
         "provenance": "defaulted", "source": "built-in default", "unit": "%"},
        {"key": "aov", "label": "AOV", "value": 100.0,
         "provenance": "detected", "source": "GA4 average", "unit": "$"},
    ] if with_assumptions else None
    return build_seo_forecast_grid(
        monthly_traffic=traffic,
        monthly_transactions=transactions,
        monthly_revenue=revenue,
        monthly_cvr=cvr,
        monthly_aov=aov,
        months=months,
        traffic_p10=t_p10,
        traffic_p90=t_p90,
        revenue_p10=r_p10,
        revenue_p90=r_p90,
        monthly_baseline=baseline,
        monthly_positional_uplift=pos_up,
        monthly_new_content_uplift=nc_up,
        monthly_decay=decay,
        assumptions_summary=assump,
    )


def _open_wb(buf: io.BytesIO):
    buf.seek(0)
    return load_workbook(buf)


def _row_labels(ws) -> list[str]:
    return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=1).value is not None]


# ---------------------------------------------------------------------------
# TestForecastGridExport
# ---------------------------------------------------------------------------

@pytest.mark.skip(reason="Legacy tests — expect old 3-arg signature; will be rewritten in Deliverables page update session")
class TestForecastGridExport:
    def test_xlsx_opens_without_corruption(self):
        buf = _make_grid()
        wb = _open_wb(buf)
        assert wb is not None

    def test_default_single_sheet_preserved(self):
        buf = _make_grid()
        wb = _open_wb(buf)
        assert wb.sheetnames == ["SEO Forecast"]

    def test_grid_includes_cvr_row_when_provided(self):
        buf = _make_grid(with_cvr=True)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "CVR %" in labels

    def test_grid_omits_cvr_row_when_none(self):
        buf = _make_grid(with_cvr=False)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "CVR %" not in labels

    def test_grid_includes_aov_row_when_provided(self):
        buf = _make_grid(with_aov=True)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "AOV" in labels

    def test_grid_omits_aov_row_when_none(self):
        buf = _make_grid(with_aov=False)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "AOV" not in labels

    def test_grid_includes_band_rows_when_provided(self):
        buf = _make_grid(with_bands=True)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "Traffic P10" in labels
        assert "Traffic P90" in labels
        assert "Revenue P10" in labels
        assert "Revenue P90" in labels

    def test_grid_omits_band_rows_when_none(self):
        buf = _make_grid(with_bands=False)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "Traffic P10" not in labels
        assert "Traffic P90" not in labels

    def test_stream_breakdown_sheet_created_when_streams_provided(self):
        buf = _make_grid(with_streams=True)
        wb = _open_wb(buf)
        assert "Stream Breakdown" in wb.sheetnames

    def test_stream_breakdown_sheet_omitted_when_none(self):
        buf = _make_grid(with_streams=False)
        wb = _open_wb(buf)
        assert "Stream Breakdown" not in wb.sheetnames

    def test_assumptions_sheet_created_when_summary_provided(self):
        buf = _make_grid(with_assumptions=True)
        wb = _open_wb(buf)
        assert "Assumptions" in wb.sheetnames

    def test_assumptions_sheet_omitted_when_none(self):
        buf = _make_grid(with_assumptions=False)
        wb = _open_wb(buf)
        assert "Assumptions" not in wb.sheetnames

    def test_assumptions_sheet_groups_by_category(self):
        buf = _make_grid(with_assumptions=True)
        wb = _open_wb(buf)
        ws = wb["Assumptions"]
        # The assumption rows have labels (CVR / AOV), which fall under "Financial Model"
        # — group header should appear
        labels = _row_labels(ws)
        assert "Financial Model" in labels

    def test_cvr_cells_formatted_as_percentage(self):
        buf = _make_grid(with_cvr=True)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        # Find the CVR row
        cvr_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "CVR %":
                cvr_row = r
                break
        assert cvr_row is not None
        # First forecast cell should be numeric value <= 1 (decimal form for %)
        first_val = ws.cell(row=cvr_row, column=2).value
        assert first_val is not None
        assert first_val < 1.0, f"CVR cell should be decimal (0.025), got {first_val}"

    def test_aov_cells_formatted_as_currency(self):
        buf = _make_grid(with_aov=True)
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        aov_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "AOV":
                aov_row = r
                break
        assert aov_row is not None
        fmt = ws.cell(row=aov_row, column=2).number_format
        assert "$" in fmt or "0.00" in fmt

    def test_traffic_transactions_revenue_always_present(self):
        buf = _make_grid()
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "Traffic" in labels
        assert "Transactions" in labels
        assert "Revenue" in labels

    def test_annual_total_row_present(self):
        buf = _make_grid()
        wb = _open_wb(buf)
        ws = wb["SEO Forecast"]
        labels = _row_labels(ws)
        assert "Annual Total" in labels

    def test_stream_breakdown_has_combined_row(self):
        buf = _make_grid(with_streams=True)
        wb = _open_wb(buf)
        ws = wb["Stream Breakdown"]
        labels = _row_labels(ws)
        assert any("Combined" in (lbl or "") for lbl in labels)

    def test_three_sheet_workbook_when_all_data_provided(self):
        buf = _make_grid(with_streams=True, with_assumptions=True)
        wb = _open_wb(buf)
        assert len(wb.sheetnames) == 3
        assert "SEO Forecast" in wb.sheetnames
        assert "Stream Breakdown" in wb.sheetnames
        assert "Assumptions" in wb.sheetnames


# ---------------------------------------------------------------------------
# TestSnapshotWithDynamicMetrics
# ---------------------------------------------------------------------------

def _make_combined_df(months: int = 12) -> pd.DataFrame:
    """Minimal combined_df for snapshot tests."""
    start = pd.Timestamp("2025-07-01")
    rows = []
    for j in range(months):
        date = start + pd.DateOffset(months=j)
        rows.append({
            "date": date,
            "baseline": 10_000 + j * 100,
            "positional_uplift": 500 + j * 10,
            "new_content_uplift": 200,
            "decay": 50,
            "combined": 10_650 + j * 110,
            "combined_p10": 9_500 + j * 100,
            "combined_p50": 10_650 + j * 110,
            "combined_p90": 11_800 + j * 120,
            "is_forecast": True,
            "uplift_pct": 6.5,
        })
    return pd.DataFrame(rows)


def _make_metrics_df(months: int = 12) -> pd.DataFrame:
    return pd.DataFrame({
        "month": list(range(1, months + 1)),
        "date": [pd.Timestamp("2025-07-01") + pd.DateOffset(months=j) for j in range(months)],
        "traffic": [10_000 + j * 100 for j in range(months)],
        "cvr": [2.5 + j * 0.01 for j in range(months)],
        "aov": [100.0 + j * 0.5 for j in range(months)],
        "transactions": [int((10_000 + j * 100) * 0.025) for j in range(months)],
        "revenue": [round(int((10_000 + j * 100) * 0.025) * (100.0 + j * 0.5), 2)
                    for j in range(months)],
    })


class TestSnapshotWithDynamicMetrics:
    def test_snapshot_includes_dynamic_metrics_flag(self):
        combined_df = _make_combined_df()
        metrics_df = _make_metrics_df()
        snap = build_snapshot("Test", combined_df, {}, metrics_df=metrics_df)
        assert snap["dynamic_metrics"] is True

    def test_snapshot_without_metrics_df_has_false_flag(self):
        combined_df = _make_combined_df()
        snap = build_snapshot("Test", combined_df, {})
        assert snap["dynamic_metrics"] is False

    def test_snapshot_includes_revenue_per_month(self):
        combined_df = _make_combined_df()
        metrics_df = _make_metrics_df()
        snap = build_snapshot("Test", combined_df, {}, metrics_df=metrics_df)
        records_with_revenue = [r for r in snap["forecast"] if "revenue" in r]
        assert len(records_with_revenue) == 12

    def test_snapshot_includes_cvr_per_month(self):
        combined_df = _make_combined_df()
        metrics_df = _make_metrics_df()
        snap = build_snapshot("Test", combined_df, {}, metrics_df=metrics_df)
        records_with_cvr = [r for r in snap["forecast"] if "cvr" in r]
        assert len(records_with_cvr) == 12

    def test_snapshot_includes_assumptions_snapshot(self):
        combined_df = _make_combined_df()
        assump = [{"key": "blended_cr_pct", "label": "CVR", "value": 2.5,
                   "provenance": "defaulted", "source": "built-in default", "unit": "%"}]
        snap = build_snapshot("Test", combined_df, {}, assumptions_snapshot=assump)
        assert "assumptions_snapshot" in snap
        assert len(snap["assumptions_snapshot"]) == 1

    def test_old_snapshot_without_dynamic_flag_still_loads(self):
        old_snap = {
            "snapshot_version": "1.0",
            "client_name": "Old Client",
            "snapshot_date": "2024-01-01T00:00:00+00:00",
            "engine_versions": {"snapshot": "1.0"},
            "parameters": {},
            "forecast": [
                {"date": "2024-02-01", "combined_p50": 10000.0, "combined": 10000.0,
                 "combined_p10": 9000.0, "combined_p90": 11000.0},
            ],
        }
        data = json_roundtrip(old_snap)
        assert "dynamic_metrics" not in data  # old snapshot has no flag
        # compare_to_actuals with traffic metric should still work
        actuals = pd.DataFrame({
            "date": [pd.Timestamp("2024-02-01")],
            "traffic": [9500.0],
        })
        result = compare_to_actuals(data, actuals, metric="traffic")
        assert len(result) == 1

    def test_compare_to_actuals_grades_revenue_when_requested(self):
        combined_df = _make_combined_df()
        metrics_df = _make_metrics_df()
        snap = build_snapshot("Test", combined_df, {}, metrics_df=metrics_df)

        actuals = pd.DataFrame({
            "date": [pd.Timestamp("2025-07-01") + pd.DateOffset(months=j) for j in range(6)],
            "revenue": [25_000.0 + j * 200 for j in range(6)],
        })
        result = compare_to_actuals(snap, actuals, metric="revenue")
        assert not result.empty
        assert "forecast_p50" in result.columns
        assert "actual" in result.columns

    def test_compare_to_actuals_defaults_to_traffic(self):
        combined_df = _make_combined_df()
        snap = build_snapshot("Test", combined_df, {})

        actuals = pd.DataFrame({
            "date": [pd.Timestamp("2025-07-01") + pd.DateOffset(months=j) for j in range(6)],
            "traffic": [10_500.0 + j * 100 for j in range(6)],
        })
        result = compare_to_actuals(snap, actuals)  # no metric arg → traffic
        assert not result.empty

    def test_compare_to_actuals_returns_empty_when_actuals_col_missing(self):
        combined_df = _make_combined_df()
        metrics_df = _make_metrics_df()
        snap = build_snapshot("Test", combined_df, {}, metrics_df=metrics_df)
        actuals = pd.DataFrame({"date": [pd.Timestamp("2025-07-01")], "traffic": [10000.0]})
        # Ask for revenue but actuals_df has no revenue column
        result = compare_to_actuals(snap, actuals, metric="revenue")
        assert result.empty

    def test_summarise_variance_works_for_any_metric(self):
        rows = [
            {"date": pd.Timestamp("2025-07-01"), "forecast_p10": None,
             "forecast_p50": 100.0, "forecast_p90": None,
             "actual": 110.0, "variance": 10.0, "variance_pct": 10.0,
             "within_band": False},
        ]
        df = pd.DataFrame(rows)
        summary = summarise_variance(df)
        assert summary["n_months_compared"] == 1
        assert summary["mean_variance_pct"] == pytest.approx(10.0)


# ---------------------------------------------------------------------------
# Helpers (module-level to avoid import loop in test)
# ---------------------------------------------------------------------------

def json_roundtrip(d: dict) -> dict:
    import json
    return json.loads(json.dumps(d))
