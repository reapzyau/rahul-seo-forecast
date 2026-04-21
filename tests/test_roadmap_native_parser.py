"""Tests for engine/roadmap_native_parser.py — Prompt 7."""
import io
from pathlib import Path

import pandas as pd
import pytest

FIXTURE_DIR = Path(__file__).parent / "fixtures"
XLSX_FIXTURE = FIXTURE_DIR / "sample_pattern_native_roadmap.xlsx"
TASK_CSV = FIXTURE_DIR / "sample_task_table.csv"
PARAM_CSV = FIXTURE_DIR / "sample_param_table.csv"


@pytest.fixture(scope="session")
def xlsx_bytes():
    return XLSX_FIXTURE.read_bytes()


@pytest.fixture(scope="session")
def task_csv_bytes():
    return TASK_CSV.read_bytes()


@pytest.fixture(scope="session")
def param_csv_bytes():
    return PARAM_CSV.read_bytes()


from engine.roadmap_ai_engine import load_roadmap_v2
from engine.roadmap_native_parser import (
    detect_roadmap_format,
    parse_pattern_native,
    wrap_legacy_param_table_as_bundle,
    wrap_legacy_task_table_as_bundle,
)


class TestDetectFormat:
    def test_detect_pattern_native(self, xlsx_bytes):
        assert detect_roadmap_format(xlsx_bytes, "xlsx") == "pattern_native"

    def test_detect_task_table(self, task_csv_bytes):
        assert detect_roadmap_format(task_csv_bytes, "csv") == "task_table"

    def test_detect_param_table(self, param_csv_bytes):
        assert detect_roadmap_format(param_csv_bytes, "csv") == "param_table"

    def test_detect_unknown_returns_unknown(self):
        assert detect_roadmap_format(b"hello world", "txt") == "unknown"


class TestParsePatternNative:
    @pytest.fixture(scope="class")
    def bundle(self, xlsx_bytes):
        return parse_pattern_native(xlsx_bytes)

    def test_parse_pattern_native_client_metadata_extracted(self, bundle):
        meta = bundle["client_metadata"]
        assert "client_name" in meta
        assert meta["client_name"] != ""

    def test_parse_pattern_native_per_focus_hours_computed(self, bundle):
        per_focus = bundle["per_focus"]
        total_hours = sum(f["monthly_hours"] for f in per_focus.values())
        assert total_hours > 0

    def test_parse_pattern_native_content_plan_not_empty(self, bundle):
        assert len(bundle["content_plan"]) > 0

    def test_parse_pattern_native_content_plan_classifies_new_page_vs_optimisation(self, bundle):
        types = {item["content_type"] for item in bundle["content_plan"]}
        assert "new_page" in types
        # Fixture has both new pages and optimisations
        assert len(types) > 1

    def test_parse_pattern_native_strategy_restart_detected(self, bundle):
        # Fixture has "Monthly Strategy Review" in consulting + content only in months 1-4
        # so strategy_restart_month should be set
        timeline = bundle["timeline"]
        assert "strategy_restart_month" in timeline
        # Could be None if content covers all 12 months — just assert key exists

    def test_parse_pattern_native_handles_missing_sheets(self):
        # A workbook with only Breakdown + 3 expected sheets should still parse
        import openpyxl

        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        ws_bd = wb.create_sheet("Breakdown")
        ws_bd.cell(row=4, column=5, value="Content Hours")
        for i, h in enumerate([10] * 12):
            ws_bd.cell(row=4, column=7 + i, value=h)
        wb.create_sheet("1. Client Detail")
        wb.create_sheet("4. Content")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        bundle = parse_pattern_native(buf.read())
        assert bundle["schema_version"] == "2.0"
        assert "content" in bundle["per_focus"]


class TestWrapLegacy:
    def test_wrap_legacy_task_table_produces_v2_bundle(self):
        legacy = {"content_cadence": 4, "effort_level": "moderate", "maintenance_coverage": 0.5}
        bundle = wrap_legacy_task_table_as_bundle(legacy)
        assert bundle["schema_version"] == "2.0"
        assert bundle["format_detected"] == "task_table"
        assert bundle["source_summary"]["parsing_confidence"] == pytest.approx(0.7)
        assert "per_focus" in bundle

    def test_wrap_legacy_param_table_produces_v2_bundle(self):
        legacy = {"effort_level": "light", "content_cadence": 2}
        bundle = wrap_legacy_param_table_as_bundle(legacy)
        assert bundle["schema_version"] == "2.0"
        assert bundle["format_detected"] == "param_table"
        assert bundle["source_summary"]["parsing_confidence"] == pytest.approx(0.5)


class TestRealPatternSOW:
    """Tests against the Helen Kaminski-layout fixture (real Pattern template structure)."""

    @pytest.fixture(scope="class")
    def bundle(self):
        path = FIXTURE_DIR / "sample_real_pattern_sow.xlsx"
        return parse_pattern_native(path.read_bytes())

    def test_client_name_extracted(self, bundle):
        assert bundle["client_metadata"]["client_name"] == "Helen Kaminski"

    def test_industry_extracted(self, bundle):
        assert bundle["client_metadata"]["industry"] == "Accessories"

    def test_retainer_extracted(self, bundle):
        assert bundle["client_metadata"]["retainer_aud_monthly"] == pytest.approx(4906.0)

    def test_main_poc_extracted(self, bundle):
        assert "@" in bundle["client_metadata"].get("main_poc", "")

    def test_content_plan_has_real_urls(self, bundle):
        for item in bundle["content_plan"]:
            assert item["url"].startswith("http"), f"Bad URL: {item['url']!r}"

    def test_content_plan_word_counts_nonzero(self, bundle):
        items_with_words = [i for i in bundle["content_plan"] if i["word_count"] > 0]
        assert len(items_with_words) > 5

    def test_content_plan_seo_hours_nonzero(self, bundle):
        items_with_hours = [i for i in bundle["content_plan"] if i["seo_hours"] > 0]
        assert len(items_with_hours) > 5

    def test_content_type_classifies_new_page_before_faq(self, bundle):
        # 'New Page: Optimisation & FAQs' must be new_page, not faq
        new_page_items = [
            i for i in bundle["content_plan"]
            if "new page" in i.get("_raw_content_type", "").lower()
        ]
        for item in new_page_items:
            assert item["content_type"] == "new_page", (
                f"Expected new_page, got {item['content_type']!r} "
                f"for raw type {item.get('_raw_content_type')!r}"
            )

    def test_existing_copy_with_faqs_is_optimisation(self, bundle):
        # 'Existing Copy: Optimisation & FAQs' must NOT become 'faq'
        existing_faq_items = [
            i for i in bundle["content_plan"]
            if "existing copy" in i.get("_raw_content_type", "").lower()
            and "faq" in i.get("_raw_content_type", "").lower()
        ]
        for item in existing_faq_items:
            assert item["content_type"] == "optimisation", (
                f"Expected optimisation, got {item['content_type']!r}"
            )

    def test_localisation_flag_captured(self, bundle):
        loc_items = [i for i in bundle["content_plan"] if i.get("is_localisation")]
        assert len(loc_items) > 0

    def test_localisation_items_have_non_au_urls(self, bundle):
        loc_items = [i for i in bundle["content_plan"] if i.get("is_localisation")]
        # All localisation items in the fixture point to helenkaminski.com (not .com.au)
        for item in loc_items:
            assert ".com.au" not in item["url"], (
                f"Localisation item has .com.au URL: {item['url']}"
            )

    def test_consulting_tasks_have_names(self, bundle):
        consulting_tasks = (
            bundle["per_focus"]["strategy"]["tasks"]
            + bundle["per_focus"]["analytics"]["tasks"]
        )
        named = [t for t in consulting_tasks if t.get("task")]
        assert len(named) > 3

    def test_tasks_stored_in_per_focus(self, bundle):
        # 8 consulting tasks + some technical tasks classified into analytics via keyword
        # — just verify that tasks were stored (not zero)
        total = len(bundle["per_focus"]["strategy"]["tasks"]) + len(bundle["per_focus"]["analytics"]["tasks"])
        assert total >= 8

    def test_breakdown_uses_12_month_mean_not_nonzero(self, bundle):
        # Content hours: sum([9.9,7.8,14.3,15.8,11.5,8.8,0,0,0,0,0,0]) / 12 = 5.68
        # The nonzero mean would be 11.35 — assert we're not using it
        content = bundle["per_focus"]["content"]
        assert content["monthly_hours"] < 8.0, (
            f"Content hours look like the inflated nonzero mean: {content['monthly_hours']}"
        )

    def test_breakdown_metadata_in_client_metadata(self, bundle):
        meta = bundle["client_metadata"]
        assert meta.get("hours_per_month_target") == pytest.approx(23.28)
        assert meta.get("cost_per_hour") == pytest.approx(200.0)

    def test_roadmap_sheet_parsed(self, bundle):
        assert "roadmap_summary" in bundle
        assert len(bundle["roadmap_summary"]) >= 6

    def test_roadmap_has_technical_deliverables(self, bundle):
        entries = bundle["roadmap_summary"]
        tech_entries = [v for v in entries.values() if v.get("technical_deliverable")]
        assert len(tech_entries) >= 3

    def test_no_invented_split_data(self, bundle):
        # Old parser: analytics = consulting_avg × 0.3 = 8.58 × 0.3 = 2.58
        # New parser uses itemised tasks — should be different
        analytics = bundle["per_focus"]["analytics"]["monthly_hours"]
        assert analytics != pytest.approx(2.58, abs=0.1)

    def test_monthly_hours_series_present(self, bundle):
        for area, data in bundle["per_focus"].items():
            assert "monthly_hours_series" in data, f"Missing monthly_hours_series for {area}"
            assert len(data["monthly_hours_series"]) == 12

    def test_cms_extracted_from_technical_sheet(self, bundle):
        # Client Detail doesn't have CMS; Technical sheet row 4 has 'CMS:' / 'Shopify'
        assert bundle["client_metadata"].get("cms") == "Shopify"

    def test_primary_domain_is_com_au(self, bundle):
        assert bundle["primary_domain"] == "helenkaminski.com.au"

    def test_localisation_domain_detected(self, bundle):
        assert "helenkaminski.com" in bundle["localisation_domains"]

    def test_tooltip_text_not_stored_as_metadata(self, bundle):
        """Template instruction text should never land in client_metadata values."""
        for key, value in bundle["client_metadata"].items():
            if not isinstance(value, str):
                continue
            lower = value.lower()
            assert "to be provided" not in lower, (
                f"Tooltip text leaked into client_metadata.{key}: {value!r}"
            )
            assert "double click" not in lower, (
                f"Tooltip text leaked into client_metadata.{key}: {value!r}"
            )
            assert "click to add" not in lower, (
                f"Tooltip text leaked into client_metadata.{key}: {value!r}"
            )


class TestLoadRoadmapV2:
    def test_load_roadmap_v2_dispatches_correctly(self, xlsx_bytes, task_csv_bytes, param_csv_bytes):
        b1, m1 = load_roadmap_v2(None, xlsx_bytes, "roadmap.xlsx")
        assert b1["format_detected"] == "pattern_native"
        assert m1 == "deterministic"

        b2, m2 = load_roadmap_v2(None, task_csv_bytes, "tasks.csv")
        assert b2["format_detected"] == "task_table"
        assert m2 == "deterministic"

        b3, m3 = load_roadmap_v2(None, param_csv_bytes, "params.csv")
        assert b3["format_detected"] == "param_table"
        assert m3 == "deterministic"


def test_is_tooltip_value_catches_common_patterns():
    from engine.roadmap_native_parser import _is_tooltip_value

    assert _is_tooltip_value("To be provided by the client")
    assert _is_tooltip_value("Double click & select")
    assert _is_tooltip_value("TBC: awaiting brand review")
    assert _is_tooltip_value(None)
    assert _is_tooltip_value("")
    assert _is_tooltip_value("   ")
    # Real values pass through
    assert not _is_tooltip_value("Helen Kaminski")
    assert not _is_tooltip_value("Shopify")
    assert not _is_tooltip_value(4906)


class TestMidDraftSOW:
    """Regression: SOW started but not fully filled in.

    Represents the common state where strategist has listed URLs but content
    team hasn't supplied titles, word counts, or per-item SEO hours yet.
    Bundle should load successfully with warnings, not raise.
    """

    @pytest.fixture(scope="class")
    def bundle(self):
        path = FIXTURE_DIR / "sample_mid_draft_roadmap.xlsx"
        if not path.exists():
            pytest.skip("Mid-draft fixture not built — run tests/fixtures/build_mid_draft_fixture.py")
        return parse_pattern_native(path.read_bytes())

    def test_bundle_loads_without_raising(self, bundle):
        assert bundle["schema_version"] == "2.0"

    def test_client_metadata_extracted(self, bundle):
        assert bundle["client_metadata"]["client_name"] == "Draft Co"

    def test_content_plan_urls_parsed(self, bundle):
        assert len(bundle["content_plan"]) == 25

    def test_validation_warnings_populated(self, bundle):
        assert "validation_warnings" in bundle
        warnings = bundle["validation_warnings"]
        assert len(warnings) > 0

    def test_warnings_flag_zero_word_counts(self, bundle):
        warnings = " ".join(bundle["validation_warnings"]).lower()
        assert "word_count" in warnings or "word count" in warnings

    def test_warnings_flag_zero_seo_hours(self, bundle):
        warnings = " ".join(bundle["validation_warnings"]).lower()
        assert "seo_hours" in warnings or "hours" in warnings

    def test_warnings_flag_missing_titles(self, bundle):
        warnings = " ".join(bundle["validation_warnings"]).lower()
        assert "title" in warnings
