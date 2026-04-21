"""Build the sample_real_pattern_sow.xlsx fixture.

Replicates the actual Pattern SOW template column layout (Helen Kaminski data).
Run directly:
    python tests/fixtures/build_real_sow_fixture.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURE_DIR = Path(__file__).parent
FIXTURE_PATH = FIXTURE_DIR / "sample_real_pattern_sow.xlsx"


def build(output_path: str | Path | None = None) -> Path:
    """Create the real-layout fixture xlsx. Returns the path written to."""
    out = Path(output_path) if output_path else FIXTURE_PATH
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── 1. Client Detail ──────────────────────────────────────────────────────
    # Real layout: col A blank, labels in col B, values in col C, rows 3–14
    ws_cd = wb.create_sheet("1. Client Detail")
    ws_cd.cell(row=1, column=2, value="SEO Retainer Agreement")

    client_rows = [
        (3,  "Client Name",                          "Helen Kaminski"),
        (4,  "Industry",                             "Accessories"),
        (5,  "Project Start Date",                   "2026-01-01"),
        (6,  "Monthly Retainer (excl tech fees)",    4906),
        (7,  "Main POC",                             "poc@helenkaminski.com.au"),
        (8,  "Pattern Client Partner",               "partner@pattern.com"),
        (9,  "Pattern SEO Lead",                     "lead@pattern.com"),
        (10, "Strategy Slides",                      "HK Strategy Deck 2026"),
        (11, "Keyword Research",                     "HK Keyword Research 2026"),
        (12, "Reporting Document",                   "HK Monthly Report"),
        (13, "TOV Document",                         "HK Brand Voice Guide.pdf"),
        (14, "Content Ways of Working",              "Content goes through HK brand review"),
    ]
    for row, label, value in client_rows:
        ws_cd.cell(row=row, column=2, value=label)
        ws_cd.cell(row=row, column=3, value=value)

    # ── Breakdown ─────────────────────────────────────────────────────────────
    # Col E labels, cols G:R (12 months), side metadata in cols B/C
    ws_bd = wb.create_sheet("Breakdown")
    ws_bd.cell(row=1, column=2, value="SEO Retainer Breakdown")

    # Side metadata in cols B/C
    ws_bd.cell(row=6, column=2, value="Monthly Retainer")
    ws_bd.cell(row=6, column=3, value=4906)
    ws_bd.cell(row=7, column=2, value="Hours P/M")
    ws_bd.cell(row=7, column=3, value=23.28)
    ws_bd.cell(row=8, column=2, value="Cost P/H")
    ws_bd.cell(row=8, column=3, value=200)

    # Month headers in row 5, cols G:R
    for i in range(12):
        ws_bd.cell(row=5, column=7 + i, value=f"Month {i + 1}")

    # Service hours — Helen Kaminski actual data
    service_rows = [
        (6,  "Consulting Hours", [8, 8, 8, 8, 8, 15, 8, 8, 8, 8, 8, 8]),
        (7,  "Technical Hours",  [0, 0, 6, 2, 4, 0,  0, 0, 0, 0, 0, 0]),
        (8,  "Content Hours",    [9.9, 7.8, 14.3, 15.8, 11.5, 8.8, 0, 0, 0, 0, 0, 0]),
        (9,  "Link Hours",       [0] * 12),
        (10, "Hours Allocated",  [17.9, 15.8, 28.3, 25.8, 23.5, 23.8, 8, 8, 8, 8, 8, 8]),
        (11, "Hours to Use",     [23.28] * 12),
    ]
    for row_num, label, hours in service_rows:
        ws_bd.cell(row=row_num, column=5, value=label)
        for i, h in enumerate(hours):
            ws_bd.cell(row=row_num, column=7 + i, value=h)

    # Utilisation % row
    ws_bd.cell(row=12, column=5, value="% of Retainer")
    utilisation = [77, 68, 122, 111, 101, 102, 34, 34, 34, 34, 34, 34]
    for i, u in enumerate(utilisation):
        ws_bd.cell(row=12, column=7 + i, value=u)

    # ── 2. Consulting ─────────────────────────────────────────────────────────
    # Real layout: header row 3, data from row 4; col A blank, cols B-E = Task/Desc/Hours/Cadence
    ws_con = wb.create_sheet("2. Consulting")
    ws_con.cell(row=3, column=2, value="Task")
    ws_con.cell(row=3, column=3, value="Description")
    ws_con.cell(row=3, column=4, value="Hours")
    ws_con.cell(row=3, column=5, value="Cadence")

    consulting_tasks = [
        ("Monthly Strategy Meeting",      "Monthly strategic direction call with client",       1.0, "Monthly"),
        ("Monthly Reporting",             "GA4 + GSC monthly report and dashboard update",      2.0, "Monthly"),
        ("Competitor Analysis",           "Competitive landscape review",                        4.0, "Quarterly"),
        ("Keyword Research & Mapping",    "Full keyword universe mapping to site architecture", 8.0, "Bi-Annual"),
        ("Content Gap Analysis",          "Identify content opportunities vs. competitors",     4.0, "Quarterly"),
        ("SEO Health Check",              "Comprehensive site health audit and scoring",         2.0, "Monthly"),
        ("QBR Presentation",              "Quarterly business review deck preparation",          3.0, "Quarterly"),
        ("Onboarding",                    "Initial strategy, brand review and setup",            8.0, "Onboarding Month"),
    ]
    for row_idx, (task, desc, hrs, cadence) in enumerate(consulting_tasks, 4):
        ws_con.cell(row=row_idx, column=2, value=task)
        ws_con.cell(row=row_idx, column=3, value=desc)
        ws_con.cell(row=row_idx, column=4, value=hrs)
        ws_con.cell(row=row_idx, column=5, value=cadence)

    # ── 3. Technical ──────────────────────────────────────────────────────────
    ws_tech = wb.create_sheet("3. Technical")
    ws_tech.cell(row=3, column=2, value="Task")
    ws_tech.cell(row=3, column=3, value="Description")
    ws_tech.cell(row=3, column=4, value="Hours")
    ws_tech.cell(row=3, column=5, value="Cadence")

    # CMS at row 4 col B/C (special metadata row)
    ws_tech.cell(row=4, column=2, value="CMS:")
    ws_tech.cell(row=4, column=3, value="Shopify")

    technical_tasks = [
        (5, "Technical SEO Audit",          "Full crawl, redirect audit, coverage analysis",       6.0, "Monthly"),
        (6, "Core Web Vitals Remediation",  "CWV fixes: LCP, CLS, FID",                            2.0, "Monthly"),
        (7, "Schema Markup Implementation", "Product, breadcrumb, org schema markup",               4.0, "Monthly"),
        (8, "Crawl Error Monitoring",       "Monitor and resolve crawl errors via GSC",             1.0, "Monthly"),
        (9, "Site Speed Optimisation",      "Image compression, lazy load, CDN config",             2.0, "Quarterly"),
    ]
    for row_num, task, desc, hrs, cadence in technical_tasks:
        ws_tech.cell(row=row_num, column=2, value=task)
        ws_tech.cell(row=row_num, column=3, value=desc)
        ws_tech.cell(row=row_num, column=4, value=hrs)
        ws_tech.cell(row=row_num, column=5, value=cadence)

    # ── 4. Content ────────────────────────────────────────────────────────────
    # Header row 7, data from row 8
    # Cols: A blank | B Month | C Month Name | D Content Name | E URL |
    #        F Total Words | G Content Type | H Keywords | I Brief Detail |
    #        J FAQ Questions | K SEO Hours | L Template | M Production | N SEO Review | O SEO Impl
    ws_cont = wb.create_sheet("4. Content")
    headers = [
        (2, "Month"), (3, "Month Name"), (4, "Content Name"), (5, "URL"),
        (6, "Total Words"), (7, "Content Type"), (8, "Keywords"),
        (9, "Content Brief Detail"), (10, "FAQ Questions"), (11, "SEO Hours"),
        (12, "Template set-up"), (13, "Production time"),
        (14, "SEO Review Time"), (15, "SEO Implementation time"),
    ]
    for col, hdr in headers:
        ws_cont.cell(row=7, column=col, value=hdr)

    # Mix of content types, including 'New Page: Optimisation & FAQs' (must be new_page not faq),
    # 'Existing Copy: Optimisation & FAQs' (must be optimisation not faq),
    # and localisation rows (col I = 'localisation')
    content_rows = [
        # M1 primary domain
        ("Month 1", "January 2026", "Homepage", "https://helenkaminski.com.au/",
         300, "New Page: Optimisation", "helen kaminski hats", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 1", "January 2026", "Womens Hats Collection", "https://helenkaminski.com.au/collections/womens-hats",
         300, "Existing Copy: Optimisation", "womens hats australia", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 1", "January 2026", "Straw Hats", "https://helenkaminski.com.au/collections/straw-hats",
         300, "Existing Copy: Optimisation & FAQs", "straw hats", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 1", "January 2026", "Summer Hats", "https://helenkaminski.com.au/collections/summer-hats",
         300, "New Page: Optimisation & FAQs", "summer hats", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 1", "January 2026", "Beach Hats", "https://helenkaminski.com.au/collections/beach-hats",
         300, "Existing Copy: Optimisation", "beach hats", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        # M1 localisation rows (helenkaminski.com, not .com.au)
        ("Month 1", "January 2026", "Homepage US", "https://helenkaminski.com/",
         150, "New Page: Optimisation", "helen kaminski hats", "localisation",
         "", 1.0, 0.5, 0.5, 0.5, 0.0),
        ("Month 1", "January 2026", "Womens Hats US", "https://helenkaminski.com/collections/womens-hats",
         150, "Existing Copy: Optimisation", "womens hats", "localisation",
         "", 1.0, 0.5, 0.5, 0.5, 0.0),
        # M2 primary domain
        ("Month 2", "February 2026", "Sun Hats", "https://helenkaminski.com.au/collections/sun-hats",
         300, "Existing Copy: Optimisation", "sun hats australia", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 2", "February 2026", "Hats Blog Post", "https://helenkaminski.com.au/blogs/news/hat-guide",
         500, "New Page: Optimisation", "how to choose a hat", "",
         "", 2.5, 0.5, 1.5, 0.5, 0.5),
        ("Month 2", "February 2026", "Hat Care Guide", "https://helenkaminski.com.au/pages/hat-care",
         300, "Existing Copy: Optimisation & FAQs", "hat care tips", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 2", "February 2026", "Raffia Hats", "https://helenkaminski.com.au/collections/raffia-hats",
         300, "Existing Copy: Optimisation", "raffia hats", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        # M2 localisation
        ("Month 2", "February 2026", "Sun Hats US", "https://helenkaminski.com/collections/sun-hats",
         150, "Existing Copy: Optimisation", "sun hats", "localisation",
         "", 1.0, 0.5, 0.5, 0.5, 0.0),
        ("Month 2", "February 2026", "Hat Care US", "https://helenkaminski.com/pages/hat-care",
         150, "Existing Copy: Optimisation & FAQs", "hat care", "localisation",
         "", 1.0, 0.5, 0.5, 0.5, 0.0),
        # M3 primary domain
        ("Month 3", "March 2026", "Winter Hats", "https://helenkaminski.com.au/collections/winter-hats",
         300, "Existing Copy: Optimisation", "winter hats australia", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 3", "March 2026", "Knit Beanies", "https://helenkaminski.com.au/collections/beanies",
         300, "New Page: Optimisation", "knit beanies australia", "",
         "", 2.0, 0.5, 1.0, 0.5, 0.5),
    ]
    for row_idx, row_data in enumerate(content_rows, 8):
        (month_str, month_name, content_name, url, words, ct, kw, brief,
         faq, seo_h, tmpl, prod, rev, impl) = row_data
        ws_cont.cell(row=row_idx, column=2, value=month_str)
        ws_cont.cell(row=row_idx, column=3, value=month_name)
        ws_cont.cell(row=row_idx, column=4, value=content_name)
        ws_cont.cell(row=row_idx, column=5, value=url)
        ws_cont.cell(row=row_idx, column=6, value=words)
        ws_cont.cell(row=row_idx, column=7, value=ct)
        ws_cont.cell(row=row_idx, column=8, value=kw)
        ws_cont.cell(row=row_idx, column=9, value=brief)
        ws_cont.cell(row=row_idx, column=10, value=faq)
        ws_cont.cell(row=row_idx, column=11, value=seo_h)
        ws_cont.cell(row=row_idx, column=12, value=tmpl)
        ws_cont.cell(row=row_idx, column=13, value=prod)
        ws_cont.cell(row=row_idx, column=14, value=rev)
        ws_cont.cell(row=row_idx, column=15, value=impl)

    # ── 5. Links ──────────────────────────────────────────────────────────────
    ws_links = wb.create_sheet("5. Links")
    ws_links.cell(row=3, column=2, value="Task")
    ws_links.cell(row=3, column=3, value="Description")
    ws_links.cell(row=3, column=4, value="Hours")
    ws_links.cell(row=3, column=5, value="Cadence")

    links_tasks = [
        (4, "Digital PR Outreach", "Pitch to AU lifestyle publications", 0.0, "Monthly"),
        (5, "Guest Post Sourcing",  "Identify and pitch guest post opportunities", 0.0, "Monthly"),
        (6, "Link Reclamation",    "Recover unlinked brand mentions", 0.0, "Quarterly"),
    ]
    for row_num, task, desc, hrs, cadence in links_tasks:
        ws_links.cell(row=row_num, column=2, value=task)
        ws_links.cell(row=row_num, column=3, value=desc)
        ws_links.cell(row=row_num, column=4, value=hrs)
        ws_links.cell(row=row_num, column=5, value=cadence)

    # ── Roadmap ───────────────────────────────────────────────────────────────
    ws_rm = wb.create_sheet("Roadmap")
    ws_rm.cell(row=1, column=2, value="SEO Roadmap 2026")

    roadmap_data = [
        (4,  "Month 1", "Technical audit, redirect fixes, schema setup",
              "Month 1", "2026-01-31", "https://helenkaminski.com.au/\nhttps://helenkaminski.com.au/collections/womens-hats"),
        (5,  "Month 2", "CWV remediation, internal linking pass",
              "Month 2", "2026-02-28", "https://helenkaminski.com.au/collections/sun-hats"),
        (6,  "Month 3", "Schema expansion, crawl error resolution",
              "Month 3", "2026-03-31", "https://helenkaminski.com.au/collections/winter-hats"),
        (7,  "Month 4", "Page speed optimisation",
              "Month 4", "2026-04-30", "https://helenkaminski.com.au/collections/autumn-hats"),
        (8,  "Month 5", "Structured data audit",
              "Month 5", "2026-05-31", "https://helenkaminski.com.au/collections/travel-hats"),
        (9,  "Month 6", "Mid-year technical review",
              "Month 6", "2026-06-30", "https://helenkaminski.com.au/collections/resort-hats"),
    ]
    for row_num, tech_month, tech_del, content_month, launch, content_del in roadmap_data:
        ws_rm.cell(row=row_num, column=2, value=tech_month)
        ws_rm.cell(row=row_num, column=4, value=tech_del)
        ws_rm.cell(row=row_num, column=7, value=content_month)
        ws_rm.cell(row=row_num, column=8, value=launch)
        ws_rm.cell(row=row_num, column=9, value=content_del)

    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Fixture written to: {path}")
