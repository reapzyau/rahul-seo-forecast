"""Build sample_mid_draft_roadmap.xlsx — a SOW that's been started but not filled in.

Represents the state where content URLs are listed but no word counts, titles,
or per-item SEO hours have been allocated yet. Tests the validation_warnings path.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURE_DIR = Path(__file__).parent
FIXTURE_PATH = FIXTURE_DIR / "sample_mid_draft_roadmap.xlsx"


def build(output_path: str | Path | None = None) -> Path:
    out = Path(output_path) if output_path else FIXTURE_PATH
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── 1. Client Detail ──
    ws_cd = wb.create_sheet("1. Client Detail")
    ws_cd.cell(row=1, column=2, value="SEO Retainer Agreement")
    client_rows = [
        (3,  "Client Name",                       "Draft Co"),
        (4,  "Industry",                          "Apparel"),
        (5,  "Project Start Date",                "2026-06-01"),
        (6,  "Monthly Retainer (excl tech fees)", 5000),
        (10, "CMS",                               "Shopify"),
    ]
    for row, label, value in client_rows:
        ws_cd.cell(row=row, column=2, value=label)
        ws_cd.cell(row=row, column=3, value=value)

    # ── Breakdown ──
    ws_bd = wb.create_sheet("Breakdown")
    ws_bd.cell(row=6, column=2, value="Monthly Retainer")
    ws_bd.cell(row=6, column=3, value=5000)
    ws_bd.cell(row=7, column=2, value="Hours P/M")
    ws_bd.cell(row=7, column=3, value=25.0)
    ws_bd.cell(row=8, column=2, value="Cost P/H")
    ws_bd.cell(row=8, column=3, value=200)
    for i in range(12):
        ws_bd.cell(row=5, column=7 + i, value=f"Month {i + 1}")
    service_rows = [
        (6,  "Consulting Hours", [5] * 6 + [0] * 6),
        (7,  "Technical Hours",  [3] * 6 + [0] * 6),
        (8,  "Content Hours",    [12] * 6 + [0] * 6),
        (9,  "Link Hours",       [0] * 12),
    ]
    for row_num, label, hours in service_rows:
        ws_bd.cell(row=row_num, column=5, value=label)
        for i, h in enumerate(hours):
            ws_bd.cell(row=row_num, column=7 + i, value=h)

    # ── 2. Consulting ──
    ws_con = wb.create_sheet("2. Consulting")
    ws_con.cell(row=3, column=2, value="Task")
    ws_con.cell(row=3, column=3, value="Description")
    ws_con.cell(row=3, column=4, value="Hours")
    ws_con.cell(row=3, column=5, value="Cadence")
    ws_con.cell(row=4, column=2, value="Monthly Strategy")
    ws_con.cell(row=4, column=4, value=5.0)
    ws_con.cell(row=4, column=5, value="Monthly")

    # ── 3. Technical ──
    ws_tech = wb.create_sheet("3. Technical")
    ws_tech.cell(row=3, column=2, value="Task")
    ws_tech.cell(row=3, column=3, value="Description")
    ws_tech.cell(row=3, column=4, value="Hours")
    ws_tech.cell(row=3, column=5, value="Cadence")
    ws_tech.cell(row=4, column=2, value="CMS:")
    ws_tech.cell(row=4, column=3, value="Shopify")

    # ── 4. Content — 25 URLs but no titles, word counts, or SEO hours ──
    ws_cont = wb.create_sheet("4. Content")
    headers = [
        (2, "Month"), (3, "Month Name"), (4, "Content Name"), (5, "URL"),
        (6, "Total Words"), (7, "Content Type"), (8, "Keywords"),
        (9, "Content Brief Detail"), (10, "FAQ Questions"), (11, "SEO Hours"),
    ]
    for col, hdr in headers:
        ws_cont.cell(row=7, column=col, value=hdr)

    for row_idx in range(8, 8 + 25):
        month_num = ((row_idx - 8) // 5) + 1  # 5 pages per month across 5 months
        ws_cont.cell(row=row_idx, column=2, value=f"Month {month_num}")
        ws_cont.cell(row=row_idx, column=3, value=f"Month {month_num}")
        # NO content name (col 4 left blank)
        ws_cont.cell(row=row_idx, column=5, value=f"https://draftco.com/page-{row_idx - 8}")
        # NO word count (col 6 blank)
        ws_cont.cell(row=row_idx, column=7, value="Existing Copy: Optimisation")
        # NO keywords, brief, FAQ, or SEO hours

    # ── 5. Links ──
    ws_links = wb.create_sheet("5. Links")
    ws_links.cell(row=3, column=2, value="Task")
    ws_links.cell(row=3, column=3, value="Description")
    ws_links.cell(row=3, column=4, value="Hours")
    ws_links.cell(row=3, column=5, value="Cadence")

    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Fixture written to: {path}")
