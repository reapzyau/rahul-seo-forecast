"""Standalone script that creates the sample_pattern_native_roadmap.xlsx fixture.

Updated to match the real Pattern SOW template column layout:
- Client Detail: labels in col B, values in col C
- Task sheets: header at row 3, data from row 4; col A blank, cols B-E
- Content sheet: header at row 7, data from row 8; col A blank, cols B-O

Run directly:
    python tests/fixtures/build_fixture.py

Or import and call build() from a pytest conftest.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

FIXTURE_DIR = Path(__file__).parent
FIXTURE_PATH = FIXTURE_DIR / "sample_pattern_native_roadmap.xlsx"


def build(output_path: str | Path | None = None) -> Path:
    """Create the fixture xlsx. Returns the path written to."""
    out = Path(output_path) if output_path else FIXTURE_PATH
    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Breakdown sheet ────────────────────────────────────────────────────────
    # Col E (index 4) has focus labels; cols G:R (cols 7-18) have monthly hours
    # Side metadata in cols B/C
    ws_bd = wb.create_sheet("Breakdown")
    ws_bd.cell(row=1, column=2, value="SEO Retainer Breakdown")

    ws_bd.cell(row=6, column=2, value="Monthly Retainer")
    ws_bd.cell(row=6, column=3, value=5000)
    ws_bd.cell(row=7, column=2, value="Hours P/M")
    ws_bd.cell(row=7, column=3, value=24.0)
    ws_bd.cell(row=8, column=2, value="Cost P/H")
    ws_bd.cell(row=8, column=3, value=200)

    for i in range(12):
        ws_bd.cell(row=5, column=7 + i, value=f"Month {i + 1}")

    service_rows = [
        (6,  "Consulting Hours", [8] * 12),
        (7,  "Technical Hours",  [12] * 12),
        (8,  "Content Hours",    [20] * 12),
        (9,  "Link Hours",       [6] * 12),
        (10, "Hours Allocated",  [46] * 12),
        (11, "Hours to Use",     [24] * 12),
    ]
    for row_num, label, hours in service_rows:
        ws_bd.cell(row=row_num, column=5, value=label)
        for i, h in enumerate(hours):
            ws_bd.cell(row=row_num, column=7 + i, value=h)

    ws_bd.cell(row=12, column=5, value="% of Retainer")
    for i in range(12):
        ws_bd.cell(row=12, column=7 + i, value=100)

    # ── 1. Client Detail sheet ─────────────────────────────────────────────────
    # Real layout: col A blank, labels in col B, values in col C
    ws_cd = wb.create_sheet("1. Client Detail")
    ws_cd.cell(row=1, column=2, value="SEO Retainer Agreement")

    client_data = [
        (3,  "Client Name",                       "Sample Retail Co"),
        (4,  "Industry",                          "Fashion"),
        (5,  "Project Start Date",                "2026-07-01"),
        (6,  "Monthly Retainer (excl tech fees)", "AUD 5000"),
        (7,  "Main POC",                          "contact@sampleretail.com"),
        (8,  "Pattern Client Partner",            "partner@pattern.com"),
        (9,  "Pattern SEO Lead",                  "lead@pattern.com"),
        (10, "CMS",                               "Shopify"),
        (11, "Strategy Slides",                   "Sample Retail Strategy 2026"),
    ]
    for row, label, value in client_data:
        ws_cd.cell(row=row, column=2, value=label)
        ws_cd.cell(row=row, column=3, value=value)

    # ── 2. Consulting sheet ────────────────────────────────────────────────────
    # Real layout: header at row 3, data from row 4; col A blank, cols B-E
    ws_con = wb.create_sheet("2. Consulting")
    ws_con.cell(row=3, column=2, value="Task")
    ws_con.cell(row=3, column=3, value="Description")
    ws_con.cell(row=3, column=4, value="Hours")
    ws_con.cell(row=3, column=5, value="Cadence")

    consulting_tasks = [
        ("Monthly Strategy Review",    "Monthly strategy direction call",         4.0, "Monthly"),
        ("Keyword Research & Mapping", "Keyword universe mapping",                 8.0, "Quarterly"),
        ("Competitor Analysis",        "Competitive landscape review",             6.0, "Quarterly"),
        ("GA4 & GSC Reporting",        "Monthly performance report and dashboard", 3.0, "Monthly"),
        ("Content Gap Analysis",       "Identify content opportunities",           8.0, "Bi-Annual"),
    ]
    for row_idx, (task, desc, hrs, cadence) in enumerate(consulting_tasks, 4):
        ws_con.cell(row=row_idx, column=2, value=task)
        ws_con.cell(row=row_idx, column=3, value=desc)
        ws_con.cell(row=row_idx, column=4, value=hrs)
        ws_con.cell(row=row_idx, column=5, value=cadence)

    # ── 3. Technical sheet ─────────────────────────────────────────────────────
    ws_tech = wb.create_sheet("3. Technical")
    ws_tech.cell(row=3, column=2, value="Task")
    ws_tech.cell(row=3, column=3, value="Description")
    ws_tech.cell(row=3, column=4, value="Hours")
    ws_tech.cell(row=3, column=5, value="Cadence")

    ws_tech.cell(row=4, column=2, value="CMS:")
    ws_tech.cell(row=4, column=3, value="Shopify")

    technical_tasks = [
        (5, "Core Web Vitals Audit",          "CWV audit and fix recommendations",     6.0, "Quarterly"),
        (6, "Crawl Error Remediation",        "Fix crawl errors from GSC",             4.0, "Monthly"),
        (7, "Schema Markup Implementation",   "Product and breadcrumb schema",          8.0, "Quarterly"),
        (8, "Site Speed Optimisation",        "Image and script optimisation",          4.0, "Monthly"),
        (9, "Internal Linking Audit",         "Silo structure internal link review",   10.0, "Bi-Annual"),
    ]
    for row_num, task, desc, hrs, cadence in technical_tasks:
        ws_tech.cell(row=row_num, column=2, value=task)
        ws_tech.cell(row=row_num, column=3, value=desc)
        ws_tech.cell(row=row_num, column=4, value=hrs)
        ws_tech.cell(row=row_num, column=5, value=cadence)

    # ── 4. Content sheet ───────────────────────────────────────────────────────
    # Real layout: header row 7, data rows 8+; col A blank, cols B-O
    ws_cont = wb.create_sheet("4. Content")
    ws_cont.cell(row=1, column=2, value="Content Production Plan")

    content_headers = [
        (2, "Month"), (3, "Month Name"), (4, "Content Name"), (5, "URL"),
        (6, "Total Words"), (7, "Content Type"), (8, "Keywords"),
        (9, "Content Brief Detail"), (10, "FAQ Questions"), (11, "SEO Hours"),
        (12, "Template set-up"), (13, "Production time"),
        (14, "SEO Review Time"), (15, "SEO Implementation time"),
    ]
    for col, hdr in content_headers:
        ws_cont.cell(row=7, column=col, value=hdr)

    content_rows = [
        ("Month 1", "July",      "Summer Fashion Guide 2026",    "https://example.com/blog/summer-fashion",
         2000, "New Page: Optimisation",           "summer fashion guide", "", "", 3.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 1", "July",      "How to Style Denim FAQ",       "https://example.com/faq/how-to-style-denim",
         800,  "Existing Copy: Optimisation & FAQs", "how to style denim", "", "", 2.0, 0.5, 0.5, 0.5, 0.5),
        ("Month 2", "August",    "Spring Trends 2026",           "https://example.com/blog/spring-trends",
         1800, "New Page: Optimisation",           "spring fashion trends", "", "", 3.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 2", "August",    "Women's Dresses Category",     "https://example.com/categories/womens-dresses",
         500,  "Existing Copy: Optimisation",      "womens dresses", "", "", 2.0, 0.5, 1.0, 0.5, 0.5),
        ("Month 3", "September", "Workwear Essentials Guide",    "https://example.com/blog/workwear-essentials",
         2200, "New Page: Optimisation",           "workwear essentials", "", "", 4.0, 0.5, 1.5, 0.5, 0.5),
        ("Month 3", "September", "Classic Blazer Product Page",  "https://example.com/products/classic-blazer",
         400,  "Existing Copy: Optimisation",      "classic blazer", "", "", 1.0, 0.5, 0.5, 0.0, 0.0),
        ("Month 4", "October",   "Autumn Layering Techniques",   "https://example.com/blog/autumn-layering",
         1900, "New Page: Optimisation",           "autumn layering", "", "", 3.0, 0.5, 1.0, 0.5, 0.5),
    ]
    for row_idx, row_data in enumerate(content_rows, 8):
        (month_str, month_name, content_name, url, words, ct, kw, brief,
         faq, seo_h, tmpl, prod, rev, impl) = row_data
        ws_cont.cell(row=row_idx, column=2, value=month_str)
        ws_cont.cell(row=row_idx, column=3, value=month_name)
        ws_cont.cell(row=row_idx, column=4, value=content_name)
        ws_cont.cell(row=row_idx, column=5, value=url)
        ws_cont.cell(row=row_idx, column=6, value=words)
        ws_cont.cell(row=row_idx, column=7, value=ct)
        ws_cont.cell(row=row_idx, column=8, value=kw)
        ws_cont.cell(row=row_idx, column=9, value=brief)
        ws_cont.cell(row=row_idx, column=10, value=faq)
        ws_cont.cell(row=row_idx, column=11, value=seo_h)
        ws_cont.cell(row=row_idx, column=12, value=tmpl)
        ws_cont.cell(row=row_idx, column=13, value=prod)
        ws_cont.cell(row=row_idx, column=14, value=rev)
        ws_cont.cell(row=row_idx, column=15, value=impl)

    # ── 5. Links sheet ─────────────────────────────────────────────────────────
    ws_links = wb.create_sheet("5. Links")
    ws_links.cell(row=3, column=2, value="Task")
    ws_links.cell(row=3, column=3, value="Description")
    ws_links.cell(row=3, column=4, value="Hours")
    ws_links.cell(row=3, column=5, value="Cadence")

    links_tasks = [
        (4, "Digital PR Outreach", "Pitch to lifestyle publications", 4.0, "Monthly"),
        (5, "Guest Post Sourcing",  "Identify guest post opportunities", 2.0, "Monthly"),
        (6, "Link Reclamation",    "Recover unlinked brand mentions",   3.0, "Quarterly"),
    ]
    for row_num, task, desc, hrs, cadence in links_tasks:
        ws_links.cell(row=row_num, column=2, value=task)
        ws_links.cell(row=row_num, column=3, value=desc)
        ws_links.cell(row=row_num, column=4, value=hrs)
        ws_links.cell(row=row_num, column=5, value=cadence)

    wb.save(out)
    return out


if __name__ == "__main__":
    path = build()
    print(f"Fixture written to: {path}")
