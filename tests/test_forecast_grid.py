"""Tests for utils/forecast_grid.py."""
import io

import openpyxl
import pandas as pd
import pytest

from tests.fixtures import make_ga4_df, make_semrush_kw_df


@pytest.fixture
def fake_scenario_results():
    """Minimal scenario_results dict with synthetic data."""
    from engine.scenario_engine import build_scenario_presets, run_three_scenarios
    ga4 = make_ga4_df(months=12, starting_traffic=10000)
    kw = make_semrush_kw_df(n=30, positions=[8] * 30)
    kw["is_branded"] = False
    presets = build_scenario_presets()
    results = run_three_scenarios(
        ga4_df=ga4, kw_df=kw, kw_existing=kw,
        presets=presets, months=6, seed=42,
    )
    return results, presets


class TestScenarioGridSheet:
    def test_returns_one_row_per_forecast_month(self, fake_scenario_results):
        from utils.forecast_grid import build_scenario_grid_sheet
        results, presets = fake_scenario_results
        df = build_scenario_grid_sheet(
            "Moderate", results["Moderate"], presets["Moderate"],
            cvr_series=2.5, aov_series=100.0,
        )
        forecast_rows = df[df["Is Forecast"]]
        assert len(forecast_rows) == 6

    def test_has_all_expected_columns(self, fake_scenario_results):
        from utils.forecast_grid import build_scenario_grid_sheet
        results, presets = fake_scenario_results
        df = build_scenario_grid_sheet(
            "Moderate", results["Moderate"], presets["Moderate"],
            cvr_series=2.5, aov_series=100.0,
        )
        expected = {
            "Month Label", "Baseline Traffic", "Positional Uplift",
            "New Content Uplift", "Decay", "Traffic P50",
            "Transactions", "Revenue", "AOV Used", "CVR Used (%)",
            "Avg Portfolio Position", "Avg Portfolio CTR (%)",
        }
        assert expected.issubset(set(df.columns))

    def test_decay_shown_as_negative_or_zero(self, fake_scenario_results):
        from utils.forecast_grid import build_scenario_grid_sheet
        results, presets = fake_scenario_results
        df = build_scenario_grid_sheet(
            "Moderate", results["Moderate"], presets["Moderate"],
            cvr_series=2.5, aov_series=100.0,
        )
        # Decay values should be ≤ 0 (subtractive)
        assert (df["Decay"] <= 0).all()

    def test_seasonal_aov_series_reflected(self, fake_scenario_results):
        from utils.forecast_grid import build_scenario_grid_sheet
        results, presets = fake_scenario_results
        aov_series = [80, 90, 100, 110, 120, 130]
        df = build_scenario_grid_sheet(
            "Moderate", results["Moderate"], presets["Moderate"],
            cvr_series=2.5, aov_series=aov_series,
        )
        forecast = df[df["Is Forecast"]].reset_index(drop=True)
        assert forecast["AOV Used"].tolist() == aov_series


class TestBuildThreeScenarioGrid:
    def test_produces_four_sheets(self, fake_scenario_results):
        from utils.forecast_grid import build_three_scenario_grid
        results, presets = fake_scenario_results
        buf = build_three_scenario_grid(
            scenario_results=results, presets=presets,
            cvr=2.5, aov=100.0, start_month=7,
        )
        wb = openpyxl.load_workbook(buf)
        assert {"Conservative", "Moderate", "Aggressive", "Comparison"}.issubset(set(wb.sheetnames))

    def test_comparison_sheet_has_three_scenario_columns(self, fake_scenario_results):
        from utils.forecast_grid import build_three_scenario_grid
        results, presets = fake_scenario_results
        buf = build_three_scenario_grid(
            scenario_results=results, presets=presets,
            cvr=2.5, aov=100.0, start_month=7,
        )
        wb = openpyxl.load_workbook(buf)
        comp = wb["Comparison"]
        # Row 3 is the header row with scenario names
        header_values = [c.value for c in comp[3]]
        assert "Conservative" in header_values
        assert "Moderate" in header_values
        assert "Aggressive" in header_values

    def test_seasonal_aov_applied_when_enabled(self, fake_scenario_results):
        from utils.forecast_grid import build_three_scenario_grid
        results, presets = fake_scenario_results
        seasonality = {m: {"traffic_mod": 0, "cr_mod": 0, "aov_mod": 0.0} for m in range(1, 13)}
        seasonality[11]["aov_mod"] = 0.20  # Nov gets a 20% AOV boost
        buf = build_three_scenario_grid(
            scenario_results=results, presets=presets,
            cvr=2.5, aov=100.0, seasonality=seasonality,
            apply_seasonal_aov=True, start_month=11,
        )
        wb = openpyxl.load_workbook(buf)
        mod_sheet = wb["Moderate"]  # noqa: F841
        # November should have AOV of 120, January 100
        buf.seek(0)
        df = pd.read_excel(buf, sheet_name="Moderate", header=3)
        aov_col = df["AOV Used"]
        # Month 1 of forecast starts in November → AOV 120
        assert aov_col.iloc[0] == pytest.approx(120.0)

    def test_writes_valid_xlsx(self, fake_scenario_results):
        from utils.forecast_grid import build_three_scenario_grid
        results, presets = fake_scenario_results
        buf = build_three_scenario_grid(
            scenario_results=results, presets=presets,
            cvr=2.5, aov=100.0, start_month=7,
        )
        data = buf.read()
        assert len(data) > 0
        assert data[:2] == b"PK"  # xlsx zip magic bytes
