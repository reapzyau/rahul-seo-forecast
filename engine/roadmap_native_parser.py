"""Deterministic parser for Pattern-native xlsx roadmap format.

Parses the multi-sheet xlsx produced by Pattern's retainer template without
requiring an AI call. Falls back to legacy scalar wrappers for simpler formats.

Pure Python — no Streamlit imports.
"""
from __future__ import annotations

import io
import re
from typing import Any

import openpyxl

# ── Sheet name constants ──────────────────────────────────────────────────────

PATTERN_NATIVE_SHEETS_REQUIRED = {"Breakdown"}
PATTERN_NATIVE_SHEETS_EXPECTED = {
    "1. Client Detail",
    "2. Consulting",
    "3. Technical",
    "4. Content",
    "5. Links",
}

# ── Normalisation maps ────────────────────────────────────────────────────────

INDUSTRY_NORMALISATION: dict[str, str] = {
    "accessories": "Accessories",
    "apparel": "Apparel",
    "clothing": "Apparel",
    "fashion": "Apparel",
    "beauty": "Beauty",
    "cosmetics": "Beauty",
    "home": "Home",
    "homewares": "Home",
    "b2b saas": "B2B SaaS",
    "saas": "B2B SaaS",
    "automotive": "Automotive",
    "travel": "Travel",
    "food": "Food & Beverage",
    "health": "Health",
    "fitness": "Health",
    "finance": "Finance",
    "fintech": "Finance",
}

# ── Effort classification thresholds (avg monthly hours) ─────────────────────

_EFFORT_HOURS_THRESHOLDS: list[tuple[float, str]] = [
    (8.0, "light"),
    (20.0, "moderate"),
    (float("inf"), "aggressive"),
]

# ── Column aliases for format detection ──────────────────────────────────────

_TASK_COL_ALIASES = {"task", "activity"}
_FOCUS_COL_ALIASES = {"focus", "area", "category"}
_OCC_COL_ALIASES = {"occurrence", "frequency"}
_HOURS_COL_ALIASES = {"hours", "hrs"}
_CADENCE_COL_ALIASES = {"cadence", "content_cadence", "posts_per_month"}
_EFFORT_COL_ALIASES = {"effort_level", "effort"}
_MAINT_COL_ALIASES = {"maintenance_coverage", "maintenance", "maint_coverage"}

# ── Cadence → monthly multiplier ─────────────────────────────────────────────

_CADENCE_MULTIPLIER: dict[str, float] = {
    "monthly": 1.0,
    "bi-monthly": 0.5,
    "bi monthly": 0.5,
    "quarterly": 1 / 3,
    "bi-annual": 1 / 6,
    "bi annual": 1 / 6,
    "6 months": 1 / 6,
    "annual": 1 / 12,
    "one-off": 1 / 12,
    "one off": 1 / 12,
    "onboarding month": 1 / 12,
}

# ── Focus-area keyword classifier ─────────────────────────────────────────────

_FOCUS_KEYWORDS: dict[str, list[str]] = {
    "analytics": ["report", "reporting", "dashboard", "ga4", "gsc", "tracking", "health check"],
    "strategy": [
        "strategy", "review", "qbr", "roadmap", "planning", "onboarding",
        "kickoff", "client meeting", "client management", "ad-hoc", "fluid",
    ],
    "technical": ["audit", "schema", "core web vitals", "cwv", "speed", "crawl",
                  "redirect", "shopify", "cms", "robots", "sitemap"],
    "on_page": ["page strategy", "internal link", "metadata", "title tag",
                "on-page", "on page", "product grid"],
    "off_page": ["link", "outreach", "digital pr", "guest post", "backlink"],
    "local": ["gmb", "google my business", "local citation", "local landing"],
    "content": ["article", "blog", "content production", "content calendar",
                "content brief", "long-form", "longform"],
}


def _classify_task_focus(task_name: str, description: str, default: str = "strategy") -> str:
    """Map a task name + description to one of the 7 focus area keys."""
    text = (task_name + " " + description).lower()
    for focus, keywords in _FOCUS_KEYWORDS.items():
        if any(kw in text for kw in keywords):
            return focus
    return default


def _cadence_to_monthly(cadence_str: str) -> float:
    """Convert a cadence string to a monthly fraction."""
    key = cadence_str.strip().lower()
    for pattern, mult in _CADENCE_MULTIPLIER.items():
        if pattern in key:
            return mult
    return 1.0  # default to monthly


# ── Format detection ──────────────────────────────────────────────────────────


def detect_roadmap_format(raw_bytes: bytes, file_extension: str) -> str:
    """Detect which roadmap format the bytes represent.

    Returns one of: "pattern_native", "task_table", "param_table", "unknown".
    """
    ext = file_extension.lower().lstrip(".")

    # Only xlsx can be Pattern-native
    if ext in ("xlsx", "xls"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
            sheet_names = set(wb.sheetnames)
            wb.close()
            if PATTERN_NATIVE_SHEETS_REQUIRED.issubset(sheet_names):
                matching_expected = sheet_names & PATTERN_NATIVE_SHEETS_EXPECTED
                if len(matching_expected) >= 3:
                    return "pattern_native"
        except Exception:
            pass

    # Try reading as a flat table (csv or xlsx)
    try:
        import pandas as pd

        buf = io.BytesIO(raw_bytes)
        if ext in ("xlsx", "xls"):
            df = pd.read_excel(buf, engine="openpyxl")
        else:
            try:
                df = pd.read_csv(buf)
            except Exception:
                return "unknown"

        cols_lower = {str(c).strip().lower() for c in df.columns}

        has_task = bool(cols_lower & _TASK_COL_ALIASES)
        has_focus = bool(cols_lower & _FOCUS_COL_ALIASES)
        has_occ = bool(cols_lower & _OCC_COL_ALIASES)
        has_hours = bool(cols_lower & _HOURS_COL_ALIASES)
        has_cadence = bool(cols_lower & _CADENCE_COL_ALIASES)
        has_effort = bool(cols_lower & _EFFORT_COL_ALIASES)
        has_maint = bool(cols_lower & _MAINT_COL_ALIASES)

        # param_table: has cadence or effort_level or maintenance_coverage
        if has_cadence or has_maint or (has_effort and not has_task and not has_focus):
            return "param_table"

        # task_table: has task + focus + occurrence + hours (at least 3 of 4)
        task_score = sum([has_task, has_focus, has_occ, has_hours])
        if task_score >= 3:
            return "task_table"

    except Exception:
        pass

    return "unknown"


# ── Empty per-focus skeleton ──────────────────────────────────────────────────


def _empty_per_focus() -> dict:
    """Return a per_focus dict with all 7 focus areas at zero/moderate defaults."""
    focus_areas = ["content", "technical", "on_page", "off_page", "local", "analytics", "strategy"]
    return {
        area: {
            "effort_level": "moderate",
            "monthly_hours": 0.0,
            "monthly_hours_series": [0.0] * 12,
            "cadence": 0,
            "task_count": 0,
            "tasks": [],
        }
        for area in focus_areas
    }


# ── Client detail parser ──────────────────────────────────────────────────────


def _parse_client_detail(ws) -> dict:
    """Parse the '1. Client Detail' worksheet into a metadata dict.

    Real template layout: labels in col B (index 1), values in col C (index 2).
    Rows 3–20 (template has grown; scan generously).
    """
    result: dict[str, Any] = {}

    for row in ws.iter_rows(min_row=3, max_row=20, values_only=True):
        if not row or len(row) < 2:
            continue
        # Labels are in col B (index 1), values in col C (index 2)
        label_cell = row[1]
        value_cell = row[2] if len(row) > 2 else None
        if label_cell is None:
            continue
        label = str(label_cell).strip().lower()
        value = value_cell

        if "client name" in label and value is not None:
            result["client_name"] = str(value).strip()

        elif "industry" in label and value is not None:
            raw = str(value).strip().lower()
            result["industry"] = INDUSTRY_NORMALISATION.get(raw, str(value).strip())

        elif "retainer" in label and value is not None:
            num_str = re.sub(r"[^0-9.]", "", str(value))
            try:
                result["retainer_aud_monthly"] = float(num_str)
            except (ValueError, TypeError):
                result["retainer_aud_monthly"] = str(value).strip()

        elif "project start" in label and value is not None:
            result["project_start_date"] = str(value).strip()

        elif label == "cms" or label.startswith("cms"):
            if value is not None:
                result["cms"] = str(value).strip()

        elif "main poc" in label and value is not None:
            result["main_poc"] = str(value).strip()

        elif "pattern client partner" in label and value is not None:
            result["pattern_client_partner"] = str(value).strip()

        elif "pattern seo lead" in label and value is not None:
            result["pattern_seo_lead"] = str(value).strip()

        elif "strategy slides" in label and value is not None:
            result["strategy_slides"] = str(value).strip()

        elif "keyword research" in label and value is not None:
            result["keyword_research_doc"] = str(value).strip()

        elif "reporting document" in label and value is not None:
            result["reporting_document"] = str(value).strip()

        elif "tov document" in label and value is not None:
            result["tov_document"] = str(value).strip()

        elif "content ways of working" in label and value is not None:
            result["content_ways_of_working"] = str(value).strip()

    return result


# ── Breakdown parser ──────────────────────────────────────────────────────────


def _parse_breakdown(ws) -> dict:
    """Parse the 'Breakdown' worksheet.

    Returns monthly hours per service type plus side metadata from cols B/C:
      - consulting/technical/content/link: list[float] of 12 monthly values
      - metadata: dict with monthly_retainer_aud, hours_per_month_target, cost_per_hour,
                  monthly_utilisation_pct (12-element list)
    """
    result: dict = {
        "consulting": [0.0] * 12,
        "technical": [0.0] * 12,
        "content": [0.0] * 12,
        "link": [0.0] * 12,
        "metadata": {},
    }

    label_map = {
        "consulting hours": "consulting",
        "technical hours": "technical",
        "content hours": "content",
        "link hours": "link",
    }

    # Side metadata labels in col B (index 1), values in col C (index 2)
    side_meta_map = {
        "monthly retainer": "monthly_retainer_aud",
        "hours p/m": "hours_per_month_target",
        "cost p/h": "cost_per_hour",
    }

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue

        # Check col B side metadata
        cell_b = row[1] if len(row) > 1 else None
        cell_c = row[2] if len(row) > 2 else None
        if cell_b is not None:
            b_lower = str(cell_b).strip().lower()
            for meta_label, meta_key in side_meta_map.items():
                if meta_label in b_lower and cell_c is not None:
                    try:
                        result["metadata"][meta_key] = float(
                            re.sub(r"[^0-9.]", "", str(cell_c)) or "0"
                        )
                    except (ValueError, TypeError):
                        pass

        if len(row) < 7:
            continue

        cell_e = row[4]  # col E (0-indexed index 4)
        if cell_e is None:
            continue

        label_lower = str(cell_e).strip().lower()
        matched_key = None
        for label_fragment, key in label_map.items():
            if label_fragment in label_lower:
                matched_key = key
                break

        if matched_key is not None:
            # Extract cols G:R (0-indexed 6:18)
            monthly_values = []
            for val in row[6:18]:
                if val is not None and val != "":
                    try:
                        monthly_values.append(float(val))
                    except (ValueError, TypeError):
                        monthly_values.append(0.0)
                else:
                    monthly_values.append(0.0)
            while len(monthly_values) < 12:
                monthly_values.append(0.0)
            result[matched_key] = monthly_values[:12]
            continue

        # Capture utilisation % row (col E contains "% of retainer")
        if "% of retainer" in label_lower or "% retainer" in label_lower:
            util_values = []
            for val in row[6:18]:
                if val is not None and val != "":
                    try:
                        util_values.append(float(val))
                    except (ValueError, TypeError):
                        util_values.append(0.0)
                else:
                    util_values.append(0.0)
            while len(util_values) < 12:
                util_values.append(0.0)
            result["metadata"]["monthly_utilisation_pct"] = util_values[:12]

    return result


# ── Effort classification ─────────────────────────────────────────────────────


def _classify_effort_hours(hours: float) -> str:
    """Classify average monthly hours into an effort level label."""
    for threshold, label in _EFFORT_HOURS_THRESHOLDS:
        if hours <= threshold:
            return label
    return "aggressive"


# ── Apply breakdown to bundle ─────────────────────────────────────────────────


def _apply_breakdown_to_bundle(
    bundle: dict,
    breakdown: dict,
    consulting_tasks: list[dict] | None = None,
    technical_tasks: list[dict] | None = None,
) -> None:
    """Map breakdown monthly hours into per_focus entries.

    Uses true 12-month mean (not nonzero mean) to avoid inflating partial-year retainers.
    Splits consulting→strategy/analytics and technical→technical/on_page proportionally
    from itemised task data when available; falls back to 70/30 only when no tasks exist.
    """
    per_focus = bundle["per_focus"]

    def _twelve_month_mean(values: list) -> float:
        vals = list(values) + [0.0] * max(0, 12 - len(values))
        return sum(vals[:12]) / 12

    def _task_proportions(tasks: list[dict], focus_a: str, focus_b: str) -> tuple[float, float]:
        """Return (prop_a, prop_b) from task monthly-equivalent hours."""
        eq: dict[str, float] = {focus_a: 0.0, focus_b: 0.0}
        for t in (tasks or []):
            classified = _classify_task_focus(t.get("task", ""), t.get("description", ""))
            if classified not in eq:
                # Route uncategorised tasks to the first focus
                classified = focus_a
            cadence = t.get("cadence", "Monthly")
            eq[classified] += t.get("hours", 0.0) * _cadence_to_monthly(cadence)
        total = eq[focus_a] + eq[focus_b]
        if total == 0:
            return 0.70, 0.30
        return eq[focus_a] / total, eq[focus_b] / total

    # Consulting → strategy / analytics
    consulting_avg = _twelve_month_mean(breakdown.get("consulting", []))
    strategy_prop, analytics_prop = _task_proportions(consulting_tasks, "strategy", "analytics")
    strategy_hours = consulting_avg * strategy_prop
    analytics_hours = consulting_avg * analytics_prop
    per_focus["strategy"]["monthly_hours"] = round(strategy_hours, 2)
    per_focus["strategy"]["effort_level"] = _classify_effort_hours(strategy_hours)
    per_focus["strategy"]["monthly_hours_series"] = [
        round(v * strategy_prop, 2) for v in breakdown.get("consulting", [0.0] * 12)
    ]
    per_focus["analytics"]["monthly_hours"] = round(analytics_hours, 2)
    per_focus["analytics"]["effort_level"] = _classify_effort_hours(analytics_hours)
    per_focus["analytics"]["monthly_hours_series"] = [
        round(v * analytics_prop, 2) for v in breakdown.get("consulting", [0.0] * 12)
    ]

    # Technical → technical / on_page
    technical_avg = _twelve_month_mean(breakdown.get("technical", []))
    tech_prop, on_page_prop = _task_proportions(technical_tasks, "technical", "on_page")
    tech_hours = technical_avg * tech_prop
    on_page_hours = technical_avg * on_page_prop
    per_focus["technical"]["monthly_hours"] = round(tech_hours, 2)
    per_focus["technical"]["effort_level"] = _classify_effort_hours(tech_hours)
    per_focus["technical"]["monthly_hours_series"] = [
        round(v * tech_prop, 2) for v in breakdown.get("technical", [0.0] * 12)
    ]
    per_focus["on_page"]["monthly_hours"] = round(on_page_hours, 2)
    per_focus["on_page"]["effort_level"] = _classify_effort_hours(on_page_hours)
    per_focus["on_page"]["monthly_hours_series"] = [
        round(v * on_page_prop, 2) for v in breakdown.get("technical", [0.0] * 12)
    ]

    # Content → content (direct)
    content_avg = _twelve_month_mean(breakdown.get("content", []))
    per_focus["content"]["monthly_hours"] = round(content_avg, 2)
    per_focus["content"]["effort_level"] = _classify_effort_hours(content_avg)
    per_focus["content"]["monthly_hours_series"] = [
        round(v, 2) for v in breakdown.get("content", [0.0] * 12)
    ]

    # Link → off_page (direct)
    link_avg = _twelve_month_mean(breakdown.get("link", []))
    per_focus["off_page"]["monthly_hours"] = round(link_avg, 2)
    per_focus["off_page"]["effort_level"] = _classify_effort_hours(link_avg)
    per_focus["off_page"]["monthly_hours_series"] = [
        round(v, 2) for v in breakdown.get("link", [0.0] * 12)
    ]

    # Propagate breakdown side metadata to client_metadata
    meta = breakdown.get("metadata", {})
    if meta:
        bundle["client_metadata"].update({
            k: v for k, v in meta.items() if k not in bundle["client_metadata"]
        })


# ── Content plan parser ───────────────────────────────────────────────────────


_MONTH_NUM_RE = re.compile(r"\d+")


def _parse_content_plan(ws) -> list[dict]:
    """Parse the '4. Content' worksheet into a list of content item dicts.

    Real template layout (header row 7, data from row 8, 0-indexed):
      row[1]=col B Month string  row[2]=col C Month Name  row[3]=col D Content Name
      row[4]=col E URL           row[5]=col F Total Words  row[6]=col G Content Type
      row[7]=col H Keywords      row[8]=col I Brief Detail row[9]=col J FAQ Questions
      row[10]=col K SEO Hours    row[11]=col L Template    row[12]=col M Production
      row[13]=col N SEO Review   row[14]=col O SEO Impl
    """

    def _safe_float(val) -> float:
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    def _safe_int(val) -> int:
        if val is None:
            return 0
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return 0

    def _get(row, idx, default=""):
        return row[idx] if len(row) > idx and row[idx] is not None else default

    items = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row:
            continue
        # URL is at col E (index 4) — use this as the empty-row guard
        url_val = _get(row, 4)
        if not url_val or str(url_val).strip() == "":
            continue

        # Parse month number from 'Month 1', 'Month  2', 'Month3', etc.
        month_raw = _get(row, 1)
        month_match = _MONTH_NUM_RE.search(str(month_raw)) if month_raw else None
        month = int(month_match.group()) if month_match else None

        month_name = str(_get(row, 2)).strip()
        title = str(_get(row, 3)).strip()
        url = str(url_val).strip()
        word_count = _safe_int(_get(row, 5, None))
        raw_content_type = str(_get(row, 6)).strip()
        keywords_raw = str(_get(row, 7)).strip()
        brief_detail = str(_get(row, 8)).strip()
        seo_hours = _safe_float(_get(row, 10, None))

        # Content type: check 'new page' BEFORE 'faq' — 'faq' substring also appears
        # in 'Existing Copy: Optimisation & FAQs' which must stay as 'optimisation'
        ct_lower = raw_content_type.lower()
        if "new page" in ct_lower:
            content_type = "new_page"
        elif "optimisation" in ct_lower or "optimization" in ct_lower:
            content_type = "optimisation"
        elif "faq" in ct_lower:
            content_type = "faq"
        else:
            content_type = "optimisation"

        keywords = [k.strip() for k in keywords_raw.split(",") if k.strip()] if keywords_raw else []
        is_localisation = "localisation" in brief_detail.lower() or "localization" in brief_detail.lower()

        time_breakdown = {
            "template_setup": _safe_float(_get(row, 11, None)),
            "production": _safe_float(_get(row, 12, None)),
            "seo_review": _safe_float(_get(row, 13, None)),
            "seo_implementation": _safe_float(_get(row, 14, None)),
        }

        items.append(
            {
                "month": month,
                "month_name": month_name,
                "url": url,
                "title": title,
                "content_type": content_type,
                "_raw_content_type": raw_content_type,
                "word_count": word_count,
                "seo_hours": seo_hours,
                "keywords": keywords,
                "is_localisation": is_localisation,
                "brief_detail": brief_detail,
                "time_breakdown": time_breakdown,
            }
        )
    return items


# ── Task sheet parser ─────────────────────────────────────────────────────────


def _parse_task_sheet(ws, default_focus: str = "strategy") -> list[dict]:
    """Parse a Consulting/Technical/Links worksheet.

    Real template layout: header at row 3, data from row 4.
      col B (index 1) = Task
      col C (index 2) = Description
      col D (index 3) = Hours
      col E (index 4) = Cadence
    """
    tasks = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        if all(v is None for v in row):
            continue

        task_val = row[1] if len(row) > 1 else None
        desc_val = row[2] if len(row) > 2 else None
        hrs_val = row[3] if len(row) > 3 else None
        cadence_val = row[4] if len(row) > 4 else None

        # Skip rows where both task and description are empty (header echoes, footers, etc.)
        task_str = str(task_val).strip() if task_val is not None else ""
        if not task_str:
            continue

        try:
            hours = float(hrs_val) if hrs_val is not None else 0.0
        except (ValueError, TypeError):
            hours = 0.0

        desc_str = str(desc_val).strip() if desc_val is not None else ""
        cadence_str = str(cadence_val).strip() if cadence_val is not None else "Monthly"
        focus = _classify_task_focus(task_str, desc_str, default=default_focus)

        tasks.append(
            {
                "task": task_str,
                "description": desc_str,
                "cadence": cadence_str,
                "hours": hours,
                "focus": focus,
            }
        )
    return tasks


# ── Roadmap sheet parser ──────────────────────────────────────────────────────


def _parse_roadmap_sheet(ws) -> dict:
    """Parse the 'Roadmap' worksheet into a dict keyed by month number.

    Template layout (each logical month occupies one row starting at row 4):
      col B (index 1) = 'Month #' label
      col D (index 3) = Technical deliverable text
      col G (index 6) = 'Month #' label (content side)
      col H (index 7) = Content launch date
      col I (index 8) = Content deliverable (URLs joined with newlines)
    """
    roadmap: dict[int, dict] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue

        # Technical side: col B has 'Month N'
        month_b = row[1] if len(row) > 1 else None
        if month_b is not None:
            m_match = _MONTH_NUM_RE.search(str(month_b))
            if m_match:
                month_num = int(m_match.group())
                tech_del = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                entry = roadmap.setdefault(month_num, {
                    "technical_deliverable": "",
                    "content_summary": "",
                    "content_launch_date": None,
                })
                if tech_del:
                    entry["technical_deliverable"] = tech_del

        # Content side: col G has 'Month N'
        month_g = row[6] if len(row) > 6 else None
        if month_g is not None:
            m_match = _MONTH_NUM_RE.search(str(month_g))
            if m_match:
                month_num = int(m_match.group())
                launch_date = row[7] if len(row) > 7 else None
                content_del = str(row[8]).strip() if len(row) > 8 and row[8] else ""
                entry = roadmap.setdefault(month_num, {
                    "technical_deliverable": "",
                    "content_summary": "",
                    "content_launch_date": None,
                })
                if launch_date is not None:
                    entry["content_launch_date"] = str(launch_date)
                if content_del:
                    entry["content_summary"] = content_del

    return roadmap


# ── Timeline finalisation ─────────────────────────────────────────────────────


def _finalise_timeline(bundle: dict) -> None:
    """Compute timeline metadata and store it in bundle['timeline']."""
    content_plan = bundle.get("content_plan", [])
    months_with_content = [
        item["month"]
        for item in content_plan
        if item.get("month") and isinstance(item["month"], (int, float))
    ]
    last_month = int(max(months_with_content)) if months_with_content else 12

    consulting_tasks = bundle.get("_consulting_tasks", [])
    has_strategy_review = any(
        "strategy review" in (t.get("task") or "").lower() for t in consulting_tasks
    )

    strategy_restart_month: int | None = None
    if last_month < 12 and has_strategy_review:
        strategy_restart_month = last_month + 1

    bundle["timeline"] = {
        "months_covered": 12,
        "strategy_restart_month": strategy_restart_month,
        "has_launch_dates": bool(months_with_content),
    }

    # Clean up the private key
    bundle.pop("_consulting_tasks", None)


# ── Global rollup finalisation ────────────────────────────────────────────────


def _finalise_global_rollup(bundle: dict) -> None:
    """Compute a global rollup across all per_focus entries."""
    per_focus = bundle.get("per_focus", {})

    total_monthly_hours = sum(f.get("monthly_hours", 0.0) for f in per_focus.values())

    # Max effort across all focus areas
    effort_order = ["light", "moderate", "aggressive"]
    effort_levels = [f.get("effort_level", "moderate") for f in per_focus.values() if f.get("monthly_hours", 0.0) > 0]
    if effort_levels:
        max_effort = max(effort_levels, key=lambda e: effort_order.index(e) if e in effort_order else 0)
    else:
        max_effort = "moderate"

    # Content cadence: number of content items per month from content_plan
    content_plan = bundle.get("content_plan", [])
    new_pages = [item for item in content_plan if item.get("content_type") == "new_page"]
    # Deduplicate by month to get posts per month average
    months_seen: dict[int, int] = {}
    for item in new_pages:
        m = item.get("month")
        if m:
            months_seen[m] = months_seen.get(m, 0) + 1
    content_cadence = round(sum(months_seen.values()) / len(months_seen)) if months_seen else 0

    # Maintenance coverage: fraction of portfolio covered by technical + on_page
    tech_hours = per_focus.get("technical", {}).get("monthly_hours", 0.0)
    on_page_hours = per_focus.get("on_page", {}).get("monthly_hours", 0.0)
    maintenance_hours = tech_hours + on_page_hours
    maintenance_coverage = round(min(maintenance_hours / max(total_monthly_hours, 1.0), 1.0), 2)

    bundle["global_rollup"] = {
        "total_monthly_hours": round(total_monthly_hours, 2),
        "effort_level": max_effort,
        "maintenance_coverage": maintenance_coverage,
        "content_cadence": content_cadence,
        "positional_effort_level": per_focus.get("on_page", {}).get("effort_level", "moderate"),
    }


# ── Main parser ───────────────────────────────────────────────────────────────


def parse_pattern_native(raw_bytes: bytes) -> dict:
    """Parse a Pattern-native xlsx roadmap into a v2 bundle dict."""
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    sheet_names = wb.sheetnames

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "pattern_native",
        "extraction_method": "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.95,
        },
        "client_metadata": {},
        "per_focus": _empty_per_focus(),
        "content_plan": [],
        "timeline": {},
        "global_rollup": {},
        "roadmap_summary": {},
        "recommendations": [],
        "gaps": [],
    }

    if "1. Client Detail" in sheet_names:
        bundle["client_metadata"] = _parse_client_detail(wb["1. Client Detail"])
        bundle["source_summary"]["sheets_parsed"].append("1. Client Detail")

    # Parse task sheets first so their data can inform the breakdown splits
    consulting_tasks: list[dict] = []
    if "2. Consulting" in sheet_names:
        consulting_tasks = _parse_task_sheet(wb["2. Consulting"], default_focus="strategy")
        bundle["_consulting_tasks"] = consulting_tasks
        # Store tasks in per_focus entries
        for t in consulting_tasks:
            focus = t.get("focus", "strategy")
            if focus not in bundle["per_focus"]:
                focus = "strategy"
            bundle["per_focus"][focus]["tasks"].append(t)
        bundle["source_summary"]["sheets_parsed"].append("2. Consulting")

    technical_tasks: list[dict] = []
    if "3. Technical" in sheet_names:
        technical_tasks = _parse_task_sheet(wb["3. Technical"], default_focus="technical")
        for t in technical_tasks:
            focus = t.get("focus", "technical")
            if focus not in bundle["per_focus"]:
                focus = "technical"
            bundle["per_focus"][focus]["tasks"].append(t)
        # CMS fallback: check row 4 col B/C of Technical sheet when Client Detail didn't supply it
        if not bundle["client_metadata"].get("cms"):
            ws_tech = wb["3. Technical"]
            for row in ws_tech.iter_rows(min_row=3, max_row=6, values_only=True):
                if not row or len(row) < 3:
                    continue
                b_val = row[1]
                c_val = row[2]
                if b_val and "cms" in str(b_val).lower() and c_val:
                    bundle["client_metadata"]["cms"] = str(c_val).strip()
                    break
        bundle["source_summary"]["sheets_parsed"].append("3. Technical")

    if "5. Links" in sheet_names:
        links_tasks = _parse_task_sheet(wb["5. Links"], default_focus="off_page")
        for t in links_tasks:
            bundle["per_focus"]["off_page"]["tasks"].append(t)
        bundle["source_summary"]["sheets_parsed"].append("5. Links")

    if "Breakdown" in sheet_names:
        breakdown = _parse_breakdown(wb["Breakdown"])
        _apply_breakdown_to_bundle(bundle, breakdown, consulting_tasks, technical_tasks)
        bundle["source_summary"]["sheets_parsed"].append("Breakdown")

    if "4. Content" in sheet_names:
        bundle["content_plan"] = _parse_content_plan(wb["4. Content"])
        bundle["source_summary"]["content_launches_detected"] = len(bundle["content_plan"])
        bundle["source_summary"]["sheets_parsed"].append("4. Content")

    if "Roadmap" in sheet_names:
        bundle["roadmap_summary"] = _parse_roadmap_sheet(wb["Roadmap"])
        bundle["source_summary"]["sheets_parsed"].append("Roadmap")

    # Update task counts
    for focus_data in bundle["per_focus"].values():
        focus_data["task_count"] = len(focus_data.get("tasks", []))

    _finalise_timeline(bundle)
    _finalise_global_rollup(bundle)
    return bundle


# ── Legacy wrappers ───────────────────────────────────────────────────────────


def wrap_legacy_task_table_as_bundle(legacy_result: dict) -> dict:
    """Wrap a legacy parse_task_table() output dict into a v2 bundle.

    Args:
        legacy_result: Dict with keys: content_cadence, effort_level,
            maintenance_coverage (and optional _monthly_hours).

    Returns:
        A v2 bundle with parsing_confidence = 0.7.
    """
    effort = legacy_result.get("effort_level", "moderate")
    cadence = legacy_result.get("content_cadence", 4)
    maintenance = legacy_result.get("maintenance_coverage", 0.0)
    monthly_hours = legacy_result.get("_monthly_hours", 0.0)

    per_focus = _empty_per_focus()
    # Apply effort to content, on_page, off_page (the focus areas that the
    # task-table loader addresses)
    for area in ("content", "on_page", "off_page"):
        per_focus[area]["effort_level"] = effort

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "task_table",
        "extraction_method": "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.7,
        },
        "client_metadata": {},
        "per_focus": per_focus,
        "content_plan": [],
        "timeline": {
            "months_covered": 12,
            "strategy_restart_month": None,
            "has_launch_dates": False,
        },
        "global_rollup": {
            "total_monthly_hours": float(monthly_hours),
            "effort_level": effort,
            "maintenance_coverage": float(maintenance),
            "content_cadence": int(cadence),
            "positional_effort_level": effort,
        },
        "recommendations": [],
        "gaps": [],
    }
    return bundle


def wrap_legacy_param_table_as_bundle(legacy_result: dict) -> dict:
    """Wrap a legacy parse_param_table() output dict into a v2 bundle.

    Args:
        legacy_result: Dict with optional keys: content_cadence, effort_level,
            maintenance_coverage.

    Returns:
        A v2 bundle with parsing_confidence = 0.5.
    """
    effort = legacy_result.get("effort_level", "moderate")
    cadence = legacy_result.get("content_cadence", 4)
    maintenance = legacy_result.get("maintenance_coverage", 0.0)

    per_focus = _empty_per_focus()

    bundle: dict = {
        "schema_version": "2.0",
        "format_detected": "param_table",
        "extraction_method": "deterministic",
        "source_summary": {
            "sheets_parsed": [],
            "parsing_confidence": 0.5,
        },
        "client_metadata": {},
        "per_focus": per_focus,
        "content_plan": [],
        "timeline": {
            "months_covered": 12,
            "strategy_restart_month": None,
            "has_launch_dates": False,
        },
        "global_rollup": {
            "total_monthly_hours": 0.0,
            "effort_level": effort,
            "maintenance_coverage": float(maintenance),
            "content_cadence": int(cadence),
            "positional_effort_level": effort,
        },
        "recommendations": [],
        "gaps": [],
    }
    return bundle
