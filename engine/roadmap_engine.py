"""SEO roadmap engine — GAZMAN-style Task/Focus/Occurrence monthly grid with xlsx export.

Replaces the legacy budget_engine.py with a richer month-by-month hour
allocation grid and colour-coded Excel export.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ── Focus area colour palette ────────────────────────────────────────────────

FOCUS_COLORS: dict[str, str] = {
    "Technical": "#2563EB",
    "Content": "#22C55E",
    "On-Page": "#F97316",
    "Off-Page": "#8B5CF6",
    "Local": "#EAB308",
    "Analytics": "#06B6D4",
    "Strategy": "#EF4444",
}

# ── Default SEO task catalogue ───────────────────────────────────────────────

DEFAULT_SEO_TASKS: list[dict] = [
    {"task": "Technical Audit",           "focus": "Technical",  "occurrence": "Bi-Annual",  "hours": 8.0},
    {"task": "Core Web Vitals Check",     "focus": "Technical",  "occurrence": "Quarterly",  "hours": 4.0},
    {"task": "Long-Form Article Production", "focus": "Content", "occurrence": "Monthly",    "hours": 10.0},
    {"task": "Content Calendar & Planning",  "focus": "Content", "occurrence": "Monthly",    "hours": 2.0},
    {"task": "Page Optimisation",         "focus": "On-Page",    "occurrence": "Monthly",    "hours": 4.0},
    {"task": "Internal Linking Audit",    "focus": "On-Page",    "occurrence": "Quarterly",  "hours": 3.0},
    {"task": "Link Building & Outreach",  "focus": "Off-Page",   "occurrence": "Monthly",    "hours": 8.0},
    {"task": "Digital PR Campaign",       "focus": "Off-Page",   "occurrence": "Quarterly",  "hours": 6.0},
    {"task": "GMB Optimisation",          "focus": "Local",      "occurrence": "Monthly",    "hours": 2.0},
    {"task": "Monthly Reporting",         "focus": "Analytics",  "occurrence": "Monthly",    "hours": 3.0},
    {"task": "Quarterly Strategy Review", "focus": "Strategy",   "occurrence": "Quarterly",  "hours": 4.0},
]


# ── Occurrence helpers ───────────────────────────────────────────────────────

def _active_months(occurrence: str, months: int) -> list[int]:
    """Return 1-indexed month numbers where a task is scheduled."""
    occ = occurrence.strip().lower().replace("-", "")
    if occ == "monthly":
        return list(range(1, months + 1))
    if occ == "bimonthly":
        return [m for m in range(1, months + 1) if m % 2 == 1]
    if occ == "quarterly":
        return [m for m in range(1, months + 1) if (m - 1) % 3 == 0]
    if occ == "biannual":
        return [m for m in range(1, months + 1) if (m - 1) % 6 == 0]
    if occ in ("annual", "oneoff", "one off"):
        return [1] if months >= 1 else []
    # Fallback: treat unknown occurrences as one-off
    return [1] if months >= 1 else []


# ── Core builder ─────────────────────────────────────────────────────────────

def build_roadmap(
    tasks: list[dict] | None = None,
    months: int = 12,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Build a GAZMAN-style SEO roadmap with monthly hour grid.

    Args:
        tasks: List of task dicts (task, focus, occurrence, hours).
               Defaults to DEFAULT_SEO_TASKS.
        months: Number of months to plan.

    Returns:
        Tuple of (task_df, monthly_df, summary).
            task_df  — one row per task (Task, Focus, Occurrence, Hours).
            monthly_df — grid with M1..M{months} columns + Total Hours row.
            summary  — dict with n_tasks, total_hours, avg_hours_per_month,
                        peak_month, peak_hours.
    """
    task_list = tasks or DEFAULT_SEO_TASKS

    # ── task_df ──────────────────────────────────────────────────────────
    task_df = pd.DataFrame([
        {
            "Task": t["task"],
            "Focus": t["focus"],
            "Occurrence": t["occurrence"],
            "Hours": t["hours"],
        }
        for t in task_list
    ])

    # ── monthly_df (grid) ────────────────────────────────────────────────
    month_cols = [f"M{m}" for m in range(1, months + 1)]
    rows = []
    for t in task_list:
        active = set(_active_months(t["occurrence"], months))
        row: dict = {
            "Task": t["task"],
            "Focus": t["focus"],
            "Occurrence": t["occurrence"],
        }
        total = 0.0
        for m in range(1, months + 1):
            hrs = t["hours"] if m in active else 0.0
            row[f"M{m}"] = hrs
            total += hrs
        row["Total Hours"] = total
        rows.append(row)

    monthly_df = pd.DataFrame(rows)

    # TOTAL row
    total_row: dict = {
        "Task": "TOTAL",
        "Focus": "",
        "Occurrence": "",
    }
    for col in month_cols:
        total_row[col] = monthly_df[col].sum()
    total_row["Total Hours"] = monthly_df["Total Hours"].sum()
    monthly_df = pd.concat(
        [monthly_df, pd.DataFrame([total_row])], ignore_index=True
    )

    # ── summary ──────────────────────────────────────────────────────────
    # Exclude TOTAL row for per-month stats
    month_sums = {
        col: monthly_df.loc[monthly_df["Task"] != "TOTAL", col].sum()
        for col in month_cols
    }
    total_hours = sum(month_sums.values())
    avg_hours = total_hours / months if months > 0 else 0.0
    peak_col = max(month_sums, key=month_sums.get) if month_sums else "M1"
    peak_month = int(peak_col.replace("M", ""))
    peak_hours = month_sums.get(peak_col, 0.0)

    summary = {
        "n_tasks": len(task_list),
        "total_hours": total_hours,
        "avg_hours_per_month": round(avg_hours, 2),
        "peak_month": peak_month,
        "peak_hours": peak_hours,
    }

    return task_df, monthly_df, summary


# ── Excel export ─────────────────────────────────────────────────────────────

def build_roadmap_xlsx(
    monthly_df: pd.DataFrame,
    summary: dict,
    hourly_rate: float = 200.0,
    client_name: str = "",
    fy_label: str = "FY26",
) -> io.BytesIO:
    """Export the roadmap grid to a styled xlsx workbook in memory.

    Args:
        monthly_df: Monthly grid DataFrame from ``build_roadmap()``.
        summary: Summary dict from ``build_roadmap()``.
        hourly_rate: Cost per hour for budget calculation.
        client_name: Optional client name for the header.
        fy_label: Fiscal year label shown in the header.

    Returns:
        BytesIO buffer containing the xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Roadmap"

    # ── Title row ────────────────────────────────────────────────────────
    title_text = f"SEO Roadmap — {fy_label}"
    if client_name:
        title_text = f"{client_name} | {title_text}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(monthly_df.columns))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    # ── Header row (row 3) ───────────────────────────────────────────────
    header_row = 3
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    for col_idx, col_name in enumerate(monthly_df.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ────────────────────────────────────────────────────────
    data_start = header_row + 1
    for row_idx, (_, row_data) in enumerate(monthly_df.iterrows(), start=data_start):
        for col_idx, col_name in enumerate(monthly_df.columns, start=1):
            value = row_data[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Colour-code the Focus column
            if col_name == "Focus" and value in FOCUS_COLORS:
                hex_color = FOCUS_COLORS[value].lstrip("#")
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True, size=10)

            # Bold the TOTAL row
            if row_data["Task"] == "TOTAL":
                cell.font = Font(bold=True, size=10)

            # Right-align numeric month columns and Total Hours
            if col_name.startswith("M") or col_name == "Total Hours":
                cell.alignment = Alignment(horizontal="right")
                # Show 0 as empty for readability
                if isinstance(value, (int, float)) and value == 0:
                    cell.value = ""

    # ── Column widths ────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 32  # Task
    ws.column_dimensions[get_column_letter(2)].width = 14  # Focus
    ws.column_dimensions[get_column_letter(3)].width = 14  # Occurrence
    for c in range(4, len(monthly_df.columns) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 10

    # ── Summary section ──────────────────────────────────────────────────
    summary_start = data_start + len(monthly_df) + 2
    label_font = Font(bold=True, size=10)

    total_hours = summary.get("total_hours", 0)
    avg_hours = summary.get("avg_hours_per_month", 0)
    total_cost = total_hours * hourly_rate

    summary_items = [
        ("Total Hours", total_hours),
        ("Avg Hours / Month", avg_hours),
        ("Hourly Rate", f"${hourly_rate:,.2f}"),
        ("Total Cost", f"${total_cost:,.2f}"),
    ]

    for i, (label, value) in enumerate(summary_items):
        r = summary_start + i
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font = label_font
        ws.cell(row=r, column=2, value=value)

    # ── Write to buffer ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
