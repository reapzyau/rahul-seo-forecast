"""Pattern Multi-Channel Plan template populator (Section 7).

Populates the SEO block of Pattern's standard Multi-Channel Plan Excel template
with monthly forecast values, adds an SEO Forecast Notes sheet with methodology,
and writes formula cells so Revenue/Transactions/ROAS update dynamically if
the analyst adjusts assumptions inside Excel.

Template structure assumed (v1 Multi-Channel Plan):
  - Sheets: "4 CHANNEL FORECAST (15%)", "4 CHANNEL FORECAST (20%)", "4 CHANNEL FORECAST (25%)"
  - SEO block at rows 49-58
  - 12 month columns as 3-col groups (Forecast/Actual/%Var) starting at col C (Jun)
  - Annual totals at col AM

Usage
-----
    from exporters.pattern_template import populate_template
    populate_template(
        template_path=Path("template.xlsx"),
        output_path=Path("client_forecast.xlsx"),
        forecasts_by_tier={"4k": df_4k, "5.5k": df_55k, "7k": df_7k},
        blended_cr=0.018,
        weighted_aov=180.0,
        overlap_months=["Jul-26", "Aug-26", ...],
        methodology_text="...",
    )
"""

from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ── Template constants ──────────────────────────────────────────────────────

# Row indices for the SEO metrics block (1-based, matching Pattern v1 template).
SEO_BLOCK_ROWS = {
    "spend":      49,
    "revenue":    50,
    "roas":       51,
    "transactions": 52,
    "aov":        53,
    "traffic":    54,
    "cvr":        55,
    "mgmt_fee":   56,
    "tech_fee":   57,
    "total":      58,
}

# Each month maps to the column index of its FORECAST sub-column (1-based).
# The group layout is: col N = Forecast, N+1 = Actual, N+2 = %Var.
MONTH_COL_INDICES = {
    "Jun": 3,  "Jul": 6,  "Aug": 9,  "Sep": 12,
    "Oct": 15, "Nov": 18, "Dec": 21, "Jan": 24,
    "Feb": 27, "Mar": 30, "Apr": 33, "May": 36,
}

ANNUAL_TOTAL_COL = 39  # column AM

# Columns whose annual total should be re-written as SUM formulas.
# (Template sometimes has hard-coded zeros here.)
ANNUAL_SUM_ROWS = [
    SEO_BLOCK_ROWS["spend"],
    SEO_BLOCK_ROWS["transactions"],
    SEO_BLOCK_ROWS["traffic"],
    SEO_BLOCK_ROWS["mgmt_fee"],
]

# Standard percentage-tier sheet names in the template.
_STANDARD_TIER_SHEETS = [
    "4 CHANNEL FORECAST (15%)",
    "4 CHANNEL FORECAST (20%)",
    "4 CHANNEL FORECAST (25%)",
]


# ── Helpers ─────────────────────────────────────────────────────────────────


def _retainer_from_tier_key(tier_key: str) -> float:
    """Parse a tier key like '4k', '5.5k', '7k' → monthly retainer as float."""
    clean = tier_key.lower().replace("k", "").strip()
    try:
        return float(clean) * 1000
    except ValueError:
        return 0.0


def _month_short(month_label: str) -> str:
    """Extract 3-char month abbreviation from labels like 'Jul-26' or 'Jul 2026'."""
    match = re.match(r"([A-Za-z]{3})", month_label)
    return match.group(1).capitalize() if match else month_label[:3]


def _add_annual_sum_formulas(ws, rows: list[int]) -> None:
    """Replace values in the annual-total column with SUM formulas over the 12 month cols."""
    forecast_col_letters = [get_column_letter(c) for c in MONTH_COL_INDICES.values()]
    for r in rows:
        terms = "+".join(f"{c}{r}" for c in forecast_col_letters)
        ws.cell(row=r, column=ANNUAL_TOTAL_COL, value=f"=SUM({terms})")


def _add_seo_forecast_notes(
    wb,
    forecasts_by_tier: dict,
    blended_cr: float,
    weighted_aov: float,
    methodology_text: str = "",
) -> None:
    """Add an 'SEO Forecast Notes' sheet with KPIs and methodology text."""
    ws = wb.create_sheet("SEO Forecast Notes")
    row = 1

    def _write(label: str, value=None):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        if value is not None:
            ws.cell(row=row, column=2, value=value)
        row += 1

    _write("SEO Forecast — Methodology Notes")
    _write("")
    _write("Revenue Assumptions")
    _write("  Blended CVR", f"{blended_cr * 100:.2f}%")
    _write("  Weighted AOV", f"${weighted_aov:,.0f}")
    _write("")
    _write("Tier Summary")
    for tier_key, df in forecasts_by_tier.items():
        if "combined_p50" in df.columns:
            annual = df["combined_p50"].sum()
            _write(f"  {tier_key} — Annual combined sessions (P50)", int(annual))
    _write("")
    _write("Methodology")
    for line in (methodology_text or "").splitlines():
        _write(line)


# ── Main entry point ─────────────────────────────────────────────────────────


def populate_template(
    template_path: Path,
    output_path: Path,
    forecasts_by_tier: dict,
    blended_cr: float,
    weighted_aov: float,
    overlap_months: list[str],
    methodology_text: str = "",
    tier_sheet_renames: dict | None = None,
) -> Path:
    """Populate a Pattern Multi-Channel Plan template with SEO forecast values.

    Args:
        template_path: Path to the client's blank Multi-Channel Plan .xlsx.
        output_path: Where to write the populated file.
        forecasts_by_tier: Dict mapping tier key (e.g. "4k") to a DataFrame with
            columns: month_name (str like "Jul-26"), combined_p50 (int sessions).
            Optional: baseline_traffic, positional_p50, new_content.
        blended_cr: Blended conversion rate (decimal, e.g. 0.018 for 1.8%).
        weighted_aov: Weighted average order value in dollars.
        overlap_months: List of month labels (e.g. ["Jul-26", "Aug-26", ...])
            that fall within the forecast horizon. Months NOT in this list are
            left at their template values (usually 0).
        methodology_text: Human-readable methodology to embed in Notes sheet.
        tier_sheet_renames: Optional dict mapping template sheet names to new names.
            Auto-inferred from forecasts_by_tier keys when not provided.

    Returns:
        output_path for chaining.
    """
    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)

    tier_keys = list(forecasts_by_tier.keys())

    # Rename tier sheets
    if tier_sheet_renames:
        rename_map = tier_sheet_renames
    else:
        rename_map = dict(zip(
            _STANDARD_TIER_SHEETS[:len(tier_keys)],
            [f"4 CHANNEL FORECAST ({k})" for k in tier_keys],
            strict=False,
        ))

    for old, new in rename_map.items():
        if old in wb.sheetnames and old != new:
            wb[old].title = new

    # Populate each tier sheet
    overlap_set = set(overlap_months)

    for tier_key, df in forecasts_by_tier.items():
        sheet_name = f"4 CHANNEL FORECAST ({tier_key})"
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        retainer = _retainer_from_tier_key(tier_key)

        # Build lookup: month_label → row
        df_by_month: dict[str, dict] = {}
        for _, row in df.iterrows():
            label = str(row.get("month_name", ""))
            df_by_month[label] = row.to_dict()

        for month_label in overlap_set:
            short = _month_short(month_label)
            col_idx = MONTH_COL_INDICES.get(short)
            if col_idx is None:
                continue
            col_letter = get_column_letter(col_idx)

            row_data = df_by_month.get(month_label)
            if row_data is None:
                continue

            traffic = int(row_data.get("combined_p50", 0))
            r = SEO_BLOCK_ROWS

            ws.cell(row=r["spend"],   column=col_idx, value=0)
            ws.cell(row=r["aov"],     column=col_idx, value=weighted_aov)
            ws.cell(row=r["traffic"], column=col_idx, value=traffic)
            ws.cell(row=r["cvr"],     column=col_idx, value=blended_cr)
            ws.cell(row=r["mgmt_fee"], column=col_idx, value=retainer)
            ws.cell(row=r["tech_fee"], column=col_idx, value=0)

            # Formula-driven metrics (update live if analyst tweaks assumptions)
            ws.cell(
                row=r["revenue"], column=col_idx,
                value=f"={col_letter}{r['aov']}*{col_letter}{r['traffic']}*{col_letter}{r['cvr']}",
            )
            ws.cell(
                row=r["transactions"], column=col_idx,
                value=f"={col_letter}{r['traffic']}*{col_letter}{r['cvr']}",
            )
            ws.cell(
                row=r["roas"], column=col_idx,
                value=f"=IFERROR({col_letter}{r['revenue']}/{col_letter}{r['mgmt_fee']},0)",
            )
            ws.cell(
                row=r["total"], column=col_idx,
                value=f"={col_letter}{r['mgmt_fee']}+{col_letter}{r['tech_fee']}",
            )

        # Annual total formulas
        _add_annual_sum_formulas(ws, ANNUAL_SUM_ROWS)

    # Add SEO Forecast Notes sheet
    _add_seo_forecast_notes(wb, forecasts_by_tier, blended_cr, weighted_aov, methodology_text)

    wb.save(output_path)
    return output_path
