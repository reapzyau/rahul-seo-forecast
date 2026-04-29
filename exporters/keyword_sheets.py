"""Keyword detail sheets for the Pattern Multi-Channel Plan export (Section 9).

Adds two types of sheets to the workbook:
  1. Top Keywords (Xk) — top 50 keywords by P50 uplift at M12 per tier
  2. Keyword Movement Summary — pivot table: keyword count and volume by
     difficulty tier × position bucket across the full quick-win pool.

Usage
-----
    from exporters.keyword_sheets import add_top_keywords_sheet, add_keyword_movement_summary
    add_top_keywords_sheet(wb, "Top Keywords (4k)", kw_df_top50, tier_label="4k")
    add_keyword_movement_summary(wb, full_quick_win_df)
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Constants ────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="2B5EA7")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="E8F0FE")
_TOTAL_FONT = Font(bold=True)

POSITION_BUCKETS = ["4-5", "6-10", "11-15", "16-20", "21-30"]
DIFFICULTY_ORDER = ["Easy", "Moderate", "Hard", "Very Hard", "Extreme"]


def _bucket_position(pos) -> str:
    """Map a position value to the appropriate bucket label."""
    try:
        p = int(pos)
    except (ValueError, TypeError):
        return "Other"
    if p <= 5:
        return "4-5"
    if p <= 10:
        return "6-10"
    if p <= 15:
        return "11-15"
    if p <= 20:
        return "16-20"
    return "21-30"


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _apply_color_scale(ws, col_letter: str, start_row: int, end_row: int) -> None:
    """Green-Yellow-Red colour scale on a column (high = green)."""
    rule = ColorScaleRule(
        start_type="min", start_color="F8696B",   # red (low)
        mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow (mid)
        end_type="max", end_color="63BE7B",        # green (high)
    )
    ws.conditional_formatting.add(
        f"{col_letter}{start_row}:{col_letter}{end_row}", rule
    )


# ── Top Keywords sheet ───────────────────────────────────────────────────────


def add_top_keywords_sheet(
    wb: Workbook,
    sheet_name: str,
    kw_df: pd.DataFrame,
    tier_label: str,
    top_n: int = 50,
) -> None:
    """Add a top-N keywords sheet sorted by P50 uplift descending.

    Required columns in kw_df: keyword, volume, kd, tier, position,
        target_position, uplift_p10, uplift_p50, uplift_p90.
    Optional: uplift (alias for uplift_p50).

    Args:
        wb: Workbook to add the sheet to.
        sheet_name: Sheet title.
        kw_df: Full keyword DataFrame from run_positional_forecast_mc.
        tier_label: Human-readable tier label shown in the sheet header.
        top_n: Number of keywords to include (default 50).
    """
    ws = wb.create_sheet(sheet_name)

    # Sort by P50 uplift descending and take top_n
    uplift_col = "uplift_p50" if "uplift_p50" in kw_df.columns else "uplift"
    df = (
        kw_df.sort_values(uplift_col, ascending=False)
        .head(top_n)
        .reset_index(drop=True)
    )

    headers = [
        "Rank", "Keyword", "Volume", "KD", "Tier",
        "Current Pos", "Target Pos", "Position Gain",
        "P10 Uplift", "P50 Uplift", "P90 Uplift",
    ]

    # Title row
    ws.cell(row=1, column=1, value=f"Top Keywords — {tier_label} Tier")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13)

    _write_header_row(ws, 2, headers)

    data_start = 3
    for idx, row in df.iterrows():
        r = data_start + idx
        cur_col = get_column_letter(6)
        tar_col = get_column_letter(7)
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=str(row.get("keyword", "")))
        ws.cell(row=r, column=3, value=int(row.get("volume", 0)))
        ws.cell(row=r, column=4, value=int(row.get("kd", 0)))
        ws.cell(row=r, column=5, value=str(row.get("tier", "")))
        ws.cell(row=r, column=6, value=int(row.get("position", 0)))
        ws.cell(row=r, column=7, value=int(row.get("target_position", 0)))
        # Position gain as formula: current_pos - target_pos
        ws.cell(row=r, column=8, value=f"={cur_col}{r}-{tar_col}{r}")
        ws.cell(row=r, column=9,  value=round(float(row.get("uplift_p10", 0)), 1))
        ws.cell(row=r, column=10, value=round(float(row.get(uplift_col, 0)), 1))
        ws.cell(row=r, column=11, value=round(float(row.get("uplift_p90", 0)), 1))

    data_end = data_start + len(df) - 1

    # TOTAL row
    total_row = data_end + 1
    ws.cell(row=total_row, column=1, value="TOTAL")
    ws.cell(row=total_row, column=1).font = _TOTAL_FONT
    ws.cell(row=total_row, column=1).fill = _TOTAL_FILL
    for col_idx, _agg_col in [(3, "volume"), (9, None), (10, None), (11, None)]:
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=total_row, column=col_idx,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
        ws.cell(row=total_row, column=col_idx).fill = _TOTAL_FILL
        ws.cell(row=total_row, column=col_idx).font = _TOTAL_FONT

    # Colour scale on P50 uplift column (col 10)
    if data_end >= data_start:
        _apply_color_scale(ws, "J", data_start, data_end)

    # Column widths
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["E"].width = 12
    for c in "ACDFGHIJ":
        ws.column_dimensions[c].width = 14


# ── Keyword Movement Summary ─────────────────────────────────────────────────


def add_keyword_movement_summary(
    wb: Workbook,
    kw_df: pd.DataFrame,
    sheet_name: str = "Keyword Movement Summary",
) -> None:
    """Add a keyword movement summary sheet with two pivot tables.

    Pivot 1: Keyword count by difficulty tier × position bucket
    Pivot 2: Total search volume by difficulty tier × position bucket

    Required columns in kw_df: tier, position, volume.

    Args:
        wb: Workbook to add the sheet to.
        kw_df: Full keyword DataFrame from the positional pool.
        sheet_name: Sheet title.
    """
    ws = wb.create_sheet(sheet_name)

    df = kw_df.copy()
    if "tier" not in df.columns or "position" not in df.columns:
        ws.cell(row=1, column=1, value="No data available — tier/position columns missing.")
        return

    df["position_bucket"] = df["position"].apply(_bucket_position)

    def _write_pivot(
        title: str,
        pivot: pd.DataFrame,
        start_row: int,
        value_label: str,
    ) -> int:
        """Write a pivot table to the sheet. Returns next available row."""
        ws.cell(row=start_row, column=1, value=title)
        ws.cell(row=start_row, column=1).font = Font(bold=True, size=12)
        start_row += 1

        # Headers: blank, then position bucket labels, then Total
        buckets = [b for b in POSITION_BUCKETS if b in pivot.columns]
        header = ["Difficulty Tier"] + buckets + ["Total"]
        _write_header_row(ws, start_row, header)
        start_row += 1

        data_start = start_row
        tiers = [t for t in DIFFICULTY_ORDER if t in pivot.index]
        for tier in tiers:
            ws.cell(row=start_row, column=1, value=tier)
            row_total = 0
            for col_idx, bucket in enumerate(buckets, 2):
                val = int(pivot.loc[tier, bucket]) if bucket in pivot.columns else 0
                ws.cell(row=start_row, column=col_idx, value=val)
                row_total += val
            ws.cell(row=start_row, column=len(header), value=row_total)
            start_row += 1

        data_end = start_row - 1

        # Column totals
        ws.cell(row=start_row, column=1, value="Total")
        ws.cell(row=start_row, column=1).font = _TOTAL_FONT
        ws.cell(row=start_row, column=1).fill = _TOTAL_FILL
        for col_idx, _bucket in enumerate(buckets, 2):
            col_letter = get_column_letter(col_idx)
            ws.cell(
                row=start_row, column=col_idx,
                value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
            )
            ws.cell(row=start_row, column=col_idx).fill = _TOTAL_FILL
            ws.cell(row=start_row, column=col_idx).font = _TOTAL_FONT
        grand_col_letter = get_column_letter(len(header))
        ws.cell(
            row=start_row, column=len(header),
            value=f"=SUM({grand_col_letter}{data_start}:{grand_col_letter}{data_end})",
        )
        ws.cell(row=start_row, column=len(header)).fill = _TOTAL_FILL
        ws.cell(row=start_row, column=len(header)).font = _TOTAL_FONT

        # Colour scale on bucket columns
        for col_idx in range(2, len(header)):
            col_letter = get_column_letter(col_idx)
            if data_end >= data_start:
                _apply_color_scale(ws, col_letter, data_start, data_end)

        return start_row + 2

    # Pivot 1: keyword count
    pivot_count = (
        df.groupby(["tier", "position_bucket"])
        .size()
        .unstack(fill_value=0)
    )
    next_row = _write_pivot("Keyword Count by Tier × Position Bucket", pivot_count, 1, "Count")

    # Pivot 2: total volume
    if "volume" in df.columns:
        pivot_volume = (
            df.groupby(["tier", "position_bucket"])["volume"]
            .sum()
            .unstack(fill_value=0)
        )
        _write_pivot("Total Search Volume by Tier × Position Bucket", pivot_volume, next_row, "Volume")

    # Column widths
    ws.column_dimensions["A"].width = 20
    for i in range(2, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14
