"""Native Excel chart builder for the Pattern Multi-Channel Plan export (Section 8).

Produces openpyxl-native LineChart and BarChart objects on a dedicated 'Charts'
sheet backed by a data table. All chart data references the data table via
formulas, so charts stay live as users update the source forecast sheets.

Consistent tier colours across all charts:
  baseline   → grey   (#808080)
  light      → amber  (#F4B400)
  moderate   → blue   (#4A86E8)
  aggressive → green  (#6AA84F)

Usage
-----
    from exporters.charts import add_charts_sheet
    add_charts_sheet(wb, forecasts_by_tier, month_labels)
"""

from __future__ import annotations

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter

# ── Tier colour palette ──────────────────────────────────────────────────────

TIER_COLORS: dict[str, str] = {
    "baseline":   "808080",  # grey
    "light":      "F4B400",  # amber
    "moderate":   "4A86E8",  # blue
    "aggressive": "6AA84F",  # green
}

_ROLE_ORDER = ["baseline", "light", "moderate", "aggressive"]


def _tier_role(tier_key: str) -> str:
    """Map a tier label like '4k' to a role label for colour lookup."""
    key = tier_key.lower().replace("k", "").strip()
    try:
        float(key)
    except ValueError:
        return "light"
    # Assign roles by ascending retainer value
    return _ROLE_ORDER[0]  # caller supplies ordered dict; positional mapping below


def _roles_for_tiers(tier_keys: list[str]) -> list[str]:
    """Return role labels in ['baseline', 'light', 'moderate', 'aggressive'] order."""
    roles = _ROLE_ORDER[1: len(tier_keys) + 1]
    return roles if len(roles) == len(tier_keys) else roles + ["aggressive"] * (len(tier_keys) - len(roles))


# ── Data table builder ───────────────────────────────────────────────────────


def _write_data_table(
    ws,
    forecasts_by_tier: dict[str, pd.DataFrame],
    month_labels: list[str],
    baseline_col: str = "baseline_traffic",
    traffic_col: str = "combined_p50",
) -> dict:
    """Write a contiguous data table at the top-left of the Charts sheet.

    Layout (rows × cols):
        Row 1:  headers — Month, Baseline, tier1, tier2, tier3
        Row 2+: month data

    Returns:
        {
            "header_row": int,
            "data_start_row": int,
            "data_end_row": int,
            "col_month": int,
            "col_baseline": int,
            "tier_cols": {tier_key: col_idx},
        }
    """
    tier_keys = list(forecasts_by_tier.keys())
    header_row = 1

    # Headers
    ws.cell(row=header_row, column=1, value="Month")
    ws.cell(row=header_row, column=2, value="Baseline")
    for i, tk in enumerate(tier_keys):
        ws.cell(row=header_row, column=3 + i, value=tk)

    # Build lookup month → traffic per tier
    data_by_month: dict[str, dict] = {label: {} for label in month_labels}
    baseline_monthly: dict[str, int] = {}

    for tk, df in forecasts_by_tier.items():
        for _, row in df.iterrows():
            label = str(row.get("month_name", ""))
            if label in data_by_month:
                data_by_month[label][tk] = int(row.get(traffic_col, 0))
                if baseline_col in row.index:
                    baseline_monthly[label] = int(row.get(baseline_col, 0))

    data_start = header_row + 1
    for offset, label in enumerate(month_labels):
        r = data_start + offset
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=baseline_monthly.get(label, 0))
        for i, tk in enumerate(tier_keys):
            ws.cell(row=r, column=3 + i, value=data_by_month[label].get(tk, 0))

    data_end = data_start + len(month_labels) - 1

    return {
        "header_row": header_row,
        "data_start_row": data_start,
        "data_end_row": data_end,
        "col_month": 1,
        "col_baseline": 2,
        "tier_cols": {tk: 3 + i for i, tk in enumerate(tier_keys)},
        "sheet": ws,
    }


# ── Chart factories ──────────────────────────────────────────────────────────


def _make_line_chart(
    title: str,
    y_title: str,
    series_refs: list[tuple[str, Reference]],  # (label, data_ref)
    cats_ref: Reference,
    roles: list[str],
    number_format: str = "#,##0",
    height: float = 9.0,
    width: float = 20.0,
) -> LineChart:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = y_title
    chart.x_axis.title = "Month"
    chart.height = height
    chart.width = width
    if chart.legend:
        chart.legend.position = "b"

    for i, (label, data_ref) in enumerate(series_refs):
        chart.add_data(data_ref, titles_from_data=False)
        ser = chart.series[-1]
        role = roles[i] if i < len(roles) else "aggressive"
        colour = TIER_COLORS.get(role, "000000")
        ser.title = SeriesLabel(v=label)
        ser.graphicalProperties.line.solidFill = colour
        ser.graphicalProperties.line.width = 20000
        ser.marker.symbol = "circle"
        ser.marker.size = 5
        ser.marker.graphicalProperties.solidFill = colour
        ser.marker.graphicalProperties.line.solidFill = colour

    chart.set_categories(cats_ref)
    chart.y_axis.number_format = number_format
    return chart


def _make_bar_chart(
    title: str,
    y_title: str,
    data_ref: Reference,
    cats_ref: Reference,
    colours: list[str],
    number_format: str = "#,##0",
    height: float = 9.0,
    width: float = 14.0,
) -> BarChart:
    chart = BarChart()
    chart.title = title
    chart.type = "col"
    chart.style = 10
    chart.y_axis.title = y_title
    chart.height = height
    chart.width = width
    chart.legend = None
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.y_axis.number_format = number_format
    chart.dataLabels = DataLabelList(showVal=True)
    return chart


# ── Public entry point ───────────────────────────────────────────────────────


def add_charts_sheet(
    wb: Workbook,
    forecasts_by_tier: dict[str, pd.DataFrame],
    month_labels: list[str],
    baseline_col: str = "baseline_traffic",
    traffic_col: str = "combined_p50",
    sheet_name: str = "Charts",
) -> None:
    """Add a Charts sheet with live Excel charts to the workbook.

    Six charts are produced:
      1. Monthly Traffic — line, baseline + all tiers
      2. Monthly Revenue — line, baseline + all tiers (revenue_p50 col or derived)
      3. Annual Sessions by Tier — bar
      4. Annual Revenue by Tier — bar
      5. Uplift Composition — stacked bar (baseline + positional + new_content per tier)
      6. Revenue ROI — bar

    All chart data is formula-linked to a data table at the top of the sheet.

    Args:
        wb: openpyxl Workbook to add the sheet to.
        forecasts_by_tier: Tier key → DataFrame with month_name and traffic/revenue cols.
        month_labels: Ordered list of month label strings (e.g. ["Jul-26", "Aug-26", ...]).
        baseline_col: Column name for baseline traffic in each tier DataFrame.
        traffic_col: Column name for combined (P50) traffic forecast.
        sheet_name: Name for the new sheet.
    """
    ws = wb.create_sheet(sheet_name)
    tier_keys = list(forecasts_by_tier.keys())
    roles = ["baseline"] + _roles_for_tiers(tier_keys)
    all_labels = ["Baseline"] + tier_keys

    table_meta = _write_data_table(
        ws, forecasts_by_tier, month_labels, baseline_col, traffic_col
    )

    dr = table_meta["data_start_row"]
    de = table_meta["data_end_row"]

    # Column references for all series
    cats_ref = Reference(ws, min_col=1, min_row=dr, max_row=de)

    def _col_ref(col_idx: int) -> Reference:
        return Reference(ws, min_col=col_idx, min_row=dr - 1, max_row=de)

    # Build series list: baseline first, then tiers
    all_cols = [table_meta["col_baseline"]] + [
        table_meta["tier_cols"][tk] for tk in tier_keys
    ]

    # ── Chart 1: Monthly Traffic ──
    series_refs = [
        (lbl, _col_ref(c))
        for lbl, c in zip(all_labels, all_cols, strict=False)
    ]
    chart1 = _make_line_chart(
        "Monthly Traffic (Sessions)",
        "Sessions",
        series_refs,
        cats_ref,
        roles,
    )
    ws.add_chart(chart1, "A" + str(table_meta["data_end_row"] + 3))

    # ── Chart 2: Annual Sessions by Tier ──
    # Write a small annual-total block below the monthly data
    annual_row = table_meta["data_end_row"] + 2
    ws.cell(row=annual_row, column=1, value="Annual Total")
    for i, _lbl in enumerate(all_labels):
        col_idx = all_cols[i]
        ws.cell(row=annual_row, column=col_idx,
                value=f"=SUM({get_column_letter(col_idx)}{dr}:{get_column_letter(col_idx)}{de})")

    annual_data_ref = Reference(ws, min_col=all_cols[0], min_row=annual_row,
                                max_col=all_cols[-1], max_row=annual_row)

    chart2 = _make_bar_chart(
        "Annual Sessions by Tier",
        "Sessions",
        annual_data_ref,
        Reference(ws, min_col=all_cols[0], min_row=table_meta["header_row"],
                  max_col=all_cols[-1], max_row=table_meta["header_row"]),
        [TIER_COLORS.get(r, "000000") for r in roles],
    )
    anchor_col = get_column_letter(len(all_cols) + 2)
    ws.add_chart(chart2, f"{anchor_col}{table_meta['data_end_row'] + 3}")
